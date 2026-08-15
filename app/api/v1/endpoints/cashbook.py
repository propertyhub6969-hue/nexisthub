import uuid
from datetime import date
from decimal import Decimal
from typing import Optional

import io
from fastapi import APIRouter, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.api.deps import get_current_context, AuthContext
from app.core.audit import record_audit
from app.core.cashbook import sync_expense_cashbook, sync_notary_fee_cashbook
from app.core.notify import notify_roles
from app.models.cashbook import AccountCategory, CashBookEntry, CashDirection, CashAccount, CashTransfer, CashReconciliation
from app.models.expense import Expense
from app.models.payment import Payment
from app.models.marketing import Client
from app.models.tax import NotaryFee, Notary
from app.models.notification import NotificationKind
from app.models.property import Project, Unit, UnitUtility
from app.models.user import UserRole
from app.schemas.marketing import Paginated
from app.schemas.cashbook import (
    CategoryResponse, CashBookEntryResponse, CashBookSummary, CashBookCategoryTotal, CashBookMonth,
    PendingExpenseRow, PendingExpenseList, MarkExpensePaidRequest,
    CashAccountCreate, CashAccountUpdate, CashAccountResponse, CashAccountsSummary,
    CashTransferCreate, CashTransferResponse, EntryAccountUpdate, ClearedUpdate,
    ReconMovement, ReconcileView, ReconcileSaveRequest, ReconciliationRow,
    MutationRow, MutationImportResult,
)
from fastapi import HTTPException, status as httpstatus

router = APIRouter()


def _paginate(items, total, page, size):
    import math
    return {"items": items, "total": total, "page": page, "size": size, "pages": math.ceil(total / size) if size else 0}


@router.get("/categories", response_model=list[CategoryResponse])
async def list_categories(ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    """Peta akun (Fase B1 — daftar pendek). Kategori bawaan sistem (code terisi) di-seed otomatis per tenant."""
    r = await db.execute(
        select(AccountCategory).where(AccountCategory.tenant_id == ctx.tenant_id, AccountCategory.is_deleted == False)  # noqa: E712
        .order_by(AccountCategory.direction, AccountCategory.name)
    )
    return r.scalars().all()


@router.get("/entries", response_model=Paginated[CashBookEntryResponse])
async def list_entries(
    direction: Optional[CashDirection] = Query(None),
    category_id: Optional[uuid.UUID] = Query(None),
    account_id: Optional[uuid.UUID] = Query(None),
    unassigned: bool = Query(False, description="hanya entri yang belum berrekening"),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=500),
    ctx: AuthContext = Depends(get_current_context),
    db: AsyncSession = Depends(get_db),
):
    """Daftar baris Buku Kas (otomatis dari pembayaran disetujui & biaya dibayar), terbaru dulu."""
    conds = [CashBookEntry.tenant_id == ctx.tenant_id]
    if direction:
        conds.append(CashBookEntry.direction == direction)
    if category_id:
        conds.append(CashBookEntry.category_id == category_id)
    if account_id:
        conds.append(CashBookEntry.account_id == account_id)
    if unassigned:
        conds.append(CashBookEntry.account_id.is_(None))
    if date_from:
        conds.append(CashBookEntry.date >= date_from)
    if date_to:
        conds.append(CashBookEntry.date <= date_to)

    total = await db.scalar(select(func.count()).select_from(CashBookEntry).where(*conds))
    rows = (await db.execute(
        select(CashBookEntry, Client.full_name, Project.name, CashAccount.name)
        .select_from(CashBookEntry)
        .options(selectinload(CashBookEntry.category))
        .outerjoin(Client, Client.id == CashBookEntry.client_id)
        .outerjoin(Project, Project.id == CashBookEntry.project_id)
        .outerjoin(CashAccount, CashAccount.id == CashBookEntry.account_id)
        .where(*conds)
        .order_by(CashBookEntry.date.desc(), CashBookEntry.created_at.desc())
        .offset((page - 1) * size).limit(size)
    )).all()

    items = []
    for entry, client_name, project_name, account_name in rows:
        item = CashBookEntryResponse.model_validate(entry).model_copy(update={
            "client_name": client_name, "project_name": project_name, "account_name": account_name,
        })
        items.append(item)
    return _paginate(items, total or 0, page, size)


@router.get("/summary", response_model=CashBookSummary)
async def cashbook_summary(
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    ctx: AuthContext = Depends(get_current_context),
    db: AsyncSession = Depends(get_db),
):
    """Rekap Buku Kas: total masuk/keluar, per kategori, tren bulanan (maks 12 bulan terakhir ada transaksi)."""
    t = ctx.tenant_id
    conds = [CashBookEntry.tenant_id == t]
    if date_from:
        conds.append(CashBookEntry.date >= date_from)
    if date_to:
        conds.append(CashBookEntry.date <= date_to)

    total_in = Decimal(await db.scalar(
        select(func.coalesce(func.sum(CashBookEntry.amount), 0)).where(*conds, CashBookEntry.direction == CashDirection.IN)
    ))
    total_out = Decimal(await db.scalar(
        select(func.coalesce(func.sum(CashBookEntry.amount), 0)).where(*conds, CashBookEntry.direction == CashDirection.OUT)
    ))

    cat_rows = (await db.execute(
        select(CashBookEntry.category_id, AccountCategory.name, CashBookEntry.direction,
               func.coalesce(func.sum(CashBookEntry.amount), 0))
        .select_from(CashBookEntry)
        .outerjoin(AccountCategory, AccountCategory.id == CashBookEntry.category_id)
        .where(*conds)
        .group_by(CashBookEntry.category_id, AccountCategory.name, CashBookEntry.direction)
    )).all()
    by_category = [
        CashBookCategoryTotal(category_id=cid, category_name=name or "Belum dikategorikan", direction=direction, total=Decimal(total))
        for cid, name, direction, total in cat_rows
    ]
    by_category.sort(key=lambda r: (r.direction.value, -r.total))

    ym = func.to_char(CashBookEntry.date, "YYYY-MM")
    month_rows = (await db.execute(
        select(
            ym.label("ym"),
            func.coalesce(func.sum(CashBookEntry.amount).filter(CashBookEntry.direction == CashDirection.IN), 0),
            func.coalesce(func.sum(CashBookEntry.amount).filter(CashBookEntry.direction == CashDirection.OUT), 0),
        )
        .select_from(CashBookEntry)
        .where(*conds)
        .group_by(ym).order_by(ym)
    )).all()
    months = [CashBookMonth(month=m, total_in=Decimal(i), total_out=Decimal(o)) for m, i, o in month_rows][-12:]

    return CashBookSummary(
        total_in=total_in, total_out=total_out, saldo=total_in - total_out,
        by_category=by_category, months=months,
    )


# ═══════════════ BIAYA MENUNGGU BAYAR (pengeluaran diajukan) ═══════════════
# Pengeluaran lahir berstatus DIAJUKAN (is_paid=False) dan BELUM masuk Buku Kas —
# lihat sync_expense_cashbook. Keuangan yang menandainya lunas beserta tanggal bayar
# sebenarnya (di lapangan tanggal bayar sering beda dari tanggal pasang/opname).
# Cakupan: SEMUA jalur pengeluaran — utilitas, opname borongan, biaya manual, dan
# biaya notaris. Ini VALIDASI, bukan persetujuan: tak ada jalur penolakan; setiap
# pengeluaran dianggap sah, langkah ini hanya mencatatkan kapan uangnya benar keluar.
# Kunci baris memakai "<sumber>:<id>" karena datanya lintas dua tabel.

_EXP_LABEL = {
    "material": "Material", "upah": "Upah", "kontraktor": "Kontraktor",
    "kelistrikan": "Kelistrikan", "air_pdam": "Air / PDAM",
    "operasional": "Operasional", "perizinan": "Perizinan", "lain": "Lain-lain",
}


def _unit_label(u) -> Optional[str]:
    if u is None:
        return None
    return f"{u.block}-{u.unit_number}" if u.block else u.unit_number


@router.get("/pending-expenses", response_model=PendingExpenseList)
async def list_pending_expenses(
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    """Semua pengeluaran yang sudah diajukan tapi belum ditandai lunas keuangan.
    Mencakup empat jalur: utilitas, opname borongan, biaya manual, dan biaya notaris."""
    today = date.today()
    out: list[PendingExpenseRow] = []
    total = Decimal(0)

    # ── 1-3. dari tabel expenses (utilitas / opname / biaya manual) ──
    rows = (await db.execute(
        select(Expense, UnitUtility, Unit, Project)
        .outerjoin(UnitUtility, UnitUtility.id == Expense.utility_id)
        .outerjoin(Unit, Unit.id == Expense.unit_id)
        .outerjoin(Project, Project.id == Expense.project_id)
        .where(Expense.tenant_id == ctx.tenant_id, Expense.is_paid == False,  # noqa: E712
               Expense.is_deleted == False)                                    # noqa: E712
    )).all()
    for exp, util, unit, proj in rows:
        total += Decimal(exp.amount or 0)
        cat = exp.category.value if exp.category else "lain"
        ref_date = (util.applied_date if util else None) or exp.expense_date
        out.append(PendingExpenseRow(
            ref=f"expense:{exp.id}", id=exp.id, description=exp.description or "-",
            category=cat, category_label=_EXP_LABEL.get(cat, cat),
            amount=Decimal(exp.amount or 0), expense_date=exp.expense_date,
            project_name=proj.name if proj else None, unit_label=_unit_label(unit),
            source="utilitas" if util else ("opname" if exp.contract_id else "biaya"),
            utility_kind=util.kind.value if util else None,
            utility_status=util.status.value if util else None,
            applied_date=util.applied_date if util else None,
            installed_date=util.installed_date if util else None,
            days_waiting=(today - ref_date).days if ref_date else None,
        ))

    # ── 4. dari tabel notary_fees ──
    fees = (await db.execute(
        select(NotaryFee, Client, Notary)
        .outerjoin(Client, Client.id == NotaryFee.client_id)
        .outerjoin(Notary, Notary.id == NotaryFee.notary_id)
        .where(NotaryFee.tenant_id == ctx.tenant_id, NotaryFee.is_paid == False,  # noqa: E712
               NotaryFee.is_deleted == False)                                      # noqa: E712
    )).all()
    for fee, client, notary in fees:
        total += Decimal(fee.amount or 0)
        out.append(PendingExpenseRow(
            ref=f"notary:{fee.id}", id=fee.id, description=fee.description or "-",
            category="biaya_notaris", category_label="Biaya Notaris/Legal",
            amount=Decimal(fee.amount or 0), expense_date=fee.fee_date,
            source="notaris",
            client_name=client.full_name if client else None,
            notary_name=notary.name if notary else None,
            days_waiting=(today - fee.fee_date).days if fee.fee_date else None,
        ))

    out.sort(key=lambda r: (r.expense_date or date.min), reverse=True)
    return PendingExpenseList(rows=out, total_amount=total)


@router.get("/pending-expenses/count")
async def pending_expenses_count(
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    """Jumlah tagihan menunggu bayar (semua jalur) — untuk badge sidebar."""
    n_exp = (await db.execute(
        select(func.count()).select_from(Expense).where(
            Expense.tenant_id == ctx.tenant_id, Expense.is_paid == False,  # noqa: E712
            Expense.is_deleted == False)                                    # noqa: E712
    )).scalar() or 0
    n_fee = (await db.execute(
        select(func.count()).select_from(NotaryFee).where(
            NotaryFee.tenant_id == ctx.tenant_id, NotaryFee.is_paid == False,  # noqa: E712
            NotaryFee.is_deleted == False)                                      # noqa: E712
    )).scalar() or 0
    return {"count": int(n_exp) + int(n_fee)}


@router.post("/pending-expenses/mark-paid")
async def mark_expenses_paid(
    payload: MarkExpensePaidRequest,
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    """Tandai sekelompok pengeluaran LUNAS dengan tanggal bayar sebenarnya → masuk Buku Kas.
    Ini VALIDASI keuangan, bukan persetujuan: tak ada jalur penolakan — semua pengeluaran
    dianggap sah, langkah ini hanya mencatatkan kapan uangnya benar-benar keluar."""
    if not payload.refs:
        return {"marked": 0}
    pd = payload.paid_date or date.today()

    exp_ids, fee_ids = [], []
    for ref in payload.refs:
        kind, _, rid = ref.partition(":")
        try:
            rid = uuid.UUID(rid)
        except ValueError:
            continue
        (exp_ids if kind == "expense" else fee_ids if kind == "notary" else []).append(rid)

    marked = 0
    if exp_ids:
        rows = (await db.execute(select(Expense).where(
            Expense.id.in_(exp_ids), Expense.tenant_id == ctx.tenant_id,
            Expense.is_paid == False, Expense.is_deleted == False,  # noqa: E712
        ))).scalars().all()
        for e in rows:
            e.is_paid = True
            e.paid_at = pd
            if payload.account_id is not None:
                e.cash_account_id = payload.account_id
        await db.flush()
        for e in rows:
            await sync_expense_cashbook(db, ctx.tenant_id, e)   # baru di sini uang tercatat keluar
        marked += len(rows)

    if fee_ids:
        fees = (await db.execute(select(NotaryFee).where(
            NotaryFee.id.in_(fee_ids), NotaryFee.tenant_id == ctx.tenant_id,
            NotaryFee.is_paid == False, NotaryFee.is_deleted == False,  # noqa: E712
        ))).scalars().all()
        for f in fees:
            f.is_paid = True
            f.paid_at = pd
            if payload.account_id is not None:
                f.cash_account_id = payload.account_id
        await db.flush()
        for f in fees:
            await sync_notary_fee_cashbook(db, ctx.tenant_id, f)
        marked += len(fees)

    if marked:
        await record_audit(db, ctx.tenant_id, ctx.user_id, "PAY", "expenses", None,
                           new_data={"count": marked, "paid_date": str(pd), "refs": payload.refs})
        # beri tahu pengaju bahwa tagihannya sudah dibayar
        await notify_roles(
            db, ctx.tenant_id, (UserRole.PRODUKSI, UserRole.MARKETING), NotificationKind.EXPENSE_PAID,
            title="Biaya sudah dibayar",
            body=f"{marked} tagihan ditandai lunas ({pd.strftime('%d/%m/%Y')})",
            link="/finance/biaya-menunggu-bayar", actor_id=ctx.user_id,
        )
    return {"marked": marked, "paid_date": str(pd)}


# ═══════════════════════ REKENING KAS/BANK (multi-rekening) ═══════════════════════
async def _compute_balances(db, tenant_id):
    """→ (dict account_id→saldo, unassigned_balance). saldo = saldo_awal + masuk − keluar ± transfer."""
    accts = (await db.execute(
        select(CashAccount).where(CashAccount.tenant_id == tenant_id, CashAccount.is_deleted == False)  # noqa: E712
        .order_by(CashAccount.sort_order, CashAccount.name))).scalars().all()
    bal = {a.id: Decimal(a.opening_balance or 0) for a in accts}
    # entri kas per rekening
    erows = (await db.execute(
        select(CashBookEntry.account_id, CashBookEntry.direction, func.coalesce(func.sum(CashBookEntry.amount), 0))
        .where(CashBookEntry.tenant_id == tenant_id).group_by(CashBookEntry.account_id, CashBookEntry.direction))).all()
    unassigned = Decimal(0)
    for acc_id, direction, total in erows:
        total = Decimal(total or 0)
        signed = total if direction == CashDirection.IN else -total
        if acc_id is None:
            unassigned += signed
        elif acc_id in bal:
            bal[acc_id] += signed
    # transfer
    for col, sign in ((CashTransfer.to_account_id, 1), (CashTransfer.from_account_id, -1)):
        trows = (await db.execute(
            select(col, func.coalesce(func.sum(CashTransfer.amount), 0))
            .where(CashTransfer.tenant_id == tenant_id, CashTransfer.is_deleted == False)  # noqa: E712
            .group_by(col))).all()
        for acc_id, total in trows:
            if acc_id in bal:
                bal[acc_id] += Decimal(total or 0) * sign
    return accts, bal, unassigned


@router.get("/accounts", response_model=CashAccountsSummary)
async def list_accounts(ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    accts, bal, unassigned = await _compute_balances(db, ctx.tenant_id)
    out = []
    for a in accts:
        r = CashAccountResponse.model_validate(a).model_copy(update={"balance": bal.get(a.id, Decimal(0))})
        out.append(r)
    total = sum((bal.get(a.id, Decimal(0)) for a in accts), Decimal(0))
    return CashAccountsSummary(accounts=out, total_balance=total, unassigned_balance=unassigned)


async def _get_account(db, tenant_id, account_id) -> CashAccount:
    a = (await db.execute(select(CashAccount).where(
        CashAccount.id == account_id, CashAccount.tenant_id == tenant_id,
        CashAccount.is_deleted == False))).scalar_one_or_none()  # noqa: E712
    if a is None:
        raise HTTPException(status_code=httpstatus.HTTP_404_NOT_FOUND, detail="Rekening tidak ditemukan")
    return a


@router.post("/accounts", response_model=CashAccountResponse, status_code=httpstatus.HTTP_201_CREATED)
async def create_account(payload: CashAccountCreate, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    data = payload.model_dump()
    if data.get("is_default"):
        await db.execute(CashAccount.__table__.update().where(CashAccount.tenant_id == ctx.tenant_id).values(is_default=False))
    a = CashAccount(tenant_id=ctx.tenant_id, **data)
    db.add(a); await db.flush()
    await record_audit(db, ctx.tenant_id, ctx.user_id, "CREATE", "cash_accounts", a.id, new_data={"name": a.name})
    await db.commit(); await db.refresh(a)
    return CashAccountResponse.model_validate(a).model_copy(update={"balance": Decimal(a.opening_balance or 0)})


@router.patch("/accounts/{account_id}", response_model=CashAccountResponse)
async def update_account(account_id: uuid.UUID, payload: CashAccountUpdate, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    a = await _get_account(db, ctx.tenant_id, account_id)
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(a, k, v)
    await db.flush(); await db.commit()
    _, bal, _ = await _compute_balances(db, ctx.tenant_id)
    return CashAccountResponse.model_validate(a).model_copy(update={"balance": bal.get(a.id, Decimal(0))})


@router.post("/accounts/{account_id}/set-default", response_model=CashAccountResponse)
async def set_default_account(account_id: uuid.UUID, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    a = await _get_account(db, ctx.tenant_id, account_id)
    await db.execute(CashAccount.__table__.update().where(CashAccount.tenant_id == ctx.tenant_id).values(is_default=False))
    a.is_default = True
    await db.flush(); await db.commit()
    return CashAccountResponse.model_validate(a).model_copy(update={"balance": Decimal(0)})


@router.delete("/accounts/{account_id}", status_code=httpstatus.HTTP_204_NO_CONTENT)
async def delete_account(account_id: uuid.UUID, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    a = await _get_account(db, ctx.tenant_id, account_id)
    from datetime import datetime, timezone
    a.is_deleted = True; a.deleted_at = datetime.now(timezone.utc); a.is_active = False; a.is_default = False
    await db.commit()


# ── Transfer antar rekening ──
@router.get("/transfers", response_model=list[CashTransferResponse])
async def list_transfers(ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    Frm = CashAccount.__table__.alias("frm"); To = CashAccount.__table__.alias("dst")
    rows = (await db.execute(
        select(CashTransfer, Frm.c.name, To.c.name)
        .outerjoin(Frm, Frm.c.id == CashTransfer.from_account_id)
        .outerjoin(To, To.c.id == CashTransfer.to_account_id)
        .where(CashTransfer.tenant_id == ctx.tenant_id, CashTransfer.is_deleted == False)  # noqa: E712
        .order_by(CashTransfer.date.desc(), CashTransfer.created_at.desc()).limit(200))).all()
    return [CashTransferResponse.model_validate(t).model_copy(update={"from_account_name": fn, "to_account_name": tn})
            for t, fn, tn in rows]


@router.post("/transfers", response_model=CashTransferResponse, status_code=httpstatus.HTTP_201_CREATED)
async def create_transfer(payload: CashTransferCreate, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    if payload.from_account_id == payload.to_account_id:
        raise HTTPException(status_code=400, detail="Rekening asal & tujuan tak boleh sama.")
    await _get_account(db, ctx.tenant_id, payload.from_account_id)
    await _get_account(db, ctx.tenant_id, payload.to_account_id)
    t = CashTransfer(tenant_id=ctx.tenant_id, **payload.model_dump())
    db.add(t); await db.flush()
    await record_audit(db, ctx.tenant_id, ctx.user_id, "TRANSFER", "cash_transfers", t.id,
                       new_data={"amount": str(t.amount)})
    await db.commit(); await db.refresh(t)
    return CashTransferResponse.model_validate(t)


# ── Pindahkan entri ke rekening lain (reassign) ──
@router.patch("/entries/{entry_id}/account", response_model=CashBookEntryResponse)
async def reassign_entry_account(entry_id: uuid.UUID, payload: EntryAccountUpdate, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    entry = (await db.execute(select(CashBookEntry).where(
        CashBookEntry.id == entry_id, CashBookEntry.tenant_id == ctx.tenant_id))).scalar_one_or_none()
    if entry is None:
        raise HTTPException(status_code=httpstatus.HTTP_404_NOT_FOUND, detail="Baris kas tidak ditemukan")
    if payload.account_id is not None:
        await _get_account(db, ctx.tenant_id, payload.account_id)
    entry.account_id = payload.account_id
    # simpan juga ke SUMBER agar tak tertimpa saat re-sync
    _src_model = {"payment": Payment, "expense": Expense, "notary_fee": NotaryFee}.get(entry.source_type)
    if _src_model is not None:
        src = (await db.execute(select(_src_model).where(
            _src_model.id == entry.source_id, _src_model.tenant_id == ctx.tenant_id))).scalar_one_or_none()
        if src is not None:
            src.cash_account_id = payload.account_id
    await db.flush(); await db.commit()
    # re-fetch dgn relasi kategori + nama akun (hindari lazy-load di luar greenlet)
    row = (await db.execute(
        select(CashBookEntry, Client.full_name, Project.name, CashAccount.name)
        .options(selectinload(CashBookEntry.category))
        .outerjoin(Client, Client.id == CashBookEntry.client_id)
        .outerjoin(Project, Project.id == CashBookEntry.project_id)
        .outerjoin(CashAccount, CashAccount.id == CashBookEntry.account_id)
        .where(CashBookEntry.id == entry_id))).first()
    e, cname, pname, acc_name = row
    return CashBookEntryResponse.model_validate(e).model_copy(update={
        "client_name": cname, "project_name": pname, "account_name": acc_name})


# ═══════════════════════ REKONSILIASI (Slice 2, manual) ═══════════════════════
@router.patch("/entries/{entry_id}/cleared")
async def set_entry_cleared(entry_id: uuid.UUID, payload: ClearedUpdate, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    e = (await db.execute(select(CashBookEntry).where(
        CashBookEntry.id == entry_id, CashBookEntry.tenant_id == ctx.tenant_id))).scalar_one_or_none()
    if e is None:
        raise HTTPException(status_code=httpstatus.HTTP_404_NOT_FOUND, detail="Baris kas tidak ditemukan")
    e.is_cleared = payload.is_cleared
    await db.commit()
    return {"ok": True}


@router.patch("/transfers/{transfer_id}/cleared")
async def set_transfer_cleared(transfer_id: uuid.UUID, payload: ClearedUpdate, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    t = (await db.execute(select(CashTransfer).where(
        CashTransfer.id == transfer_id, CashTransfer.tenant_id == ctx.tenant_id))).scalar_one_or_none()
    if t is None:
        raise HTTPException(status_code=httpstatus.HTTP_404_NOT_FOUND, detail="Transfer tidak ditemukan")
    t.is_cleared = payload.is_cleared
    await db.commit()
    return {"ok": True}


async def _reconcile_data(db, tenant_id, account_id, as_of):
    """Kumpulkan gerakan rekening s/d as_of + hitung saldo buku & saldo cleared."""
    a = await _get_account(db, tenant_id, account_id)
    opening = Decimal(a.opening_balance or 0)
    movements = []
    book = opening
    cleared = opening
    # entri kas
    erows = (await db.execute(
        select(CashBookEntry).where(
            CashBookEntry.tenant_id == tenant_id, CashBookEntry.account_id == account_id,
            CashBookEntry.date <= as_of).order_by(CashBookEntry.date))).scalars().all()
    for e in erows:
        amt = Decimal(e.amount or 0)
        signed = amt if e.direction == CashDirection.IN else -amt
        book += signed
        if e.is_cleared:
            cleared += signed
        movements.append(ReconMovement(id=e.id, kind="entry", date=e.date, description=e.description,
                                       direction="in" if e.direction == CashDirection.IN else "out",
                                       amount=amt, is_cleared=e.is_cleared))
    # transfer masuk/keluar rekening
    trows = (await db.execute(
        select(CashTransfer).where(
            CashTransfer.tenant_id == tenant_id, CashTransfer.is_deleted == False,  # noqa: E712
            CashTransfer.date <= as_of,
            ((CashTransfer.from_account_id == account_id) | (CashTransfer.to_account_id == account_id)))
        .order_by(CashTransfer.date))).scalars().all()
    for t in trows:
        amt = Decimal(t.amount or 0)
        is_in = t.to_account_id == account_id
        signed = amt if is_in else -amt
        book += signed
        if t.is_cleared:
            cleared += signed
        movements.append(ReconMovement(id=t.id, kind="transfer", date=t.date,
                                       description=(t.notes or "Transfer antar rekening"),
                                       direction="in" if is_in else "out", amount=amt, is_cleared=t.is_cleared))
    movements.sort(key=lambda m: m.date)
    return a, opening, book, cleared, movements


@router.get("/accounts/{account_id}/reconcile", response_model=ReconcileView)
async def reconcile_view(account_id: uuid.UUID, as_of: date = Query(...), ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    a, opening, book, cleared, movements = await _reconcile_data(db, ctx.tenant_id, account_id, as_of)
    return ReconcileView(account_id=a.id, account_name=a.name, as_of=as_of,
                         opening_balance=opening, book_balance=book, cleared_balance=cleared, movements=movements)


@router.post("/accounts/{account_id}/reconcile", response_model=ReconciliationRow)
async def save_reconciliation(account_id: uuid.UUID, payload: ReconcileSaveRequest, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    a, opening, book, cleared, _ = await _reconcile_data(db, ctx.tenant_id, account_id, payload.statement_date)
    diff = Decimal(payload.statement_balance) - cleared
    rec = CashReconciliation(tenant_id=ctx.tenant_id, account_id=a.id, statement_date=payload.statement_date,
                             statement_balance=payload.statement_balance, book_balance=book,
                             cleared_balance=cleared, difference=diff, note=payload.note, created_by_id=ctx.user_id)
    db.add(rec); await db.flush()
    await record_audit(db, ctx.tenant_id, ctx.user_id, "RECONCILE", "cash_accounts", a.id,
                       new_data={"statement_date": str(payload.statement_date), "difference": str(diff)})
    await db.commit(); await db.refresh(rec)
    return ReconciliationRow.model_validate(rec)


@router.get("/accounts/{account_id}/reconciliations", response_model=list[ReconciliationRow])
async def list_reconciliations(account_id: uuid.UUID, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    rows = (await db.execute(
        select(CashReconciliation).where(
            CashReconciliation.tenant_id == ctx.tenant_id, CashReconciliation.account_id == account_id)
        .order_by(CashReconciliation.statement_date.desc()).limit(50))).scalars().all()
    return [ReconciliationRow.model_validate(r) for r in rows]


# ═══════════════════════ IMPOR MUTASI BANK (Slice 3, auto-cocok cleared) ═══════════════════════
def _mut_field(h):
    from app.api.v1.endpoints.import_data import _norm_header
    n = _norm_header(h)
    if n.startswith("tanggal") or n.startswith("tgl") or "date" in n: return "date"
    if "debit" in n or "keluar" in n or "debet" in n: return "debit"
    if "kredit" in n or "credit" in n or "masuk" in n: return "kredit"
    if n.startswith("jumlah") or n.startswith("nominal") or n.startswith("mutasi") or n.startswith("nilai"): return "amount"
    if "keterangan" in n or "uraian" in n or "berita" in n or n.startswith("deskrip"): return "desc"
    if n.startswith("tipe") or n.startswith("dc") or n.startswith("d/c") or n == "dk": return "dc"
    return None


def _read_mutations(contents: bytes):
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="File bukan Excel (.xlsx) yang valid.")
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return []
    col = {}
    for i, h in enumerate(header):
        f = _mut_field(h)
        if f and f not in col:
            col[f] = i
    if "date" not in col or not ({"debit", "kredit"} & set(col)) and "amount" not in col:
        raise HTTPException(status_code=400, detail="Header wajib: Tanggal + (Debit/Kredit atau Jumlah).")
    out = []
    rn = 1
    for raw in it:
        rn += 1
        def g(k):
            i = col.get(k)
            return raw[i] if (i is not None and i < len(raw)) else None
        rec = {"date": g("date"), "debit": g("debit"), "kredit": g("kredit"),
               "amount": g("amount"), "desc": g("desc"), "dc": g("dc")}
        if all((v is None or str(v).strip() == "") for v in rec.values()):
            continue
        if str(rec.get("desc") or "").strip().upper().startswith("CONTOH"):
            continue
        out.append((rn, rec))
    return out


@router.post("/accounts/{account_id}/mutations", response_model=MutationImportResult)
async def import_mutations(
    account_id: uuid.UUID,
    file: UploadFile = File(...),
    dry_run: bool = Query(True),
    tolerance_days: int = Query(5, ge=0, le=31),
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    """Impor mutasi rekening koran (xlsx) → cocokkan otomatis ke transaksi rekening ini (arah+jumlah,
    tanggal ±toleransi) → tandai CLEARED. dry_run=true hanya pratinjau. Baris bank yang tak cocok
    dilaporkan (mungkin transaksi belum tercatat: biaya admin, bunga, dll)."""
    from app.api.v1.endpoints.import_data import _to_decimal, _to_date
    a = await _get_account(db, ctx.tenant_id, account_id)
    contents = await file.read()
    rows = _read_mutations(contents)

    # gerakan rekening yg BELUM cleared (entri + transfer)
    movements = []
    erows = (await db.execute(select(CashBookEntry).where(
        CashBookEntry.tenant_id == ctx.tenant_id, CashBookEntry.account_id == a.id,
        CashBookEntry.is_cleared == False))).scalars().all()  # noqa: E712
    for e in erows:
        movements.append({"obj": e, "kind": "entry", "date": e.date,
                          "dir": "in" if e.direction == CashDirection.IN else "out",
                          "amt": Decimal(e.amount or 0), "desc": e.description})
    trows = (await db.execute(select(CashTransfer).where(
        CashTransfer.tenant_id == ctx.tenant_id, CashTransfer.is_deleted == False,  # noqa: E712
        CashTransfer.is_cleared == False,
        ((CashTransfer.from_account_id == a.id) | (CashTransfer.to_account_id == a.id))))).scalars().all()
    for tr in trows:
        movements.append({"obj": tr, "kind": "transfer", "date": tr.date,
                          "dir": "in" if tr.to_account_id == a.id else "out",
                          "amt": Decimal(tr.amount or 0), "desc": tr.notes or "Transfer antar rekening"})

    used = set()
    result_rows = []
    to_clear = []
    matched = already = nomatch = 0
    for rn, rec in rows:
        errs = []
        try:
            bdate = _to_date(rec["date"])
        except ValueError:
            bdate = None
        # arah & jumlah
        deb = kre = None
        try:
            deb = _to_decimal(rec["debit"])
            kre = _to_decimal(rec["kredit"])
        except Exception:
            errs.append("angka Debit/Kredit tak valid")
        amt = None; direction = None
        if kre and kre > 0:
            direction, amt = "in", kre
        elif deb and deb > 0:
            direction, amt = "out", deb
        elif rec["amount"] not in (None, ""):
            try:
                av = _to_decimal(rec["amount"])
            except Exception:
                av = None
            dc = str(rec["dc"] or "").strip().lower()
            if av is not None:
                if dc in ("d", "db", "debit", "k", "keluar") or av < 0:
                    direction, amt = "out", abs(av)
                else:
                    direction, amt = "in", abs(av)
        desc = str(rec["desc"] or "").strip() or "(mutasi)"
        if amt is None or amt <= 0:
            errs.append("jumlah/ arah tak terbaca")

        if errs:
            result_rows.append(MutationRow(row=rn, tgl=bdate, description=desc, direction=direction or "-",
                                           amount=amt or Decimal(0), status="error", note="; ".join(errs)))
            continue

        # cari kandidat: arah sama, jumlah sama, dalam toleransi tanggal, belum dipakai
        cands = []
        for idx, m in enumerate(movements):
            if idx in used or m["dir"] != direction or m["amt"] != amt:
                continue
            if bdate is not None and abs((m["date"] - bdate).days) > tolerance_days:
                continue
            cands.append((idx, m))
        if not cands:
            nomatch += 1
            result_rows.append(MutationRow(row=rn, tgl=bdate, description=desc, direction=direction,
                                           amount=amt, status="no_match", note="tak ada transaksi cocok (mungkin belum dicatat)"))
            continue
        # pilih tanggal terdekat
        cands.sort(key=lambda c: abs((c[1]["date"] - bdate).days) if bdate else 0)
        idx, m = cands[0]
        used.add(idx)
        matched += 1
        to_clear.append(m)
        result_rows.append(MutationRow(row=rn, tgl=bdate, description=desc, direction=direction, amount=amt,
                                       status="matched", entry_id=m["obj"].id, entry_kind=m["kind"],
                                       entry_desc=m["desc"], note=("beberapa kandidat" if len(cands) > 1 else None)))

    if not dry_run and to_clear:
        from sqlalchemy import update as _update
        e_ids = [m["obj"].id for m in to_clear if m["kind"] == "entry"]
        t_ids = [m["obj"].id for m in to_clear if m["kind"] == "transfer"]
        if e_ids:
            await db.execute(_update(CashBookEntry).where(CashBookEntry.id.in_(e_ids)).values(is_cleared=True))
        if t_ids:
            await db.execute(_update(CashTransfer).where(CashTransfer.id.in_(t_ids)).values(is_cleared=True))
        await record_audit(db, ctx.tenant_id, ctx.user_id, "IMPORT_MUTATION", "cash_accounts", a.id,
                           new_data={"matched": matched})
        await db.commit()

    return MutationImportResult(dry_run=dry_run, total=len(result_rows), matched=matched,
                                already_cleared=already, no_match=nomatch, rows=result_rows)


@router.get("/mutations/template")
async def mutations_template(ctx: AuthContext = Depends(get_current_context)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "MUTASI"
    hdr = ["Tanggal", "Keterangan", "Debit (keluar)", "Kredit (masuk)"]
    fill = PatternFill("solid", fgColor="1E3A5F")
    widths = [14, 40, 16, 16]
    for i, (h, w) in enumerate(zip(hdr, widths), start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = fill; c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[chr(64 + i)].width = w
    ex = ["20/05/2025", "CONTOH — hapus baris ini. DP dari Budi", "", 25000000]
    for i, v in enumerate(ex, start=1):
        cc = ws.cell(row=2, column=i, value=v)
        cc.font = Font(name="Arial", italic=True, size=10, color="C08A2B")
        if i == 4:
            cc.number_format = "#,##0"
    ws.freeze_panes = "A2"
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": 'attachment; filename="Template_Mutasi_Bank.xlsx"'})


# ═══════════════════════ BIAYA OPERASIONAL (overhead perusahaan, non-proyek) ═══════════════════════
from app.models.opex import OpexCategory, OperationalExpense  # noqa: E402
from app.core.cashbook import sync_opex_cashbook, seed_default_opex_categories  # noqa: E402
from app.schemas.opex import (  # noqa: E402
    OpexCategoryCreate, OpexCategoryUpdate, OpexCategoryResponse,
    OperationalExpenseCreate, OperationalExpenseUpdate, OperationalExpenseResponse,
    OpexCategoryTotal, OpexList,
)

_ONOTDEL = lambda m: m.is_deleted == False  # noqa: E731, E712


@router.get("/opex-categories", response_model=list[OpexCategoryResponse])
async def list_opex_categories(ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    await seed_default_opex_categories(db, ctx.tenant_id)   # tenant lama tanpa kategori → isi default
    await db.commit()
    rows = (await db.execute(select(OpexCategory).where(
        OpexCategory.tenant_id == ctx.tenant_id, _ONOTDEL(OpexCategory), OpexCategory.is_active == True)  # noqa: E712
        .order_by(OpexCategory.sort_order, OpexCategory.name))).scalars().all()
    return rows


@router.post("/opex-categories", response_model=OpexCategoryResponse, status_code=201)
async def create_opex_category(payload: OpexCategoryCreate, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    nxt = (await db.execute(select(func.coalesce(func.max(OpexCategory.sort_order), 0)).where(
        OpexCategory.tenant_id == ctx.tenant_id))).scalar() or 0
    c = OpexCategory(tenant_id=ctx.tenant_id, name=payload.name.strip(), sort_order=nxt + 1)
    db.add(c); await db.commit(); await db.refresh(c)
    return c


@router.patch("/opex-categories/{cid}", response_model=OpexCategoryResponse)
async def update_opex_category(cid: uuid.UUID, payload: OpexCategoryUpdate, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    c = (await db.execute(select(OpexCategory).where(OpexCategory.id == cid, OpexCategory.tenant_id == ctx.tenant_id, _ONOTDEL(OpexCategory)))).scalar_one_or_none()
    if c is None:
        from fastapi import HTTPException
        raise HTTPException(404, "Kategori tidak ditemukan")
    for f, v in payload.model_dump(exclude_unset=True).items():
        setattr(c, f, v.strip() if f == "name" and isinstance(v, str) else v)
    await db.commit(); await db.refresh(c)
    return c


@router.delete("/opex-categories/{cid}", status_code=204)
async def delete_opex_category(cid: uuid.UUID, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    c = (await db.execute(select(OpexCategory).where(OpexCategory.id == cid, OpexCategory.tenant_id == ctx.tenant_id))).scalar_one_or_none()
    if c is not None:
        c.is_deleted = True
        await db.commit()


async def _opex_resp(o: OperationalExpense) -> OperationalExpenseResponse:
    r = OperationalExpenseResponse.model_validate(o)
    r.category_name = o.category.name if o.category else None
    return r


@router.get("/opex", response_model=OpexList)
async def list_opex(date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
                    category_id: Optional[uuid.UUID] = Query(None),
                    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    conds = [OperationalExpense.tenant_id == ctx.tenant_id, _ONOTDEL(OperationalExpense)]
    if date_from:
        conds.append(OperationalExpense.expense_date >= date.fromisoformat(date_from))
    if date_to:
        conds.append(OperationalExpense.expense_date <= date.fromisoformat(date_to))
    if category_id:
        conds.append(OperationalExpense.opex_category_id == category_id)
    rows = (await db.execute(select(OperationalExpense).options(selectinload(OperationalExpense.category))
            .where(*conds).order_by(OperationalExpense.expense_date.desc().nullslast(), OperationalExpense.created_at.desc()))).scalars().all()
    total = sum((Decimal(r.amount) for r in rows if r.is_paid), Decimal(0))
    total_unpaid = sum((Decimal(r.amount) for r in rows if not r.is_paid), Decimal(0))
    agg: dict[str, Decimal] = {}
    for r in rows:
        if r.is_paid:
            key = r.category.name if r.category else "Tanpa kategori"
            agg[key] = agg.get(key, Decimal(0)) + Decimal(r.amount)
    by_cat = [OpexCategoryTotal(name=k, total=v) for k, v in sorted(agg.items(), key=lambda x: -x[1])]
    return OpexList(rows=[await _opex_resp(r) for r in rows], total=total, total_unpaid=total_unpaid, by_category=by_cat)


@router.post("/opex", response_model=OperationalExpenseResponse, status_code=201)
async def create_opex(payload: OperationalExpenseCreate, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    d = payload.model_dump()
    is_paid = d.pop("is_paid", True)
    o = OperationalExpense(tenant_id=ctx.tenant_id, created_by_id=ctx.user_id, is_paid=is_paid,
                           paid_at=(d.get("expense_date") or date.today()) if is_paid else None, **d)
    db.add(o); await db.flush()
    await sync_opex_cashbook(db, ctx.tenant_id, o)
    await record_audit(db, ctx.tenant_id, ctx.user_id, "CREATE", "operational_expenses", o.id,
                       new_data={"amount": str(o.amount), "desc": o.description})
    await db.commit()
    o = (await db.execute(select(OperationalExpense).options(selectinload(OperationalExpense.category)).where(OperationalExpense.id == o.id))).scalar_one()
    return await _opex_resp(o)


@router.patch("/opex/{oid}", response_model=OperationalExpenseResponse)
async def update_opex(oid: uuid.UUID, payload: OperationalExpenseUpdate, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    o = (await db.execute(select(OperationalExpense).where(OperationalExpense.id == oid, OperationalExpense.tenant_id == ctx.tenant_id, _ONOTDEL(OperationalExpense)))).scalar_one_or_none()
    if o is None:
        from fastapi import HTTPException
        raise HTTPException(404, "Biaya tidak ditemukan")
    data = payload.model_dump(exclude_unset=True)
    for f, v in data.items():
        setattr(o, f, v)
    if o.is_paid and o.paid_at is None:
        o.paid_at = o.expense_date or date.today()
    if not o.is_paid:
        o.paid_at = None
    await db.flush()
    await sync_opex_cashbook(db, ctx.tenant_id, o)
    await db.commit()
    o = (await db.execute(select(OperationalExpense).options(selectinload(OperationalExpense.category)).where(OperationalExpense.id == o.id))).scalar_one()
    return await _opex_resp(o)


@router.delete("/opex/{oid}", status_code=204)
async def delete_opex(oid: uuid.UUID, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    o = (await db.execute(select(OperationalExpense).where(OperationalExpense.id == oid, OperationalExpense.tenant_id == ctx.tenant_id))).scalar_one_or_none()
    if o is not None:
        o.is_deleted = True
        await db.flush()
        await sync_opex_cashbook(db, ctx.tenant_id, o)
        await db.commit()
