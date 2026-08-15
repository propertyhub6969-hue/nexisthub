import re
import secrets
import uuid
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from pydantic import BaseModel
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.api.deps import get_current_context, AuthContext
from app.models.marketing import Lead, Prospect, Client, ProspectStatus, ClientStatus, ClientPaymentType
from app.models.property import Project, Unit, UnitStatus
from app.models.payment import Payment, PaymentSchedule, ScheduleStatus, PaymentSource, PaymentApprovalStatus
from app.models.kpr import KprApplication, KprStage, Bank
from app.models.construction import UnitConstruction, ConstructionStage, ConstructionProgressLog
from app.models.tax import TaxRecord, TaxType, TaxStatus, MonthlyTaxShareLink, NotaryFee
from app.models.document import Document
from app.models.cashbook import CashBookEntry, AccountCategory, CashDirection, CashAccount
from app.models.expense import Expense, ExpenseCategory
from app.models.contractor import ContractorContract

_EXPENSE_LABEL = {
    ExpenseCategory.MATERIAL: "Material", ExpenseCategory.UPAH: "Upah",
    ExpenseCategory.KONTRAKTOR: "Kontraktor", ExpenseCategory.KELISTRIKAN: "Kelistrikan",
    ExpenseCategory.OPERASIONAL: "Operasional", ExpenseCategory.PERIZINAN: "Perizinan",
    ExpenseCategory.AIR_PDAM: "Air / PDAM",
    ExpenseCategory.LAIN: "Lain-lain",
}

router = APIRouter()

# Tahap yang menandakan pengajuan sudah DISETUJUI bank (SP3K = surat persetujuan kredit ke atas).
APPROVED_STAGES = (KprStage.SP3K, KprStage.PERSIAPAN_AKAD, KprStage.AKAD_KREDIT, KprStage.PENCAIRAN)


class DashboardStats(BaseModel):
    leads_total: int
    prospects_active: int
    clients_total: int
    units_total: int
    units_available: int
    units_booked: int
    units_sold: int
    units_held_no_client: int = 0   # unit ditahan booking tapi belum ada data Pembeli
    payments_this_month: Decimal
    total_contract: Decimal
    total_paid: Decimal
    outstanding: Decimal
    overdue_count: int


async def _count(db, model, *conds) -> int:
    return await db.scalar(select(func.count()).select_from(model).where(*conds)) or 0


@router.get("/dashboard", response_model=DashboardStats)
async def get_dashboard(
    ctx: AuthContext = Depends(get_current_context),
    db: AsyncSession = Depends(get_db),
):
    t = ctx.tenant_id
    month_start = date.today().replace(day=1)

    leads_total = await _count(db, Lead, Lead.tenant_id == t)
    prospects_active = await _count(db, Prospect, Prospect.tenant_id == t, Prospect.status == ProspectStatus.ACTIVE)
    clients_total = await _count(db, Client, Client.tenant_id == t, Client.is_deleted == False)  # noqa: E712

    units_total = await _count(db, Unit, Unit.tenant_id == t)
    units_available = await _count(db, Unit, Unit.tenant_id == t, Unit.status == UnitStatus.AVAILABLE)
    units_booked = await _count(db, Unit, Unit.tenant_id == t, Unit.status == UnitStatus.BOOKED)
    # Unit ditahan (Booking/DP atau Terjual) TAPI belum punya data Pembeli — biasanya berasal dari
    # permintaan booking agen yang sudah diterima namun belum dilanjutkan jadi Pembeli.
    # Sengaja ditampilkan supaya unit tak diam-diam tertahan tanpa pemilik data.
    units_held_no_client = await db.scalar(
        select(func.count()).select_from(Unit).where(
            Unit.tenant_id == t,
            Unit.status.in_([UnitStatus.BOOKED, UnitStatus.SOLD, UnitStatus.HANDOVER]),
            ~select(Client.id).where(Client.unit_id == Unit.id, Client.is_deleted == False).exists(),  # noqa: E712
        )
    ) or 0
    units_sold = await _count(db, Unit, Unit.tenant_id == t,
                              Unit.status.in_([UnitStatus.SOLD, UnitStatus.HANDOVER]))

    _approved = Payment.approval_status == PaymentApprovalStatus.APPROVED
    payments_this_month = Decimal(await db.scalar(
        select(func.coalesce(func.sum(Payment.amount), 0)).where(
            Payment.tenant_id == t, Payment.is_deleted == False,  # noqa: E712
            Payment.payment_date >= month_start, _approved)
    ))
    total_contract = Decimal(await db.scalar(
        select(func.coalesce(func.sum(Client.contract_value), 0)).where(
            Client.tenant_id == t, Client.is_deleted == False,  # noqa: E712
            Client.status != ClientStatus.INACTIVE)
    ))
    total_paid = Decimal(await db.scalar(
        select(func.coalesce(func.sum(Payment.amount), 0)).where(
            Payment.tenant_id == t, Payment.is_deleted == False, _approved)  # noqa: E712
    ))
    overdue_count = await _count(db, PaymentSchedule, PaymentSchedule.tenant_id == t,
                                 PaymentSchedule.is_deleted == False,  # noqa: E712
                                 PaymentSchedule.status == ScheduleStatus.PENDING,
                                 PaymentSchedule.due_date < date.today())

    return DashboardStats(
        leads_total=leads_total, prospects_active=prospects_active, clients_total=clients_total,
        units_total=units_total, units_available=units_available, units_booked=units_booked, units_sold=units_sold,
        units_held_no_client=units_held_no_client,
        payments_this_month=payments_this_month, total_contract=total_contract, total_paid=total_paid,
        outstanding=total_contract - total_paid, overdue_count=overdue_count,
    )


# ═══════════════════════ LAPORAN: REJECTION-RATE KPR PER BANK ═══════════════════════
class KprRejectionBank(BaseModel):
    bank_id: Optional[uuid.UUID]
    bank_name: str
    total: int          # total pengajuan (semua tahap) ke bank ini
    rejected: int       # jumlah ditolak
    approved: int       # sudah disetujui (SP3K/Akad/Pencairan), belum ditolak
    in_process: int     # masih proses, belum ada keputusan
    rejection_rate: float   # rejected / total * 100 (dibulatkan 1 desimal)
    avg_days_to_akad: Optional[float] = None  # rata² lama pemberkasan Collect Berkas→Akad (hari)
    akad_samples: int = 0                      # jumlah pengajuan yang dipakai utk rata² durasi


class KprRejectionReport(BaseModel):
    banks: list[KprRejectionBank]
    total: int
    rejected: int
    approved: int
    in_process: int
    rejection_rate: float
    avg_days_to_akad: Optional[float] = None
    akad_samples: int = 0


@router.get("/kpr-rejection", response_model=KprRejectionReport)
async def kpr_rejection(
    ctx: AuthContext = Depends(get_current_context),
    db: AsyncSession = Depends(get_db),
):
    """Rejection-rate pengajuan KPR per bank — bantu developer pilih bank penyalur yang paling tinggi approval-nya."""
    t = ctx.tenant_id
    rows = (await db.execute(
        select(KprApplication).options(selectinload(KprApplication.bank))
        .where(KprApplication.tenant_id == t, KprApplication.is_deleted == False)  # noqa: E712
    )).scalars().all()

    # agregasi per bank
    buckets: dict[Optional[uuid.UUID], dict] = {}
    for k in rows:
        key = k.bank_id
        b = buckets.setdefault(key, {
            "bank_id": key,
            "bank_name": (k.bank.name if k.bank else "(Tanpa bank)"),
            "total": 0, "rejected": 0, "approved": 0, "in_process": 0,
            "akad_days_sum": 0, "akad_samples": 0,
        })
        b["total"] += 1
        if k.rejected_date is not None:
            b["rejected"] += 1
        elif k.stage in APPROVED_STAGES:
            b["approved"] += 1
        else:
            b["in_process"] += 1
        # durasi pemberkasan: Collect Berkas (submitted_date) → Akad (akad_date)
        if k.submitted_date is not None and k.akad_date is not None:
            days = (k.akad_date - k.submitted_date).days
            if days >= 0:
                b["akad_days_sum"] += days
                b["akad_samples"] += 1

    def rate(rejected: int, total: int) -> float:
        return round(rejected / total * 100, 1) if total else 0.0

    def avg_days(b: dict) -> Optional[float]:
        return round(b["akad_days_sum"] / b["akad_samples"], 1) if b["akad_samples"] else None

    banks = [
        KprRejectionBank(
            bank_id=b["bank_id"], bank_name=b["bank_name"], total=b["total"],
            rejected=b["rejected"], approved=b["approved"], in_process=b["in_process"],
            rejection_rate=rate(b["rejected"], b["total"]),
            avg_days_to_akad=avg_days(b), akad_samples=b["akad_samples"],
        )
        for b in buckets.values()
    ]
    # urutkan: rejection-rate tertinggi dulu, lalu terbanyak pengajuan; bank tanpa nama di akhir
    banks.sort(key=lambda x: (x.bank_id is None, -x.rejection_rate, -x.total, x.bank_name.lower()))

    total = sum(b.total for b in banks)
    rejected = sum(b.rejected for b in banks)
    approved = sum(b.approved for b in banks)
    in_process = sum(b.in_process for b in banks)
    akad_days_sum = sum(bk["akad_days_sum"] for bk in buckets.values())
    akad_samples = sum(bk["akad_samples"] for bk in buckets.values())

    return KprRejectionReport(
        banks=banks, total=total, rejected=rejected, approved=approved, in_process=in_process,
        rejection_rate=rate(rejected, total),
        avg_days_to_akad=(round(akad_days_sum / akad_samples, 1) if akad_samples else None),
        akad_samples=akad_samples,
    )


# ═══════════════════════ LAPORAN: ARUS KAS (pembeli vs bank + retensi) ═══════════════════════
class CashflowMonth(BaseModel):
    month: str              # "YYYY-MM"
    from_buyer: Decimal
    from_bank: Decimal
    total: Decimal


class CashflowCategoryTotal(BaseModel):
    category_name: str
    direction: str                # 'in' | 'out'
    total: Decimal


class CashflowOutMonth(BaseModel):
    month: str                    # "YYYY-MM"
    by_category: list[Decimal]    # sejajar dgn out_category_names
    total: Decimal


class CashflowBreakdownItem(BaseModel):
    label: str
    total: Decimal


class CashflowReport(BaseModel):
    total_contract: Decimal       # total nilai kontrak (pembeli aktif)
    from_buyer: Decimal           # kas masuk dari pembeli (DP/cicilan)
    from_bank: Decimal            # kas masuk dari bank (pencairan KPR)
    total_in: Decimal             # total kas masuk
    kpr_plafond_total: Decimal    # total komitmen plafon KPR
    buyer_remaining: Decimal      # sisa kewajiban pembeli (piutang pembeli)
    retention_remaining: Decimal  # retensi menunggu pencairan bank
    months: list[CashflowMonth]   # tren bulanan (kronologis, maks 12 bln terakhir yang ada transaksi)
    # Ringkasan Buku Kas per kategori (ledger riil: pembayaran approved + biaya/notaris dibayar).
    # Beda basis dgn angka penjualan di atas — ini termasuk kas KELUAR.
    ledger_in: Decimal = Decimal(0)
    ledger_out: Decimal = Decimal(0)
    ledger_saldo: Decimal = Decimal(0)
    by_category: list[CashflowCategoryTotal] = []
    # Tren bulanan kas KELUAR dipecah per kategori (kolom = out_category_names)
    out_category_names: list[str] = []
    out_months: list[CashflowOutMonth] = []
    # Rincian kategori Biaya Notaris/Legal per jenis jasa (AJB, BBN, Balik Nama, dst)
    notary_breakdown: list[CashflowBreakdownItem] = []
    # Rincian kategori Biaya Operasional per jenis biaya (Material, Upah, Kontraktor, dst)
    expense_breakdown: list[CashflowBreakdownItem] = []


@router.get("/cashflow", response_model=CashflowReport)
async def cashflow(
    cat_from: Optional[date] = Query(None),   # filter periode KHUSUS ringkasan kategori (ledger)
    cat_to: Optional[date] = Query(None),
    ctx: AuthContext = Depends(get_current_context),
    db: AsyncSession = Depends(get_db),
):
    """Arus kas se-tenant: kas masuk dari pembeli vs bank, ditambah sisa kewajiban pembeli & retensi bank."""
    t = ctx.tenant_id
    # hanya pembayaran yang sudah disetujui finance — pending/rejected belum dihitung sbg kas
    notdel_p = (Payment.is_deleted == False) & (Payment.approval_status == PaymentApprovalStatus.APPROVED)  # noqa: E712

    # Hanya hitung pembayaran milik PEMBELI yang masih ada (bukan soft-deleted/orphan). Pembatalan
    # deal pakai status INACTIVE (tetap terhitung sbg kas diterima), soft-delete = data keliru → dikecualikan.
    async def _sum_payments(source: PaymentSource) -> Decimal:
        return Decimal(await db.scalar(
            select(func.coalesce(func.sum(Payment.amount), 0))
            .select_from(Payment).join(Client, Client.id == Payment.client_id)
            .where(Payment.tenant_id == t, Payment.source == source, notdel_p,
                   Client.is_deleted == False)  # noqa: E712
        ))

    from_buyer = await _sum_payments(PaymentSource.PEMBELI)
    from_bank = await _sum_payments(PaymentSource.BANK)
    total_in = from_buyer + from_bank

    total_contract = Decimal(await db.scalar(
        select(func.coalesce(func.sum(Client.contract_value), 0)).where(
            Client.tenant_id == t, Client.is_deleted == False,  # noqa: E712
            Client.status != ClientStatus.INACTIVE)
    ))

    # Sisa kewajiban pembeli & retensi — dihitung per-pembeli lalu dijumlah (clamp ≥ 0),
    # konsisten dgn ringkasan pembayaran per-pembeli (harga − dari_pembeli − plafon; plafon − dari_bank).
    clients = (await db.execute(
        select(Client.id, Client.contract_value).where(
            Client.tenant_id == t, Client.is_deleted == False,  # noqa: E712
            Client.status != ClientStatus.INACTIVE)
    )).all()

    buyer_rows = (await db.execute(
        select(Payment.client_id, func.coalesce(func.sum(Payment.amount), 0))
        .where(Payment.tenant_id == t, Payment.source == PaymentSource.PEMBELI, notdel_p)
        .group_by(Payment.client_id)
    )).all()
    bank_rows = (await db.execute(
        select(Payment.client_id, func.coalesce(func.sum(Payment.amount), 0))
        .where(Payment.tenant_id == t, Payment.source == PaymentSource.BANK, notdel_p)
        .group_by(Payment.client_id)
    )).all()
    buyer_by_client = {cid: Decimal(v) for cid, v in buyer_rows}
    bank_by_client = {cid: Decimal(v) for cid, v in bank_rows}

    # plafon KPR terbaru per pembeli — hanya DIHITUNG sbg komitmen bila stage ≥ Akad Kredit
    # (sebelum akad pinjaman belum final → plafon belum menutup kewajiban pembeli / belum retensi).
    kpr_rows = (await db.execute(
        select(KprApplication.client_id, KprApplication.plafond, KprApplication.stage)
        .where(KprApplication.tenant_id == t, KprApplication.is_deleted == False)  # noqa: E712
        .order_by(KprApplication.client_id, KprApplication.created_at.desc())
    )).all()
    committed_by_client: dict = {}
    for cid, plaf, stage in kpr_rows:
        if cid not in committed_by_client:   # baris pertama per client = terbaru (created_at desc)
            committed_by_client[cid] = Decimal(plaf or 0) if stage in (KprStage.AKAD_KREDIT, KprStage.PENCAIRAN) else Decimal(0)

    kpr_plafond_total = Decimal(0)
    buyer_remaining = Decimal(0)
    retention_remaining = Decimal(0)
    for cid, price in clients:
        price = Decimal(price or 0)
        committed = committed_by_client.get(cid, Decimal(0))
        b_paid = buyer_by_client.get(cid, Decimal(0))
        bank_paid = bank_by_client.get(cid, Decimal(0))
        kpr_plafond_total += committed
        buyer_remaining += max(price - b_paid - committed, Decimal(0))
        if committed > 0:
            retention_remaining += max(committed - bank_paid, Decimal(0))

    # Tren bulanan (kas masuk per bulan, pisah sumber)
    ym = func.to_char(Payment.payment_date, "YYYY-MM")
    month_rows = (await db.execute(
        select(
            ym.label("ym"),
            func.coalesce(func.sum(Payment.amount).filter(Payment.source == PaymentSource.PEMBELI), 0),
            func.coalesce(func.sum(Payment.amount).filter(Payment.source == PaymentSource.BANK), 0),
        )
        .select_from(Payment).join(Client, Client.id == Payment.client_id)
        .where(Payment.tenant_id == t, notdel_p, Payment.payment_date.isnot(None),
               Client.is_deleted == False)  # noqa: E712
        .group_by(ym).order_by(ym)
    )).all()
    months = [
        CashflowMonth(month=m, from_buyer=Decimal(fb), from_bank=Decimal(bk), total=Decimal(fb) + Decimal(bk))
        for m, fb, bk in month_rows
    ][-12:]   # maks 12 bulan terakhir yang ada transaksi

    # Ringkasan Buku Kas per kategori (ledger riil — termasuk kas keluar biaya/notaris).
    # Filter periode opsional berlaku KHUSUS di sini (angka penjualan di atas tetap all-time).
    cat_conds = [CashBookEntry.tenant_id == t]
    if cat_from:
        cat_conds.append(CashBookEntry.date >= cat_from)
    if cat_to:
        cat_conds.append(CashBookEntry.date <= cat_to)
    cat_rows = (await db.execute(
        select(AccountCategory.name, CashBookEntry.direction, func.coalesce(func.sum(CashBookEntry.amount), 0))
        .select_from(CashBookEntry)
        .outerjoin(AccountCategory, AccountCategory.id == CashBookEntry.category_id)
        .where(*cat_conds)
        .group_by(AccountCategory.name, CashBookEntry.direction)
    )).all()
    by_category = [
        CashflowCategoryTotal(category_name=name or "Belum dikategorikan", direction=direction.value, total=Decimal(total))
        for name, direction, total in cat_rows
    ]
    by_category.sort(key=lambda r: (r.direction, -r.total))
    ledger_in = sum((c.total for c in by_category if c.direction == CashDirection.IN.value), Decimal(0))
    ledger_out = sum((c.total for c in by_category if c.direction == CashDirection.OUT.value), Decimal(0))

    # Tren bulanan kas KELUAR per kategori (hormati filter periode yg sama)
    ymc = func.to_char(CashBookEntry.date, "YYYY-MM")
    out_rows = (await db.execute(
        select(ymc.label("ym"), AccountCategory.name, func.coalesce(func.sum(CashBookEntry.amount), 0))
        .select_from(CashBookEntry)
        .outerjoin(AccountCategory, AccountCategory.id == CashBookEntry.category_id)
        .where(*cat_conds, CashBookEntry.direction == CashDirection.OUT)
        .group_by(ymc, AccountCategory.name).order_by(ymc)
    )).all()
    cat_totals: dict = {}
    month_map: dict = {}
    for ym, name, total in out_rows:
        nm = name or "Belum dikategorikan"
        cat_totals[nm] = cat_totals.get(nm, Decimal(0)) + Decimal(total)
        month_map.setdefault(ym, {})[nm] = Decimal(total)
    out_category_names = [nm for nm, _ in sorted(cat_totals.items(), key=lambda x: -x[1])]
    out_months = [
        CashflowOutMonth(
            month=ym,
            by_category=[month_map[ym].get(nm, Decimal(0)) for nm in out_category_names],
            total=sum(month_map[ym].values(), Decimal(0)),
        )
        for ym in sorted(month_map.keys())[-12:]
    ]

    # Rincian kategori Biaya Notaris/Legal per jenis jasa (join balik ke NotaryFee via source_id)
    notary_rows = (await db.execute(
        select(NotaryFee.description, func.coalesce(func.sum(CashBookEntry.amount), 0))
        .select_from(CashBookEntry)
        .join(NotaryFee, NotaryFee.id == CashBookEntry.source_id)
        .where(*cat_conds, CashBookEntry.source_type == "notary_fee")
        .group_by(NotaryFee.description).order_by(func.coalesce(func.sum(CashBookEntry.amount), 0).desc())
    )).all()
    notary_breakdown = [CashflowBreakdownItem(label=desc or "—", total=Decimal(total)) for desc, total in notary_rows]

    # Rincian kategori Biaya Operasional per jenis biaya (join balik ke Expense via source_id)
    expense_rows = (await db.execute(
        select(Expense.category, func.coalesce(func.sum(CashBookEntry.amount), 0))
        .select_from(CashBookEntry)
        .join(Expense, Expense.id == CashBookEntry.source_id)
        .where(*cat_conds, CashBookEntry.source_type == "expense")
        .group_by(Expense.category).order_by(func.coalesce(func.sum(CashBookEntry.amount), 0).desc())
    )).all()
    expense_breakdown = [
        CashflowBreakdownItem(label=_EXPENSE_LABEL.get(cat, str(cat)), total=Decimal(total))
        for cat, total in expense_rows
    ]

    return CashflowReport(
        total_contract=total_contract, from_buyer=from_buyer, from_bank=from_bank, total_in=total_in,
        kpr_plafond_total=kpr_plafond_total, buyer_remaining=buyer_remaining,
        retention_remaining=retention_remaining, months=months,
        ledger_in=ledger_in, ledger_out=ledger_out, ledger_saldo=ledger_in - ledger_out,
        by_category=by_category, out_category_names=out_category_names, out_months=out_months,
        notary_breakdown=notary_breakdown, expense_breakdown=expense_breakdown,
    )


# ═══════════════════════ LAPORAN: REKAP PENJUALAN PER PROYEK ═══════════════════════
class SalesProject(BaseModel):
    project_id: uuid.UUID
    project_name: str
    units_total: int
    units_available: int
    units_booked: int
    units_sold: int          # sold + handover
    buyers: int              # pembeli aktif/selesai (bukan batal)
    contract_value: Decimal
    cash_in: Decimal
    remaining: Decimal


class SalesRecapReport(BaseModel):
    projects: list[SalesProject]
    units_total: int
    units_sold: int
    buyers: int
    contract_value: Decimal
    cash_in: Decimal
    remaining: Decimal


@router.get("/sales-recap", response_model=SalesRecapReport)
async def sales_recap(
    ctx: AuthContext = Depends(get_current_context),
    db: AsyncSession = Depends(get_db),
):
    """Rekap penjualan per proyek: status unit, jumlah pembeli, nilai kontrak, kas masuk, sisa."""
    t = ctx.tenant_id

    projects = (await db.execute(
        select(Project.id, Project.name).where(Project.tenant_id == t).order_by(Project.name)
    )).all()

    # jumlah unit per (proyek, status)
    unit_rows = (await db.execute(
        select(Unit.project_id, Unit.status, func.count()).where(Unit.tenant_id == t)
        .group_by(Unit.project_id, Unit.status)
    )).all()
    units_by_proj: dict = {}
    for pid, st, cnt in unit_rows:
        d = units_by_proj.setdefault(pid, {s: 0 for s in UnitStatus})
        d[st] = cnt

    # pembeli aktif/selesai per proyek + nilai kontrak
    client_rows = (await db.execute(
        select(Client.id, Client.project_id, Client.contract_value).where(
            Client.tenant_id == t, Client.is_deleted == False,  # noqa: E712
            Client.status != ClientStatus.INACTIVE)
    )).all()
    # kas masuk per pembeli (hanya yang sudah disetujui finance)
    pay_rows = (await db.execute(
        select(Payment.client_id, func.coalesce(func.sum(Payment.amount), 0))
        .where(Payment.tenant_id == t, Payment.is_deleted == False,  # noqa: E712
               Payment.approval_status == PaymentApprovalStatus.APPROVED)
        .group_by(Payment.client_id)
    )).all()
    paid_by_client = {cid: Decimal(v) for cid, v in pay_rows}

    buyers_by_proj: dict = {}
    contract_by_proj: dict = {}
    cash_by_proj: dict = {}
    for cid, pid, cv in client_rows:
        buyers_by_proj[pid] = buyers_by_proj.get(pid, 0) + 1
        contract_by_proj[pid] = contract_by_proj.get(pid, Decimal(0)) + Decimal(cv or 0)
        cash_by_proj[pid] = cash_by_proj.get(pid, Decimal(0)) + paid_by_client.get(cid, Decimal(0))

    rows: list[SalesProject] = []
    for pid, name in projects:
        u = units_by_proj.get(pid, {})
        available = u.get(UnitStatus.AVAILABLE, 0)
        booked = u.get(UnitStatus.BOOKED, 0)
        sold = u.get(UnitStatus.SOLD, 0) + u.get(UnitStatus.HANDOVER, 0)
        contract = contract_by_proj.get(pid, Decimal(0))
        cash = cash_by_proj.get(pid, Decimal(0))
        rows.append(SalesProject(
            project_id=pid, project_name=name,
            units_total=available + booked + sold, units_available=available,
            units_booked=booked, units_sold=sold,
            buyers=buyers_by_proj.get(pid, 0),
            contract_value=contract, cash_in=cash, remaining=contract - cash,
        ))

    return SalesRecapReport(
        projects=rows,
        units_total=sum(r.units_total for r in rows),
        units_sold=sum(r.units_sold for r in rows),
        buyers=sum(r.buyers for r in rows),
        contract_value=sum((r.contract_value for r in rows), Decimal(0)),
        cash_in=sum((r.cash_in for r in rows), Decimal(0)),
        remaining=sum((r.remaining for r in rows), Decimal(0)),
    )


# ═══════════════════════ LAPORAN: PROGRES KONSTRUKSI ═══════════════════════
CONSTRUCTION_REMINDER_DAYS = 7  # samakan dgn frontend Construction.tsx isLate()


class ConstructionProject(BaseModel):
    project_id: uuid.UUID
    project_name: str
    units_total: int
    avg_percent: float
    done: int            # selesai (stage=selesai atau percent>=100)
    in_progress: int     # sudah mulai, belum selesai
    not_started: int     # persiapan & 0%
    overdue_target: int  # target_date lewat & belum selesai
    late_update: int     # belum update progres > 7 hari (unit belum selesai)


class ConstructionProgressReport(BaseModel):
    projects: list[ConstructionProject]
    units_total: int
    done: int
    overdue_target: int
    late_update: int
    avg_percent: float
    stage_counts: dict[str, int]


@router.get("/construction-progress", response_model=ConstructionProgressReport)
async def construction_progress(
    ctx: AuthContext = Depends(get_current_context),
    db: AsyncSession = Depends(get_db),
):
    """Progres pembangunan per proyek: rata-rata %, tahap, selesai & keterlambatan.
    Unit tanpa baris UnitConstruction dianggap tahap persiapan / 0% (konsisten dgn list_construction)."""
    t = ctx.tenant_id
    today = date.today()

    projects = (await db.execute(
        select(Project.id, Project.name).where(Project.tenant_id == t).order_by(Project.name)
    )).all()
    proj_name = {pid: name for pid, name in projects}

    units = (await db.execute(
        select(Unit.id, Unit.project_id).where(Unit.tenant_id == t)
    )).all()

    cons = (await db.execute(
        select(UnitConstruction.unit_id, UnitConstruction.stage, UnitConstruction.percent,
               UnitConstruction.start_date, UnitConstruction.target_date, UnitConstruction.finish_date)
        .where(UnitConstruction.tenant_id == t)
    )).all()
    cmap = {r[0]: r for r in cons}

    log_rows = (await db.execute(
        select(ConstructionProgressLog.unit_id, func.max(ConstructionProgressLog.log_date))
        .where(ConstructionProgressLog.tenant_id == t, ConstructionProgressLog.is_deleted == False)  # noqa: E712
        .group_by(ConstructionProgressLog.unit_id)
    )).all()
    last_log = {uid: d for uid, d in log_rows}

    # akumulator per proyek
    agg: dict = {pid: {"units": 0, "pct_sum": 0, "done": 0, "in_progress": 0,
                       "not_started": 0, "overdue_target": 0, "late_update": 0}
                 for pid, _ in projects}
    stage_counts = {s.value: 0 for s in ConstructionStage}

    for uid, pid in units:
        a = agg.get(pid)
        if a is None:  # unit proyek yg tak ada di daftar (harusnya tak terjadi)
            continue
        c = cmap.get(uid)
        stage = c[1] if c else ConstructionStage.PERSIAPAN
        pct = c[2] if c else 0
        target = c[4] if c else None
        start = c[3] if c else None
        is_done = stage == ConstructionStage.SELESAI or pct >= 100

        a["units"] += 1
        a["pct_sum"] += pct
        stage_counts[stage.value] += 1
        if is_done:
            a["done"] += 1
        elif stage == ConstructionStage.PERSIAPAN and pct == 0:
            a["not_started"] += 1
        else:
            a["in_progress"] += 1

        if not is_done and target is not None and target < today:
            a["overdue_target"] += 1
        if not is_done:
            ref = last_log.get(uid) or start
            if ref is not None and (today - ref).days > CONSTRUCTION_REMINDER_DAYS:
                a["late_update"] += 1

    rows: list[ConstructionProject] = []
    for pid, _ in projects:
        a = agg[pid]
        n = a["units"]
        rows.append(ConstructionProject(
            project_id=pid, project_name=proj_name[pid],
            units_total=n, avg_percent=round(a["pct_sum"] / n, 1) if n else 0.0,
            done=a["done"], in_progress=a["in_progress"], not_started=a["not_started"],
            overdue_target=a["overdue_target"], late_update=a["late_update"],
        ))

    total_units = sum(r.units_total for r in rows)
    total_pct = sum(a["pct_sum"] for a in agg.values())
    return ConstructionProgressReport(
        projects=rows,
        units_total=total_units,
        done=sum(r.done for r in rows),
        overdue_target=sum(r.overdue_target for r in rows),
        late_update=sum(r.late_update for r in rows),
        avg_percent=round(total_pct / total_units, 1) if total_units else 0.0,
        stage_counts=stage_counts,
    )


# ═══════════════════════ LAPORAN: TUNGGAKAN / AGING PIUTANG ═══════════════════════
class AgingClient(BaseModel):
    client_id: uuid.UUID
    full_name: str
    project_name: Optional[str]
    unit_label: Optional[str]
    overdue_count: int       # jumlah termin telat
    outstanding: Decimal     # total tunggakan (nominal termin − sudah dibayar)
    max_days: int            # keterlambatan terlama (hari)
    bucket: str              # kategori umur berdasarkan max_days


class AgingReport(BaseModel):
    clients: list[AgingClient]
    total_outstanding: Decimal
    bucket_1_30: Decimal
    bucket_31_60: Decimal
    bucket_61_90: Decimal
    bucket_90p: Decimal
    overdue_clients: int
    overdue_schedules: int


def _aging_bucket(days: int) -> str:
    if days <= 30:
        return "1-30"
    if days <= 60:
        return "31-60"
    if days <= 90:
        return "61-90"
    return "90+"


@router.get("/aging", response_model=AgingReport)
async def aging(
    ctx: AuthContext = Depends(get_current_context),
    db: AsyncSession = Depends(get_db),
):
    """Tunggakan/aging piutang: termin (schedule) yang PENDING & lewat jatuh tempo, dikelompokkan umur & per pembeli."""
    t = ctx.tenant_id
    today = date.today()

    sched_rows = (await db.execute(
        select(PaymentSchedule.id, PaymentSchedule.client_id, PaymentSchedule.amount, PaymentSchedule.due_date)
        .where(PaymentSchedule.tenant_id == t, PaymentSchedule.is_deleted == False,  # noqa: E712
               PaymentSchedule.status == ScheduleStatus.PENDING,
               PaymentSchedule.client_id.isnot(None),
               PaymentSchedule.due_date.isnot(None), PaymentSchedule.due_date < today)
    )).all()

    # sudah dibayar per schedule (pembayaran yang dialokasikan ke termin ini, hanya yang disetujui)
    paid_rows = (await db.execute(
        select(Payment.schedule_id, func.coalesce(func.sum(Payment.amount), 0))
        .where(Payment.tenant_id == t, Payment.is_deleted == False,  # noqa: E712
               Payment.schedule_id.isnot(None), Payment.approval_status == PaymentApprovalStatus.APPROVED)
        .group_by(Payment.schedule_id)
    )).all()
    paid_by_sched = {sid: Decimal(v) for sid, v in paid_rows}

    # info pembeli + proyek + unit (untuk label)
    cli_rows = (await db.execute(
        select(Client.id, Client.full_name, Project.name, Unit.block, Unit.unit_number)
        .select_from(Client)
        .outerjoin(Project, Project.id == Client.project_id)
        .outerjoin(Unit, Unit.id == Client.unit_id)
        .where(Client.tenant_id == t)
    )).all()
    cli_info = {
        cid: (name, proj, (f"{blk} " if blk else "") + (unum or "") if unum or blk else None)
        for cid, name, proj, blk, unum in cli_rows
    }

    per_client: dict = {}
    bucket_totals = {"1-30": Decimal(0), "31-60": Decimal(0), "61-90": Decimal(0), "90+": Decimal(0)}
    for sid, cid, amount, due in sched_rows:
        outstanding = Decimal(amount or 0) - paid_by_sched.get(sid, Decimal(0))
        if outstanding <= 0:
            continue
        days = (today - due).days
        bucket_totals[_aging_bucket(days)] += outstanding
        c = per_client.setdefault(cid, {"outstanding": Decimal(0), "count": 0, "max_days": 0})
        c["outstanding"] += outstanding
        c["count"] += 1
        c["max_days"] = max(c["max_days"], days)

    clients: list[AgingClient] = []
    for cid, c in per_client.items():
        name, proj, unit_label = cli_info.get(cid, (None, None, None))
        clients.append(AgingClient(
            client_id=cid, full_name=name or "—", project_name=proj, unit_label=unit_label,
            overdue_count=c["count"], outstanding=c["outstanding"], max_days=c["max_days"],
            bucket=_aging_bucket(c["max_days"]),
        ))
    # yang paling parah dulu (nominal terbesar), lalu terlama
    clients.sort(key=lambda x: (-x.outstanding, -x.max_days))

    return AgingReport(
        clients=clients,
        total_outstanding=sum((c.outstanding for c in clients), Decimal(0)),
        bucket_1_30=bucket_totals["1-30"], bucket_31_60=bucket_totals["31-60"],
        bucket_61_90=bucket_totals["61-90"], bucket_90p=bucket_totals["90+"],
        overdue_clients=len(clients),
        overdue_schedules=sum(c.overdue_count for c in clients),
    )


# ═══════════════════════ GRAFIK PENJUALAN PER BULAN ═══════════════════════
class SalesMonthly(BaseModel):
    month: str        # "YYYY-MM"
    count: int        # jumlah unit terjual (pembeli) bulan itu
    value: Decimal    # nilai penjualan (Σ harga jual)


@router.get("/sales-monthly", response_model=list[SalesMonthly])
async def sales_monthly(
    project_id: Optional[uuid.UUID] = Query(None),
    year: Optional[int] = Query(None, ge=2000, le=2100),
    ctx: AuthContext = Depends(get_current_context),
    db: AsyncSession = Depends(get_db),
):
    """Penjualan per bulan (unit terjual + nilai) berdasarkan tanggal kontrak pembeli
    (fallback tanggal entri). Opsional filter per proyek & tahun.
    Tanpa `year` → 12 bulan terakhir yang ada penjualan (rolling). Dengan `year` → Jan-Des penuh
    tahun itu (termasuk bulan tanpa penjualan), supaya sumbu grafik konsisten satu tahun."""
    t = ctx.tenant_id
    date_col = func.coalesce(Client.contract_date, func.date(Client.created_at))
    ym = func.to_char(date_col, "YYYY-MM")
    conds = [Client.tenant_id == t, Client.is_deleted == False,  # noqa: E712
             Client.status != ClientStatus.INACTIVE]
    if project_id:
        conds.append(Client.project_id == project_id)
    if year:
        conds.append(func.extract("year", date_col) == year)
    rows = (await db.execute(
        select(ym.label("ym"), func.count(), func.coalesce(func.sum(Client.contract_value), 0))
        .where(*conds).group_by(ym).order_by(ym)
    )).all()
    by_month = {m: SalesMonthly(month=m, count=c, value=Decimal(v)) for m, c, v in rows}
    if year:
        return [by_month.get(f"{year}-{mo:02d}") or SalesMonthly(month=f"{year}-{mo:02d}", count=0, value=Decimal(0)) for mo in range(1, 13)]
    return list(by_month.values())[-12:]


# ═══════════════════════ LAPORAN: PAJAK BULANAN (PPh) ═══════════════════════
SHM_RE = re.compile(r'shm|hgb|sertifikat', re.I)
PBB_RE = re.compile(r'pbb', re.I)
SIKASEP_RE = re.compile(r'sikasep|sikumbang', re.I)


class MonthlyTaxRow(BaseModel):
    client_id: uuid.UUID
    name: str
    nik: Optional[str] = None
    location: Optional[str] = None       # nama proyek
    unit_number: Optional[str] = None    # blok-nomor
    category: Optional[str] = None       # subsidi | komersial
    base_amount: Optional[Decimal] = None  # Nilai AJB
    amount: Optional[Decimal] = None       # Jumlah PPh
    ppn_amount: Optional[Decimal] = None   # Jumlah PPN (dari TaxRecord PPN klien ini, bila ada)
    bphtb_amount: Optional[Decimal] = None  # Jumlah BPHTB (dari TaxRecord BPHTB klien ini, bila ada)
    ntpn: Optional[str] = None
    shm_number: Optional[str] = None     # dari Dokumen Legalitas unit (SHM/HGB)
    pbb_number: Optional[str] = None     # dari Dokumen Legalitas unit (PBB)
    sikumbang_number: Optional[str] = None  # KIR — No. SiKasep/SiKumbang, dari KPR pembeli (kosong utk cash)
    notary_name: Optional[str] = None
    tax_date: Optional[date] = None


class MonthlyTaxReport(BaseModel):
    month: str
    rows: list[MonthlyTaxRow]
    total_count: int
    total_base_amount: Decimal
    total_amount: Decimal
    total_ppn_amount: Decimal
    total_bphtb_amount: Decimal


async def _build_monthly_tax_report(db: AsyncSession, t: uuid.UUID, month: str, project_id: Optional[uuid.UUID]) -> MonthlyTaxReport:
    """Rekap PPh bulanan per pembeli (nama/NIK/lokasi/kategori/AJB/jumlah/NTPN/No. SiKumbang/notaris)
    — utk lapor ke akuntan/kantor pajak. Hanya baris PPh yang SUDAH ada tanggalnya (tax_date) di bulan
    terpilih; baris belum bayar (tanpa tanggal) sengaja tak ikut. Dipakai endpoint biasa & tautan publik."""
    try:
        year, mon = int(month[:4]), int(month[5:7])
    except (ValueError, IndexError):
        year = mon = 0
    if not (1 <= mon <= 12):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Format bulan harus YYYY-MM")
    start = date(year, mon, 1)
    end = date(year + 1, 1, 1) if mon == 12 else date(year, mon + 1, 1)

    conds = [
        TaxRecord.tenant_id == t, TaxRecord.is_deleted == False,  # noqa: E712
        TaxRecord.tax_type == TaxType.PPH,
        TaxRecord.tax_date.isnot(None), TaxRecord.tax_date >= start, TaxRecord.tax_date < end,
    ]
    rows = (await db.execute(
        select(TaxRecord).options(selectinload(TaxRecord.notary)).where(*conds).order_by(TaxRecord.tax_date)
    )).scalars().all()

    client_ids = {r.client_id for r in rows}
    clients: dict = {}
    if client_ids:
        cconds = [Client.id.in_(client_ids)]
        if project_id:
            cconds.append(Client.project_id == project_id)
        for c in (await db.execute(select(Client).where(*cconds))).scalars().all():
            clients[c.id] = c

    unit_ids = {c.unit_id for c in clients.values() if c.unit_id}
    units: dict = {}
    if unit_ids:
        for u in (await db.execute(select(Unit).where(Unit.id.in_(unit_ids)))).scalars().all():
            units[u.id] = u

    proj_ids = {c.project_id for c in clients.values() if c.project_id}
    proj_names: dict = {}
    if proj_ids:
        for pid, pname in (await db.execute(select(Project.id, Project.name).where(Project.id.in_(proj_ids)))).all():
            proj_names[pid] = pname

    # Jumlah PPN & BPHTB per klien (tak terikat bulan — cukup nilai klien ybs, kalau ada)
    ppn_by_client: dict = {}
    bphtb_by_client: dict = {}
    if client_ids:
        ppn_rows = (await db.execute(
            select(TaxRecord.client_id, TaxRecord.amount)
            .where(TaxRecord.client_id.in_(client_ids), TaxRecord.tenant_id == t,
                   TaxRecord.is_deleted == False, TaxRecord.tax_type == TaxType.PPN)  # noqa: E712
        )).all()
        ppn_by_client = {cid: amt for cid, amt in ppn_rows}
        bphtb_rows = (await db.execute(
            select(TaxRecord.client_id, TaxRecord.amount)
            .where(TaxRecord.client_id.in_(client_ids), TaxRecord.tenant_id == t,
                   TaxRecord.is_deleted == False, TaxRecord.tax_type == TaxType.BPHTB)  # noqa: E712
        )).all()
        bphtb_by_client = {cid: amt for cid, amt in bphtb_rows}

    # No. SHM & No. PBB & No. SiKasep/SiKumbang — dari Dokumen Legalitas unit (doc_type teks bebas, dicocokkan pola sama FE)
    shm_by_unit: dict = {}
    pbb_by_unit: dict = {}
    sikumbang_by_unit: dict = {}
    if unit_ids:
        doc_rows = (await db.execute(
            select(Document.unit_id, Document.doc_type, Document.name)
            .where(Document.unit_id.in_(unit_ids), Document.tenant_id == t, Document.is_deleted == False)  # noqa: E712
        )).all()
        for uid, doc_type, dname in doc_rows:
            if SHM_RE.search(doc_type or '') and uid not in shm_by_unit:
                shm_by_unit[uid] = dname
            elif PBB_RE.search(doc_type or '') and uid not in pbb_by_unit:
                pbb_by_unit[uid] = dname
            elif SIKASEP_RE.search(doc_type or '') and uid not in sikumbang_by_unit:
                sikumbang_by_unit[uid] = dname

    result_rows: list[MonthlyTaxRow] = []
    for r in rows:
        c = clients.get(r.client_id)
        if c is None:   # tersaring project_id, atau klien sudah dihapus
            continue
        u = units.get(c.unit_id) if c.unit_id else None
        result_rows.append(MonthlyTaxRow(
            client_id=r.client_id, name=c.full_name, nik=c.nik,
            location=proj_names.get(c.project_id),
            unit_number=("-".join(x for x in [u.block, u.unit_number] if x) if u else None),
            category=r.category, base_amount=r.base_amount, amount=r.amount,
            ppn_amount=ppn_by_client.get(r.client_id), bphtb_amount=bphtb_by_client.get(r.client_id), ntpn=r.ntpn,
            shm_number=shm_by_unit.get(c.unit_id) if c.unit_id else None,
            pbb_number=pbb_by_unit.get(c.unit_id) if c.unit_id else None,
            sikumbang_number=sikumbang_by_unit.get(c.unit_id) if c.unit_id else None,
            notary_name=r.notary.name if r.notary else None, tax_date=r.tax_date,
        ))

    return MonthlyTaxReport(
        month=month, rows=result_rows, total_count=len(result_rows),
        total_base_amount=sum((x.base_amount or Decimal(0) for x in result_rows), Decimal(0)),
        total_amount=sum((x.amount or Decimal(0) for x in result_rows), Decimal(0)),
        total_ppn_amount=sum((x.ppn_amount or Decimal(0) for x in result_rows), Decimal(0)),
        total_bphtb_amount=sum((x.bphtb_amount or Decimal(0) for x in result_rows), Decimal(0)),
    )


@router.get("/monthly-tax", response_model=MonthlyTaxReport)
async def monthly_tax(
    month: str = Query(..., description="YYYY-MM"),
    project_id: Optional[uuid.UUID] = Query(None),
    ctx: AuthContext = Depends(get_current_context),
    db: AsyncSession = Depends(get_db),
):
    return await _build_monthly_tax_report(db, ctx.tenant_id, month, project_id)


# ═══════════════════════ CHECKLIST PAJAK BELUM DIURUS ═══════════════════════
# "Selesai" = ada TaxRecord dgn status validasi/dtp/bebas (terminal). "Belum" = tak ada baris
# sama sekali, ATAU ada baris tapi status masih belum/dibayar (masih berjalan, belum tuntas).
TAX_COMPLETE_STATUSES = (TaxStatus.VALIDASI, TaxStatus.DTP, TaxStatus.BEBAS)


class TaxChecklistItem(BaseModel):
    has_record: bool
    status: str   # nilai TaxStatus, atau 'belum_ada' bila tak ada baris sama sekali
    is_complete: bool


class TaxChecklistRow(BaseModel):
    client_id: uuid.UUID
    full_name: str
    unit_label: Optional[str] = None
    project_name: Optional[str] = None
    contract_date: Optional[date] = None
    days_since_contract: Optional[int] = None
    pph: TaxChecklistItem
    bphtb: TaxChecklistItem
    ppn: TaxChecklistItem
    incomplete_count: int   # 0-3, dari pph/bphtb/ppn yang belum tuntas


class TaxChecklistReport(BaseModel):
    rows: list[TaxChecklistRow]
    total_clients: int
    total_incomplete_clients: int


def _tax_item(rec: Optional[TaxRecord]) -> TaxChecklistItem:
    if rec is None:
        return TaxChecklistItem(has_record=False, status="belum_ada", is_complete=False)
    return TaxChecklistItem(has_record=True, status=rec.status.value, is_complete=rec.status in TAX_COMPLETE_STATUSES)


@router.get("/tax-checklist", response_model=TaxChecklistReport)
async def tax_checklist(
    project_id: Optional[uuid.UUID] = Query(None),
    only_incomplete: bool = Query(True, description="Hanya tampilkan pembeli dgn minimal 1 jenis pajak belum tuntas"),
    ctx: AuthContext = Depends(get_current_context),
    db: AsyncSession = Depends(get_db),
):
    """Checklist per-pembeli: status PPh/BPHTB/PPN mana yang belum diurus (tak ada baris,
    atau masih belum/dibayar — belum divalidasi/DTP/bebas). Pembeli batal (INACTIVE) dikecualikan."""
    t = ctx.tenant_id
    cconds = [Client.tenant_id == t, Client.is_deleted == False, Client.status != ClientStatus.INACTIVE]  # noqa: E712
    if project_id:
        cconds.append(Client.project_id == project_id)
    clients = (await db.execute(select(Client).where(*cconds).order_by(Client.contract_date))).scalars().all()
    if not clients:
        return TaxChecklistReport(rows=[], total_clients=0, total_incomplete_clients=0)

    client_ids = [c.id for c in clients]
    tax_rows = (await db.execute(
        select(TaxRecord).where(TaxRecord.tenant_id == t, TaxRecord.client_id.in_(client_ids),
                                TaxRecord.is_deleted == False)  # noqa: E712
    )).scalars().all()
    by_key: dict = {(r.client_id, r.tax_type): r for r in tax_rows}

    unit_ids = {c.unit_id for c in clients if c.unit_id}
    units = {u.id: u for u in (await db.execute(select(Unit).where(Unit.id.in_(unit_ids)))).scalars().all()} if unit_ids else {}
    proj_ids = {c.project_id for c in clients if c.project_id}
    proj_names = dict((await db.execute(select(Project.id, Project.name).where(Project.id.in_(proj_ids)))).all()) if proj_ids else {}

    today = date.today()
    rows: list[TaxChecklistRow] = []
    for c in clients:
        pph = _tax_item(by_key.get((c.id, TaxType.PPH)))
        bphtb = _tax_item(by_key.get((c.id, TaxType.BPHTB)))
        ppn = _tax_item(by_key.get((c.id, TaxType.PPN)))
        incomplete = sum(1 for it in (pph, bphtb, ppn) if not it.is_complete)
        if only_incomplete and incomplete == 0:
            continue
        u = units.get(c.unit_id) if c.unit_id else None
        rows.append(TaxChecklistRow(
            client_id=c.id, full_name=c.full_name,
            unit_label=("-".join(x for x in [u.block, u.unit_number] if x) if u else None),
            project_name=proj_names.get(c.project_id),
            contract_date=c.contract_date,
            days_since_contract=(today - c.contract_date).days if c.contract_date else None,
            pph=pph, bphtb=bphtb, ppn=ppn, incomplete_count=incomplete,
        ))
    rows.sort(key=lambda r: (-(r.days_since_contract or 0), -r.incomplete_count))

    return TaxChecklistReport(
        rows=rows, total_clients=len(clients),
        total_incomplete_clients=sum(1 for c in clients
            if sum(1 for tt in (TaxType.PPH, TaxType.BPHTB, TaxType.PPN)
                   if not _tax_item(by_key.get((c.id, tt))).is_complete) > 0),
    )


# ── Tautan bagikan Laporan Pajak Bulanan ke pihak luar (mis. konsultan pajak), tanpa login ──
class ShareLinkCreate(BaseModel):
    month: str
    project_id: Optional[uuid.UUID] = None
    expires_days: int = 30


class ShareLinkResponse(BaseModel):
    id: uuid.UUID
    token: str
    month: str
    project_id: Optional[uuid.UUID] = None
    project_name: Optional[str] = None
    expires_at: datetime
    revoked_at: Optional[datetime] = None
    last_accessed_at: Optional[datetime] = None
    access_count: int
    is_active: bool
    created_at: datetime

    class Config:
        from_attributes = True


@router.get("/monthly-tax/share", response_model=list[ShareLinkResponse])
async def list_monthly_tax_share_links(ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    """Daftar tautan bagikan yang pernah dibuat tenant ini (termasuk yang sudah expired/dicabut, utk histori)."""
    rows = (await db.execute(
        select(MonthlyTaxShareLink).where(MonthlyTaxShareLink.tenant_id == ctx.tenant_id)
        .order_by(MonthlyTaxShareLink.created_at.desc())
    )).scalars().all()
    return rows


@router.post("/monthly-tax/share", response_model=ShareLinkResponse, status_code=status.HTTP_201_CREATED)
async def create_monthly_tax_share_link(payload: ShareLinkCreate, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    """Buat tautan bertoken (tanpa login) utk bagikan Laporan Pajak Bulanan satu bulan ke pihak luar."""
    proj_name = None
    if payload.project_id:
        proj_name = await db.scalar(select(Project.name).where(Project.id == payload.project_id, Project.tenant_id == ctx.tenant_id))
        if proj_name is None:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, detail="Proyek tidak ditemukan")
    days = max(1, min(365, payload.expires_days))
    link = MonthlyTaxShareLink(
        tenant_id=ctx.tenant_id, token=secrets.token_urlsafe(32), month=payload.month,
        project_id=payload.project_id, project_name_snapshot=proj_name or "Semua Proyek",
        created_by=ctx.user_id, expires_at=datetime.now(timezone.utc) + timedelta(days=days),
    )
    db.add(link)
    await db.flush()
    await db.refresh(link)
    return link


@router.delete("/monthly-tax/share/{link_id}", status_code=status.HTTP_204_NO_CONTENT)
async def revoke_monthly_tax_share_link(link_id: uuid.UUID, ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    """Cabut tautan bagikan — begitu dicabut, tautan tak bisa diakses lagi (walau belum expired)."""
    link = (await db.execute(
        select(MonthlyTaxShareLink).where(MonthlyTaxShareLink.id == link_id, MonthlyTaxShareLink.tenant_id == ctx.tenant_id)
    )).scalar_one_or_none()
    if link is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Tautan tidak ditemukan")
    link.revoked_at = datetime.now(timezone.utc)
    await db.flush()


# ═══════════════════════ LABA / RUGI PER PROYEK ═══════════════════════
# Bukan akuntansi formal (tak ada jurnal/neraca) — laporan OPERASIONAL: berapa
# untung tiap proyek & unit. Fase B2 double-entry sengaja tidak dibangun.
#
# ★ DUA KEPUTUSAN YANG MENENTUKAN ANGKANYA:
# 1. Basis ACCRUAL — biaya dihitung saat TERJADI, bukan saat dibayar. Ini menyamakan
#    angkanya dengan RAB & Kebocoran (rab.py::_realisasi_map); kalau dipakai basis kas,
#    dua laporan akan berbeda dan pengguna kehilangan kepercayaan pada keduanya.
# 2. Biaya unit BELUM TERJUAL = PERSEDIAAN (modal tertanam), TIDAK dikurangkan dari laba.
#    Kalau dikurangkan, proyek yang sedang membangun banyak unit selalu tampak rugi —
#    padahal uangnya berubah jadi rumah, bukan hilang.
#
# Logika biaya SENGAJA menyalin rab.py::_realisasi_map (transfer & retur-vendor bukan
# biaya; retur-unit mengurangi biaya unit asal). Kalau aturan itu berubah di sana,
# ubah di sini juga — silang-cek keduanya lewat total realisasi per unit.

_PROFIT_COST_GROUP = {
    ExpenseCategory.MATERIAL: "material",
    ExpenseCategory.UPAH: "upah", ExpenseCategory.KONTRAKTOR: "upah",
    ExpenseCategory.KELISTRIKAN: "utilitas", ExpenseCategory.AIR_PDAM: "utilitas",
    ExpenseCategory.OPERASIONAL: "lain", ExpenseCategory.PERIZINAN: "lain",
    ExpenseCategory.LAIN: "lain",
}


class ProjectProfitRow(BaseModel):
    project_id: uuid.UUID
    project_name: str
    units_total: int
    units_sold: int                 # unit yang sudah punya pembeli aktif
    revenue_contract: Decimal       # Σ nilai kontrak pembeli proyek ini
    revenue_cash: Decimal           # kas masuk yang sudah disetujui keuangan
    cost_sold: Decimal              # biaya unit yang sudah terjual
    cost_general: Decimal           # biaya umum proyek (tak melekat ke unit manapun)
    cost_notary: Decimal            # biaya notaris pembeli proyek ini
    profit: Decimal                 # revenue_contract − (cost_sold + cost_general + cost_notary)
    margin_pct: Optional[Decimal] = None
    inventory_value: Decimal        # biaya unit BELUM terjual — modal tertanam, bukan biaya
    clients_without_unit: int       # pembeli tanpa unit_id → pendapatannya tak punya lawan biaya


class ProjectProfitReport(BaseModel):
    rows: list[ProjectProfitRow]
    revenue_contract: Decimal
    revenue_cash: Decimal
    cost_total: Decimal
    profit: Decimal
    inventory_value: Decimal


class UnitProfitRow(BaseModel):
    unit_id: uuid.UUID
    unit_label: str
    unit_status: str
    client_name: Optional[str] = None
    contract_value: Optional[Decimal] = None
    cost_material: Decimal
    cost_upah: Decimal
    cost_utilitas: Decimal
    cost_lain: Decimal
    cost_total: Decimal
    profit: Optional[Decimal] = None     # None bila belum terjual (tak ada pendapatan)
    margin_pct: Optional[Decimal] = None
    is_sold: bool


class ProjectProfitDetail(BaseModel):
    project_id: uuid.UUID
    project_name: str
    rows: list[UnitProfitRow]
    cost_general: Decimal
    cost_notary: Decimal
    revenue_unattributed: Decimal   # kontrak pembeli yang unit_id-nya kosong


async def _project_cost_map(db, tenant_id, project_id):
    """dict[unit_id | None][grup] = Decimal. Grup: material/upah/utilitas/lain.
    Salinan aturan rab.py::_realisasi_map — accrual, transfer & retur-vendor bukan biaya."""
    from app.models.stock import StockMovement, MovementType, MovementSource
    res: dict = {}

    def bucket(uid):
        return res.setdefault(uid, {"material": Decimal(0), "upah": Decimal(0),
                                    "utilitas": Decimal(0), "lain": Decimal(0)})

    movs = (await db.execute(select(StockMovement).where(
        StockMovement.tenant_id == tenant_id, StockMovement.project_id == project_id,
        StockMovement.is_deleted == False))).scalars().all()  # noqa: E712
    NOT_A_COST = (MovementSource.RETURN_VENDOR, MovementSource.TRANSFER_OUT)
    for m in movs:
        nilai = Decimal(m.quantity) * Decimal(m.unit_price)
        if m.movement_type == MovementType.OUT and m.source not in NOT_A_COST:
            bucket(m.unit_id)["material"] += nilai
        elif m.movement_type == MovementType.IN and m.source == MovementSource.RETURN_UNIT:
            bucket(m.unit_id)["material"] -= nilai   # retur dari unit → kurangi biaya unit asal

    exps = (await db.execute(select(Expense).where(
        Expense.tenant_id == tenant_id, Expense.project_id == project_id,
        Expense.is_deleted == False))).scalars().all()  # noqa: E712
    for e in exps:
        bucket(e.unit_id)[_PROFIT_COST_GROUP.get(e.category, "lain")] += Decimal(e.amount)
    return res


def _unit_label(u) -> str:
    return "-".join(x for x in [u.block, u.unit_number] if x) or "?"


def _margin_pct(profit: Decimal, revenue: Decimal) -> Optional[Decimal]:
    if not revenue or revenue == 0:
        return None
    return (profit / revenue * 100).quantize(Decimal("0.1"))


async def _clients_of_project(db, tenant_id, project_id=None):
    """Pembeli aktif (bukan batal/terhapus) + nilai kontraknya."""
    conds = [Client.tenant_id == tenant_id, Client.is_deleted == False,  # noqa: E712
             Client.status != ClientStatus.INACTIVE]
    if project_id is not None:
        conds.append(Client.project_id == project_id)
    return (await db.execute(
        select(Client.id, Client.project_id, Client.unit_id,
               Client.contract_value, Client.full_name).where(*conds)
    )).all()


@router.get("/project-profit", response_model=ProjectProfitReport)
async def project_profit(
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    """Laba/rugi per proyek. Pendapatan = nilai kontrak pembeli aktif; biaya = accrual
    (material terdistribusi + biaya + opname + utilitas) + biaya notaris."""
    t = ctx.tenant_id
    projects = (await db.execute(
        select(Project.id, Project.name).where(Project.tenant_id == t).order_by(Project.name)
    )).all()

    units = (await db.execute(select(Unit).where(Unit.tenant_id == t))).scalars().all()
    units_by_proj: dict = {}
    for u in units:
        units_by_proj.setdefault(u.project_id, []).append(u)

    clients = await _clients_of_project(db, t)
    sold_unit_ids = {c.unit_id for c in clients if c.unit_id is not None}

    pay_rows = (await db.execute(
        select(Payment.client_id, func.coalesce(func.sum(Payment.amount), 0))
        .where(Payment.tenant_id == t, Payment.is_deleted == False,  # noqa: E712
               Payment.approval_status == PaymentApprovalStatus.APPROVED)
        .group_by(Payment.client_id)
    )).all()
    paid_by_client = {cid: Decimal(v) for cid, v in pay_rows}

    fee_rows = (await db.execute(
        select(NotaryFee.client_id, func.coalesce(func.sum(NotaryFee.amount), 0))
        .where(NotaryFee.tenant_id == t, NotaryFee.is_deleted == False)  # noqa: E712
        .group_by(NotaryFee.client_id)
    )).all()
    fee_by_client = {cid: Decimal(v) for cid, v in fee_rows}

    rows = []
    for pid, pname in projects:
        pclients = [c for c in clients if c.project_id == pid]
        revenue = sum((Decimal(c.contract_value or 0) for c in pclients), Decimal(0))
        cash = sum((paid_by_client.get(c.id, Decimal(0)) for c in pclients), Decimal(0))
        notary = sum((fee_by_client.get(c.id, Decimal(0)) for c in pclients), Decimal(0))

        cost_map = await _project_cost_map(db, t, pid)
        cost_sold = Decimal(0)
        inventory = Decimal(0)
        for uid, grp in cost_map.items():
            total = sum(grp.values(), Decimal(0))
            if uid is None:
                continue                       # biaya umum ditangani terpisah
            if uid in sold_unit_ids:
                cost_sold += total
            else:
                inventory += total             # ★ persediaan, BUKAN pengurang laba
        general = sum(cost_map.get(None, {}).values(), Decimal(0))

        profit = revenue - (cost_sold + general + notary)
        punits = units_by_proj.get(pid, [])
        rows.append(ProjectProfitRow(
            project_id=pid, project_name=pname,
            units_total=len(punits),
            units_sold=len([u for u in punits if u.id in sold_unit_ids]),
            revenue_contract=revenue, revenue_cash=cash,
            cost_sold=cost_sold, cost_general=general, cost_notary=notary,
            profit=profit, margin_pct=_margin_pct(profit, revenue),
            inventory_value=inventory,
            clients_without_unit=len([c for c in pclients if c.unit_id is None]),
        ))

    return ProjectProfitReport(
        rows=rows,
        revenue_contract=sum((r.revenue_contract for r in rows), Decimal(0)),
        revenue_cash=sum((r.revenue_cash for r in rows), Decimal(0)),
        cost_total=sum((r.cost_sold + r.cost_general + r.cost_notary for r in rows), Decimal(0)),
        profit=sum((r.profit for r in rows), Decimal(0)),
        inventory_value=sum((r.inventory_value for r in rows), Decimal(0)),
    )


class BizOpexCat(BaseModel):
    name: str
    total: Decimal


class BusinessPnL(BaseModel):
    year: int
    pendapatan: Decimal          # nilai kontrak unit terjual (akad) tahun ini
    units_sold: int
    hpp_unit: Decimal            # biaya bangun unit terjual + alokasi biaya umum proyek
    hpp_notaris: Decimal
    hpp_total: Decimal
    laba_kotor: Decimal
    biaya_operasional: Decimal
    opex_by_category: list[BizOpexCat]
    laba_usaha: Decimal
    margin_pct: Optional[Decimal] = None


@router.get("/business-pnl", response_model=BusinessPnL)
async def business_pnl(year: int = Query(...), month: int = Query(None, ge=1, le=12),
                       ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    """Laba/Rugi Usaha per tahun (manajerial, accrual matching-at-sale). Pendapatan = kontrak unit
    yang AKAD tahun ini; HPP = biaya bangun unit tsb + alokasi biaya umum proyek + notaris; dikurangi
    Biaya Operasional (opex) tahun ini. Unit belum terjual tetap persediaan (tak jadi beban)."""
    from app.models.opex import OperationalExpense
    from datetime import date as _d
    from calendar import monthrange
    t = ctx.tenant_id
    if month:
        p_start, p_end = _d(year, month, 1), _d(year, month, monthrange(year, month)[1])
    else:
        p_start, p_end = _d(year, 1, 1), _d(year, 12, 31)
    yclients = (await db.execute(select(Client.id, Client.project_id, Client.unit_id, Client.contract_value).where(
        Client.tenant_id == t, Client.is_deleted == False, Client.status != ClientStatus.INACTIVE,  # noqa: E712
        Client.contract_date >= p_start, Client.contract_date <= p_end))).all()
    revenue = sum((Decimal(c.contract_value or 0) for c in yclients), Decimal(0))

    fee_rows = (await db.execute(
        select(NotaryFee.client_id, func.coalesce(func.sum(NotaryFee.amount), 0))
        .where(NotaryFee.tenant_id == t, NotaryFee.is_deleted == False).group_by(NotaryFee.client_id))).all()  # noqa: E712
    fee_by_client = {cid: Decimal(v) for cid, v in fee_rows}

    # biaya per proyek (hanya proyek yg ada penjualan tahun ini) + alokasi biaya umum per unit
    proj_ids = {c.project_id for c in yclients if c.project_id}
    units = (await db.execute(select(Unit).where(Unit.tenant_id == t))).scalars().all()
    units_per_proj: dict = {}
    for u in units:
        units_per_proj[u.project_id] = units_per_proj.get(u.project_id, 0) + 1
    hpp_unit = Decimal(0)
    for pid in proj_ids:
        cmap = await _project_cost_map(db, t, pid)
        general = sum(cmap.get(None, {}).values(), Decimal(0))
        gen_per_unit = general / Decimal(max(units_per_proj.get(pid, 0), 1))
        for c in yclients:
            if c.project_id == pid and c.unit_id is not None:
                direct = sum(cmap.get(c.unit_id, {}).values(), Decimal(0))
                hpp_unit += direct + gen_per_unit
    notary = sum((fee_by_client.get(c.id, Decimal(0)) for c in yclients), Decimal(0))
    hpp_total = hpp_unit + notary
    laba_kotor = revenue - hpp_total

    # biaya operasional tahun ini (dibayar)
    orows = (await db.execute(select(OperationalExpense).options(selectinload(OperationalExpense.category)).where(
        OperationalExpense.tenant_id == t, OperationalExpense.is_deleted == False,  # noqa: E712
        OperationalExpense.is_paid == True,  # noqa: E712
        OperationalExpense.expense_date >= p_start, OperationalExpense.expense_date <= p_end))).scalars().all()
    opex_total = sum((Decimal(o.amount) for o in orows), Decimal(0))
    agg: dict = {}
    for o in orows:
        k = o.category.name if o.category else "Tanpa kategori"
        agg[k] = agg.get(k, Decimal(0)) + Decimal(o.amount)
    opex_by_cat = [BizOpexCat(name=k, total=v) for k, v in sorted(agg.items(), key=lambda x: -x[1])]

    laba_usaha = laba_kotor - opex_total
    return BusinessPnL(
        year=year, pendapatan=revenue, units_sold=len([c for c in yclients if c.unit_id]),
        hpp_unit=hpp_unit, hpp_notaris=notary, hpp_total=hpp_total, laba_kotor=laba_kotor,
        biaya_operasional=opex_total, opex_by_category=opex_by_cat, laba_usaha=laba_usaha,
        margin_pct=_margin_pct(laba_usaha, revenue),
    )


class FinancialPosition(BaseModel):
    # Harta / Aset
    kas_bank: Decimal
    persediaan: Decimal            # modal tertanam: biaya unit belum terjual
    piutang_pembeli: Decimal
    retensi_bank: Decimal
    total_aset: Decimal
    # Kewajiban
    biaya_belum_dibayar: Decimal   # biaya proyek menunggu bayar
    hutang_notaris: Decimal
    opex_belum_dibayar: Decimal
    total_kewajiban: Decimal
    # Bersih
    kekayaan_bersih: Decimal


@router.get("/financial-position", response_model=FinancialPosition)
async def financial_position(ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db)):
    """Posisi Keuangan Ringkas (snapshot manajerial, bukan neraca formal). Menggabungkan data yg sudah ada:
    Kas+Bank, Persediaan (modal tertanam unit belum terjual), Piutang Pembeli, Retensi Bank − Kewajiban."""
    from app.models.opex import OperationalExpense
    t = ctx.tenant_id

    # ── Kas & Bank: Σ saldo awal + entri masuk − entri keluar (transfer net nol) ──
    opening = Decimal(await db.scalar(select(func.coalesce(func.sum(CashAccount.opening_balance), 0)).where(
        CashAccount.tenant_id == t, CashAccount.is_deleted == False)) or 0)  # noqa: E712
    ein = Decimal(await db.scalar(select(func.coalesce(func.sum(CashBookEntry.amount), 0)).where(
        CashBookEntry.tenant_id == t, CashBookEntry.direction == CashDirection.IN)) or 0)
    eout = Decimal(await db.scalar(select(func.coalesce(func.sum(CashBookEntry.amount), 0)).where(
        CashBookEntry.tenant_id == t, CashBookEntry.direction == CashDirection.OUT)) or 0)
    kas_bank = opening + ein - eout

    # ── Persediaan (biaya unit belum terjual) — sama aturan dgn Laba/Rugi Proyek ──
    clients_full = await _clients_of_project(db, t)
    sold_unit_ids = {c.unit_id for c in clients_full if c.unit_id is not None}
    proj_ids = [pid for (pid,) in (await db.execute(select(Project.id).where(Project.tenant_id == t))).all()]
    persediaan = Decimal(0)
    for pid in proj_ids:
        cmap = await _project_cost_map(db, t, pid)
        for uid, grp in cmap.items():
            if uid is not None and uid not in sold_unit_ids:
                persediaan += sum(grp.values(), Decimal(0))

    # ── Piutang Pembeli & Retensi Bank (pisah, tak dobel) — pola finance_summary ──
    notdel_p = (Payment.is_deleted == False) & (Payment.approval_status == PaymentApprovalStatus.APPROVED)  # noqa: E712
    clients = (await db.execute(select(Client.id, Client.contract_value).where(
        Client.tenant_id == t, Client.is_deleted == False, Client.status != ClientStatus.INACTIVE))).all()  # noqa: E712
    buyer_by = {cid: Decimal(v) for cid, v in (await db.execute(
        select(Payment.client_id, func.coalesce(func.sum(Payment.amount), 0))
        .where(Payment.tenant_id == t, Payment.source == PaymentSource.PEMBELI, notdel_p).group_by(Payment.client_id))).all()}
    bank_by = {cid: Decimal(v) for cid, v in (await db.execute(
        select(Payment.client_id, func.coalesce(func.sum(Payment.amount), 0))
        .where(Payment.tenant_id == t, Payment.source == PaymentSource.BANK, notdel_p).group_by(Payment.client_id))).all()}
    committed_by: dict = {}
    for cid, plaf, stage in (await db.execute(select(KprApplication.client_id, KprApplication.plafond, KprApplication.stage)
                             .where(KprApplication.tenant_id == t, KprApplication.is_deleted == False)  # noqa: E712
                             .order_by(KprApplication.client_id, KprApplication.created_at.desc()))).all():
        if cid not in committed_by:
            committed_by[cid] = Decimal(plaf or 0) if stage in (KprStage.AKAD_KREDIT, KprStage.PENCAIRAN) else Decimal(0)
    piutang = Decimal(0); retensi = Decimal(0)
    for cid, price in clients:
        price = Decimal(price or 0); committed = committed_by.get(cid, Decimal(0))
        piutang += max(price - buyer_by.get(cid, Decimal(0)) - committed, Decimal(0))
        if committed > 0:
            retensi += max(committed - bank_by.get(cid, Decimal(0)), Decimal(0))

    # ── Kewajiban ──
    biaya_bd = Decimal(await db.scalar(select(func.coalesce(func.sum(Expense.amount), 0)).where(
        Expense.tenant_id == t, Expense.is_deleted == False, Expense.is_paid == False)) or 0)  # noqa: E712
    notaris_bd = Decimal(await db.scalar(select(func.coalesce(func.sum(NotaryFee.amount), 0)).where(
        NotaryFee.tenant_id == t, NotaryFee.is_deleted == False, NotaryFee.is_paid == False)) or 0)  # noqa: E712
    opex_bd = Decimal(await db.scalar(select(func.coalesce(func.sum(OperationalExpense.amount), 0)).where(
        OperationalExpense.tenant_id == t, OperationalExpense.is_deleted == False, OperationalExpense.is_paid == False)) or 0)  # noqa: E712

    total_aset = kas_bank + persediaan + piutang + retensi
    total_kewajiban = biaya_bd + notaris_bd + opex_bd
    return FinancialPosition(
        kas_bank=kas_bank, persediaan=persediaan, piutang_pembeli=piutang, retensi_bank=retensi,
        total_aset=total_aset, biaya_belum_dibayar=biaya_bd, hutang_notaris=notaris_bd,
        opex_belum_dibayar=opex_bd, total_kewajiban=total_kewajiban,
        kekayaan_bersih=total_aset - total_kewajiban,
    )


@router.get("/project-profit/{project_id}", response_model=ProjectProfitDetail)
async def project_profit_detail(
    project_id: uuid.UUID,
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    """Rincian laba/rugi per unit dalam satu proyek."""
    t = ctx.tenant_id
    proj = (await db.execute(
        select(Project).where(Project.id == project_id, Project.tenant_id == t)
    )).scalar_one_or_none()
    if proj is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Proyek tidak ditemukan")

    units = (await db.execute(
        select(Unit).where(Unit.project_id == project_id, Unit.tenant_id == t)
        .order_by(Unit.block, Unit.unit_number)
    )).scalars().all()
    clients = await _clients_of_project(db, t, project_id)
    client_by_unit = {c.unit_id: c for c in clients if c.unit_id is not None}

    fee_rows = (await db.execute(
        select(func.coalesce(func.sum(NotaryFee.amount), 0)).where(
            NotaryFee.tenant_id == t, NotaryFee.is_deleted == False,  # noqa: E712
            NotaryFee.client_id.in_([c.id for c in clients] or [uuid.UUID(int=0)]))
    )).scalar() or 0

    cost_map = await _project_cost_map(db, t, project_id)
    rows = []
    for u in units:
        grp = cost_map.get(u.id, {})
        m = grp.get("material", Decimal(0)); up = grp.get("upah", Decimal(0))
        ut = grp.get("utilitas", Decimal(0)); ln = grp.get("lain", Decimal(0))
        total = m + up + ut + ln
        c = client_by_unit.get(u.id)
        cv = Decimal(c.contract_value or 0) if c else None
        profit = (cv - total) if cv is not None else None
        rows.append(UnitProfitRow(
            unit_id=u.id, unit_label=_unit_label(u), unit_status=u.status.value,
            client_name=c.full_name if c else None, contract_value=cv,
            cost_material=m, cost_upah=up, cost_utilitas=ut, cost_lain=ln, cost_total=total,
            profit=profit, margin_pct=_margin_pct(profit, cv) if cv is not None else None,
            is_sold=c is not None,
        ))

    return ProjectProfitDetail(
        project_id=project_id, project_name=proj.name, rows=rows,
        cost_general=sum(cost_map.get(None, {}).values(), Decimal(0)),
        cost_notary=Decimal(fee_rows),
        revenue_unattributed=sum(
            (Decimal(c.contract_value or 0) for c in clients if c.unit_id is None), Decimal(0)),
    )


# ═══════════════════════ RINGKASAN SPPR / KPR PER PROYEK (dashboard seksi B) ═══════════════════════
# "SPPR" = pengajuan KPR. "Disetujui bank" = tahap sudah SP3K+ (pakai APPROVED_STAGES yg sama dgn
# laporan lain agar konsisten). "Ditolak" = KprApplication.is_rejected (rejected_date terisi).
# Metode pembayaran = rincian pembeli aktif per ClientPaymentType (cash/kpr) — ditampilkan sbg BARIS,
# bukan donut (keputusan user).

class KprMethodCount(BaseModel):
    method: str          # "kpr" | "cash"
    label: str
    count: int
    pct: float


class ProjectKprRow(BaseModel):
    project_id: uuid.UUID
    project_name: str
    total_sppr: int          # semua pengajuan KPR proyek ini
    approved_bank: int       # sudah SP3K+ (belum ditolak)
    not_approved: int        # belum SP3K (belum ditolak)
    rejected: int            # ditolak bank
    methods: list[KprMethodCount]   # rincian pembeli aktif per cara beli


class KprSummaryReport(BaseModel):
    projects: list[ProjectKprRow]
    sppr_active_total: int   # total pengajuan aktif (belum ditolak) — utk KPI atas


@router.get("/kpr-summary", response_model=KprSummaryReport)
async def kpr_summary(
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    t = ctx.tenant_id
    projects = (await db.execute(
        select(Project.id, Project.name).where(Project.tenant_id == t).order_by(Project.name)
    )).all()

    # KPR → Client (utk project_id). Ambil stage + rejected_date.
    kpr_rows = (await db.execute(
        select(KprApplication.stage, KprApplication.rejected_date, Client.project_id)
        .join(Client, Client.id == KprApplication.client_id)
        .where(KprApplication.tenant_id == t, KprApplication.is_deleted == False,  # noqa: E712
               Client.is_deleted == False)                                          # noqa: E712
    )).all()

    # pembeli aktif per proyek + cara beli (utk metode bayar)
    client_rows = (await db.execute(
        select(Client.project_id, Client.payment_type).where(
            Client.tenant_id == t, Client.is_deleted == False,   # noqa: E712
            Client.status != ClientStatus.INACTIVE)
    )).all()

    from collections import defaultdict
    kpr_by_proj: dict = defaultdict(lambda: {"total": 0, "approved": 0, "belum": 0, "rejected": 0})
    for stage, rejected_date, pid in kpr_rows:
        d = kpr_by_proj[pid]
        d["total"] += 1
        if rejected_date is not None:
            d["rejected"] += 1
        elif stage in APPROVED_STAGES:
            d["approved"] += 1
        else:
            d["belum"] += 1

    method_by_proj: dict = defaultdict(lambda: defaultdict(int))
    for pid, ptype in client_rows:
        method_by_proj[pid][ptype.value if ptype else "cash"] += 1

    _METHOD_LABEL = {"kpr": "KPR", "cash": "Cash / Inhouse"}
    out, active_total = [], 0
    for pid, pname in projects:
        d = kpr_by_proj.get(pid, {"total": 0, "approved": 0, "belum": 0, "rejected": 0})
        active_total += d["total"] - d["rejected"]
        mtot = sum(method_by_proj.get(pid, {}).values())
        methods = [
            KprMethodCount(method=m, label=_METHOD_LABEL.get(m, m.title()), count=c,
                           pct=round(c / mtot * 100, 1) if mtot else 0.0)
            for m, c in sorted(method_by_proj.get(pid, {}).items(), key=lambda x: -x[1])
        ]
        out.append(ProjectKprRow(
            project_id=pid, project_name=pname,
            total_sppr=d["total"], approved_bank=d["approved"],
            not_approved=d["belum"], rejected=d["rejected"], methods=methods,
        ))
    return KprSummaryReport(projects=out, sppr_active_total=active_total)


# ═══════════════════════ RINGKASAN KEUANGAN BERFILTER (dashboard strip) ═══════════════════════
# Filter lokasi (proyek) + bulan. "Uang Masuk" mengikuti bulan terpilih; "Sisa Piutang" & "Total
# Terbayar" SELALU akumulatif (seluruh), bulan tak berlaku — hanya ikut filter lokasi.
# Konsisten dgn /dashboard: hanya pembayaran approved dihitung; piutang = kontrak − terbayar.

class FinanceSummary(BaseModel):
    month: str            # "YYYY-MM" yang dipakai
    cash_in: Decimal      # uang masuk bulan terpilih (+ lokasi)
    outstanding: Decimal  # sisa piutang seluruh (+ lokasi)
    total_paid: Decimal   # total terbayar seluruh (+ lokasi)
    overdue_count: int     # termin terlambat (+ lokasi)
    retention: Decimal = Decimal(0)  # retensi bank seluruh (+ lokasi) — plafon akad − cair


@router.get("/finance-summary", response_model=FinanceSummary)
async def finance_summary(
    project_id: Optional[uuid.UUID] = Query(None),
    month: Optional[str] = Query(None, description="YYYY-MM; default bulan berjalan"),
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    t = ctx.tenant_id
    # rentang bulan
    today = date.today()
    try:
        y, m = (int(x) for x in month.split("-")) if month else (today.year, today.month)
        mstart = date(y, m, 1)
    except (ValueError, AttributeError):
        mstart = today.replace(day=1); y, m = mstart.year, mstart.month
    mend = date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)
    mlabel = f"{y:04d}-{m:02d}"

    _approved = Payment.approval_status == PaymentApprovalStatus.APPROVED
    # id pembeli dalam lokasi (semua bila project_id kosong) — dipakai utk membatasi payment/schedule
    client_conds = [Client.tenant_id == t, Client.is_deleted == False]  # noqa: E712
    if project_id is not None:
        client_conds.append(Client.project_id == project_id)
    client_ids_sq = select(Client.id).where(*client_conds)

    # uang masuk bulan terpilih (approved, dalam lokasi)
    cash_in = Decimal(await db.scalar(
        select(func.coalesce(func.sum(Payment.amount), 0)).where(
            Payment.tenant_id == t, Payment.is_deleted == False, _approved,  # noqa: E712
            Payment.payment_date >= mstart, Payment.payment_date < mend,
            Payment.client_id.in_(client_ids_sq))
    ) or 0)

    # total terbayar (seluruh, approved, dalam lokasi)
    total_paid = Decimal(await db.scalar(
        select(func.coalesce(func.sum(Payment.amount), 0)).where(
            Payment.tenant_id == t, Payment.is_deleted == False, _approved,  # noqa: E712
            Payment.client_id.in_(client_ids_sq))
    ) or 0)

    # total kontrak pembeli aktif (dalam lokasi) → sisa piutang
    contract_conds = [Client.tenant_id == t, Client.is_deleted == False,  # noqa: E712
                      Client.status != ClientStatus.INACTIVE]
    if project_id is not None:
        contract_conds.append(Client.project_id == project_id)
    total_contract = Decimal(await db.scalar(
        select(func.coalesce(func.sum(Client.contract_value), 0)).where(*contract_conds)
    ) or 0)
    outstanding = total_contract - total_paid
    if outstanding < 0:
        outstanding = Decimal(0)

    overdue_count = int(await db.scalar(
        select(func.count()).select_from(PaymentSchedule).where(
            PaymentSchedule.tenant_id == t, PaymentSchedule.is_deleted == False,  # noqa: E712
            PaymentSchedule.status == ScheduleStatus.PENDING,
            PaymentSchedule.due_date < today,
            PaymentSchedule.client_id.in_(client_ids_sq))
    ) or 0)

    # retensi bank (seluruh + lokasi): plafon KPR yang sudah akad − yang sudah cair.
    # Konsisten dgn tab Retensi Bank & ringkasan pembayaran (hanya AKAD_KREDIT/PENCAIRAN).
    kpr_rows = (await db.execute(
        select(KprApplication.id, KprApplication.plafond).where(
            KprApplication.tenant_id == t, KprApplication.is_deleted == False,  # noqa: E712
            KprApplication.stage.in_((KprStage.AKAD_KREDIT, KprStage.PENCAIRAN)),
            KprApplication.plafond.isnot(None), KprApplication.plafond > 0,
            KprApplication.client_id.in_(client_ids_sq))
    )).all()
    retention = Decimal(0)
    if kpr_rows:
        kpr_ids = [r[0] for r in kpr_rows]
        disb_rows = (await db.execute(
            select(Payment.kpr_id, func.coalesce(func.sum(Payment.amount), 0)).where(
                Payment.tenant_id == t, Payment.is_deleted == False, _approved,  # noqa: E712
                Payment.kpr_id.in_(kpr_ids)).group_by(Payment.kpr_id)
        )).all()
        disbursed_by_kpr = {kid: Decimal(v) for kid, v in disb_rows}
        for kid, plaf in kpr_rows:
            ret = Decimal(plaf or 0) - disbursed_by_kpr.get(kid, Decimal(0))
            if ret > 0:
                retention += ret

    return FinanceSummary(month=mlabel, cash_in=cash_in, outstanding=outstanding,
                          total_paid=total_paid, overdue_count=overdue_count, retention=retention)


# ── Rincian tiap tile strip Keuangan (dialog saat angka diklik) ──
class FinanceDetailRow(BaseModel):
    name: str                             # pembeli / (untuk cash_in) pembeli
    project_name: Optional[str] = None
    unit_label: Optional[str] = None
    bank_name: Optional[str] = None
    date: Optional[date] = None
    source_label: Optional[str] = None    # Pembeli/Bank (cash_in) atau nama termin (overdue)
    amount: Optional[Decimal] = None      # angka utama baris
    secondary: Optional[Decimal] = None   # kontrak (outstanding) / plafon (retention)
    tertiary: Optional[Decimal] = None    # terbayar (outstanding) / cair (retention)
    note: Optional[str] = None            # "N hari" (overdue)


@router.get("/finance-detail", response_model=list[FinanceDetailRow])
async def finance_detail(
    kind: str = Query(..., description="cash_in|paid|outstanding|retention|overdue"),
    project_id: Optional[uuid.UUID] = Query(None),
    month: Optional[str] = Query(None, description="YYYY-MM; hanya utk cash_in"),
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    """Daftar data di balik tiap tile strip Keuangan. Filter lokasi & bulan sama dgn
    finance-summary (cash_in ikut bulan; sisanya akumulatif, hanya lokasi)."""
    t = ctx.tenant_id
    today = date.today()
    _notdel_p = Payment.is_deleted == False  # noqa: E712
    _approved = Payment.approval_status == PaymentApprovalStatus.APPROVED
    client_conds = [Client.tenant_id == t, Client.is_deleted == False]  # noqa: E712
    if project_id is not None:
        client_conds.append(Client.project_id == project_id)
    client_ids_sq = select(Client.id).where(*client_conds)
    proj_name = dict((await db.execute(
        select(Project.id, Project.name).where(Project.tenant_id == t))).all())

    if kind == "cash_in":
        try:
            y, m = (int(x) for x in month.split("-")) if month else (today.year, today.month)
            mstart = date(y, m, 1)
        except (ValueError, AttributeError):
            mstart = today.replace(day=1); y, m = mstart.year, mstart.month
        mend = date(y + 1, 1, 1) if m == 12 else date(y, m + 1, 1)
        rows = (await db.execute(
            select(Payment.payment_date, Payment.amount, Payment.source, Client.full_name, Client.project_id)
            .join(Client, Client.id == Payment.client_id)
            .where(Payment.tenant_id == t, _notdel_p, _approved,
                   Payment.payment_date >= mstart, Payment.payment_date < mend,
                   Payment.client_id.in_(client_ids_sq))
            .order_by(Payment.payment_date.desc())
        )).all()
        return [FinanceDetailRow(
            name=fn or "—", project_name=proj_name.get(pid), date=pdate,
            source_label="Bank" if src == PaymentSource.BANK else "Pembeli",
            amount=Decimal(amt or 0)) for pdate, amt, src, fn, pid in rows]

    if kind == "paid":
        rows = (await db.execute(
            select(Client.full_name, Client.project_id, func.coalesce(func.sum(Payment.amount), 0))
            .join(Payment, Payment.client_id == Client.id)
            .where(*client_conds, _notdel_p, _approved)
            .group_by(Client.id, Client.full_name, Client.project_id)
            .order_by(func.coalesce(func.sum(Payment.amount), 0).desc())
        )).all()
        return [FinanceDetailRow(name=fn or "—", project_name=proj_name.get(pid), amount=Decimal(tot or 0))
                for fn, pid, tot in rows if Decimal(tot or 0) > 0]

    if kind == "outstanding":
        oc = [*client_conds, Client.status != ClientStatus.INACTIVE]
        crows = (await db.execute(
            select(Client.id, Client.full_name, Client.project_id, Client.unit_id, Client.contract_value).where(*oc)
        )).all()
        cids = [r[0] for r in crows]
        paid_map = {}
        if cids:
            prows = (await db.execute(
                select(Payment.client_id, func.coalesce(func.sum(Payment.amount), 0)).where(
                    Payment.tenant_id == t, _notdel_p, _approved, Payment.client_id.in_(cids)
                ).group_by(Payment.client_id)
            )).all()
            paid_map = {cid: Decimal(v) for cid, v in prows}
        unit_ids = [r[3] for r in crows if r[3]]
        unit_map = {}
        if unit_ids:
            urows = (await db.execute(
                select(Unit.id, Unit.block, Unit.unit_number).where(Unit.id.in_(unit_ids)))).all()
            unit_map = {uid: "-".join(x for x in [b, n] if x) or None for uid, b, n in urows}
        out = []
        for cid, fn, pid, uid, contract in crows:
            paid = paid_map.get(cid, Decimal(0))
            rem = Decimal(contract or 0) - paid
            if rem > 0:
                out.append(FinanceDetailRow(
                    name=fn or "—", project_name=proj_name.get(pid), unit_label=unit_map.get(uid),
                    amount=rem, secondary=Decimal(contract or 0), tertiary=paid))
        out.sort(key=lambda r: r.amount or Decimal(0), reverse=True)
        return out

    if kind == "retention":
        kpr_rows = (await db.execute(
            select(KprApplication.id, KprApplication.plafond, Client.full_name, Bank.name)
            .join(Client, Client.id == KprApplication.client_id)
            .outerjoin(Bank, Bank.id == KprApplication.bank_id)
            .where(KprApplication.tenant_id == t, KprApplication.is_deleted == False,  # noqa: E712
                   KprApplication.stage.in_((KprStage.AKAD_KREDIT, KprStage.PENCAIRAN)),
                   KprApplication.plafond.isnot(None), KprApplication.plafond > 0,
                   KprApplication.client_id.in_(client_ids_sq))
        )).all()
        disbursed_by_kpr = {}
        if kpr_rows:
            kpr_ids = [r[0] for r in kpr_rows]
            drows = (await db.execute(
                select(Payment.kpr_id, func.coalesce(func.sum(Payment.amount), 0)).where(
                    Payment.tenant_id == t, _notdel_p, _approved, Payment.kpr_id.in_(kpr_ids)
                ).group_by(Payment.kpr_id)
            )).all()
            disbursed_by_kpr = {kid: Decimal(v) for kid, v in drows}
        out = []
        for kid, plaf, fn, bname in kpr_rows:
            disb = disbursed_by_kpr.get(kid, Decimal(0))
            ret = Decimal(plaf or 0) - disb
            if ret > 0:
                out.append(FinanceDetailRow(name=fn or "—", bank_name=bname or "Tanpa Bank",
                                            amount=ret, secondary=Decimal(plaf or 0), tertiary=disb))
        out.sort(key=lambda r: r.amount or Decimal(0), reverse=True)
        return out

    if kind == "overdue":
        rows = (await db.execute(
            select(PaymentSchedule.due_date, PaymentSchedule.amount, PaymentSchedule.label,
                   Client.full_name, Client.project_id)
            .join(Client, Client.id == PaymentSchedule.client_id)
            .where(PaymentSchedule.tenant_id == t, PaymentSchedule.is_deleted == False,  # noqa: E712
                   PaymentSchedule.status == ScheduleStatus.PENDING,
                   PaymentSchedule.due_date < today,
                   PaymentSchedule.client_id.in_(client_ids_sq))
            .order_by(PaymentSchedule.due_date.asc())
        )).all()
        return [FinanceDetailRow(
            name=fn or "—", project_name=proj_name.get(pid), date=due, source_label=label,
            amount=Decimal(amt or 0), note=f"{(today - due).days} hari") for due, amt, label, fn, pid in rows]

    raise HTTPException(status_code=400, detail="kind tidak dikenal")


# ── Rincian pengajuan KPR per proyek (utk dialog klik angka SPPR di dashboard) ──
_KPR_STAGE_LABEL = {
    KprStage.COLLECT_BERKAS: "Collect Berkas",
    KprStage.BERKAS_MASUK_BANK: "Berkas Masuk Bank",
    KprStage.SP3K: "SP3K",
    KprStage.PERSIAPAN_AKAD: "Persiapan Akad",
    KprStage.AKAD_KREDIT: "Akad Kredit",
    KprStage.PENCAIRAN: "Pencairan",
}


class KprDetailRow(BaseModel):
    client_id: uuid.UUID
    client_name: str
    project_name: Optional[str] = None   # diisi & ditampilkan saat filter "Semua"
    unit_label: Optional[str] = None
    bank_name: Optional[str] = None
    stage: str
    stage_label: str
    bucket: str                     # "approved" | "belum" | "rejected"
    plafond: Optional[Decimal] = None
    submitted_date: Optional[date] = None
    sp3k_date: Optional[date] = None


@router.get("/kpr-detail", response_model=list[KprDetailRow])
async def kpr_detail(
    project_id: Optional[uuid.UUID] = Query(None, description="kosong = semua proyek"),
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    """Daftar pengajuan KPR — dipakai dialog saat angka SPPR di dashboard diklik.
    project_id kosong = semua proyek (tenant-wide). Frontend menyaring per `bucket`."""
    from app.models.kpr import Bank
    conds = [KprApplication.tenant_id == ctx.tenant_id, KprApplication.is_deleted == False,  # noqa: E712
             Client.is_deleted == False]                                                      # noqa: E712
    if project_id is not None:
        conds.append(Client.project_id == project_id)
    rows = (await db.execute(
        select(KprApplication, Client.id, Client.full_name, Client.unit_number,
               Unit.block, Unit.unit_number, Bank.name, Project.name)
        .join(Client, Client.id == KprApplication.client_id)
        .outerjoin(Unit, Unit.id == Client.unit_id)
        .outerjoin(Bank, Bank.id == KprApplication.bank_id)
        .outerjoin(Project, Project.id == Client.project_id)
        .where(*conds)
        .order_by(Project.name, Client.full_name)
    )).all()

    out = []
    for k, cid, cname, c_unitnum, blk, u_num, bank, pname in rows:
        if k.rejected_date is not None:
            bucket = "rejected"
        elif k.stage in APPROVED_STAGES:
            bucket = "approved"
        else:
            bucket = "belum"
        label = "-".join(x for x in [blk, u_num] if x) or c_unitnum or None
        out.append(KprDetailRow(
            client_id=cid, client_name=cname, project_name=pname, unit_label=label,
            bank_name=bank, stage=k.stage.value, stage_label=_KPR_STAGE_LABEL.get(k.stage, k.stage.value),
            bucket=bucket, plafond=k.plafond, submitted_date=k.submitted_date, sp3k_date=k.sp3k_date,
        ))
    return out


# ── Rincian unit per proyek (utk dialog klik angka Penjualan di dashboard) ──
_UNIT_STATUS_LABEL = {
    UnitStatus.AVAILABLE: "Tersedia", UnitStatus.BOOKED: "Booking / DP",
    UnitStatus.SOLD: "Terjual", UnitStatus.HANDOVER: "Serah Terima",
}
_SOLD_STATUSES = (UnitStatus.SOLD, UnitStatus.HANDOVER)   # sama dgn units_sold di sales-recap


class UnitDetailRow(BaseModel):
    unit_id: uuid.UUID
    unit_label: str
    project_name: Optional[str] = None   # diisi & ditampilkan saat filter "Semua"
    unit_type: Optional[str] = None
    price: Optional[Decimal] = None
    status: str
    status_label: str
    bucket: str                    # "terjual" | "belum"
    client_id: Optional[uuid.UUID] = None
    client_name: Optional[str] = None
    payment_type: Optional[str] = None        # "cash" | "kpr" (cara beli pembeli)
    payment_type_label: Optional[str] = None
    cash_in: Optional[Decimal] = None         # uang masuk pembeli (approved)
    remaining: Optional[Decimal] = None       # sisa = kontrak − terbayar (clamp≥0)


@router.get("/units-detail", response_model=list[UnitDetailRow])
async def units_detail(
    project_id: Optional[uuid.UUID] = Query(None, description="kosong = semua proyek"),
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    """Daftar unit + pembeli aktif (bila ada) — dipakai dialog saat angka Penjualan diklik.
    project_id kosong = semua proyek (tenant-wide). Frontend menyaring per bucket.
    Kolom uang masuk/sisa/cara beli konsisten dgn strip Keuangan (hanya approved,
    sisa = kontrak − terbayar)."""
    t = ctx.tenant_id
    proj_name = dict((await db.execute(
        select(Project.id, Project.name).where(Project.tenant_id == t))).all())

    # pembeli aktif per unit (anti-dobel 409 → maks satu aktif per unit)
    cl_conds = [Client.tenant_id == t, Client.is_deleted == False,  # noqa: E712
                Client.status != ClientStatus.INACTIVE, Client.unit_id.isnot(None)]
    if project_id is not None:
        cl_conds.append(Client.project_id == project_id)
    client_rows = (await db.execute(
        select(Client.unit_id, Client.id, Client.full_name, Client.payment_type, Client.contract_value)
        .where(*cl_conds)
    )).all()
    client_by_unit: dict = {}
    client_ids = []
    for uid, cid, name, ptype, contract in client_rows:
        client_by_unit[uid] = {"id": cid, "name": name, "ptype": ptype, "contract": Decimal(contract or 0)}
        client_ids.append(cid)

    # uang masuk per pembeli (hanya approved) — sama dgn semua laporan lain
    paid_by_client: dict = {}
    if client_ids:
        pay_rows = (await db.execute(
            select(Payment.client_id, func.coalesce(func.sum(Payment.amount), 0)).where(
                Payment.tenant_id == t, Payment.is_deleted == False,  # noqa: E712
                Payment.approval_status == PaymentApprovalStatus.APPROVED,
                Payment.client_id.in_(client_ids)).group_by(Payment.client_id)
        )).all()
        paid_by_client = {cid: Decimal(v) for cid, v in pay_rows}

    u_conds = [Unit.tenant_id == t]
    if project_id is not None:
        u_conds.append(Unit.project_id == project_id)
    units = (await db.execute(
        select(Unit).where(*u_conds).order_by(Unit.block, Unit.unit_number)
    )).scalars().all()

    out = []
    for u in units:
        c = client_by_unit.get(u.id)
        cid = cname = ptype = ptype_label = None
        cash_in = remaining = None
        if c:
            cid, cname = c["id"], c["name"]
            paid = paid_by_client.get(cid, Decimal(0))
            cash_in = paid
            rem = c["contract"] - paid
            remaining = rem if rem > 0 else Decimal(0)
            ptype = c["ptype"].value if c["ptype"] else "cash"
            ptype_label = "KPR" if ptype == "kpr" else "Cash"
        out.append(UnitDetailRow(
            unit_id=u.id, unit_label="-".join(x for x in [u.block, u.unit_number] if x) or "?",
            project_name=proj_name.get(u.project_id),
            unit_type=u.unit_type, price=u.price, status=u.status.value,
            status_label=_UNIT_STATUS_LABEL.get(u.status, u.status.value),
            bucket="terjual" if u.status in _SOLD_STATUSES else "belum",
            client_id=cid, client_name=cname,
            payment_type=ptype, payment_type_label=ptype_label,
            cash_in=cash_in, remaining=remaining,
        ))
    return out


# ═══════════════════════ RETENSI BANK ═══════════════════════
class BankRetentionRow(BaseModel):
    bank_id: Optional[uuid.UUID]
    bank_name: str
    kpr_count: int
    plafond: Decimal    # plafon yang sudah AKAD (komit)
    disbursed: Decimal  # sudah cair
    retention: Decimal  # sisa ditahan = plafon − cair (≥ 0)


class BankRetentionReport(BaseModel):
    total_plafond: Decimal
    total_disbursed: Decimal
    total_retention: Decimal
    banks: list[BankRetentionRow]


@router.get("/bank-retention", response_model=BankRetentionReport)
async def bank_retention(
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    """Sisa retensi (dana ditahan bank penyalur) per bank. Retensi HANYA dihitung untuk
    KPR yang sudah AKAD KREDIT / PENCAIRAN — sebelum akad, plafon belum komit. Retensi =
    plafon − total cair (uang masuk Bank approved bertaut KPR ini), konsisten dgn
    ringkasan pembayaran & halaman KPR."""
    t = ctx.tenant_id
    kpr_rows = (await db.execute(
        select(KprApplication.id, KprApplication.bank_id, Bank.name, KprApplication.plafond)
        .outerjoin(Bank, Bank.id == KprApplication.bank_id)
        .where(KprApplication.tenant_id == t, KprApplication.is_deleted == False,  # noqa: E712
               KprApplication.stage.in_((KprStage.AKAD_KREDIT, KprStage.PENCAIRAN)),
               KprApplication.plafond.isnot(None), KprApplication.plafond > 0)
    )).all()
    if not kpr_rows:
        return BankRetentionReport(total_plafond=Decimal(0), total_disbursed=Decimal(0),
                                   total_retention=Decimal(0), banks=[])

    kpr_ids = [r[0] for r in kpr_rows]
    disb_rows = (await db.execute(
        select(Payment.kpr_id, func.coalesce(func.sum(Payment.amount), 0)).where(
            Payment.tenant_id == t, Payment.is_deleted == False,  # noqa: E712
            Payment.approval_status == PaymentApprovalStatus.APPROVED,
            Payment.kpr_id.in_(kpr_ids)).group_by(Payment.kpr_id)
    )).all()
    disbursed_by_kpr = {kid: Decimal(v) for kid, v in disb_rows}

    banks: dict = {}
    for kid, bid, bname, plaf in kpr_rows:
        plaf = Decimal(plaf or 0)
        disb = disbursed_by_kpr.get(kid, Decimal(0))
        ret = plaf - disb
        if ret < 0:
            ret = Decimal(0)
        b = banks.get(bid)
        if b is None:
            b = banks[bid] = {"bank_id": bid, "bank_name": bname or "Tanpa Bank", "kpr_count": 0,
                              "plafond": Decimal(0), "disbursed": Decimal(0), "retention": Decimal(0)}
        b["kpr_count"] += 1
        b["plafond"] += plaf
        b["disbursed"] += disb
        b["retention"] += ret

    rows = sorted((BankRetentionRow(**b) for b in banks.values()),
                  key=lambda x: x.retention, reverse=True)
    return BankRetentionReport(
        total_plafond=sum((r.plafond for r in rows), Decimal(0)),
        total_disbursed=sum((r.disbursed for r in rows), Decimal(0)),
        total_retention=sum((r.retention for r in rows), Decimal(0)),
        banks=rows,
    )


# ═══════════════════════ PROYEKSI KAS (cash forecast) ═══════════════════════
class CashProjMonth(BaseModel):
    month: str            # "YYYY-MM"
    termin_in: Decimal    # perkiraan termin pembeli jatuh tempo bulan ini (sisa)
    count: int


class CashProjection(BaseModel):
    current_cash: Decimal          # saldo semua rekening sekarang
    overdue_termin: Decimal        # termin sudah lewat jatuh tempo, belum lunas (masih ditagih)
    months: list[CashProjMonth]    # perkiraan masuk N bulan ke depan
    beyond_termin: Decimal         # termin jatuh tempo setelah horizon
    unscheduled_termin: Decimal    # termin tanpa tanggal jatuh tempo
    retention_expected: Decimal    # retensi bank belum cair (akan masuk, tak terjadwal)
    expenses_unpaid: Decimal       # biaya menunggu bayar (kewajiban keluar)
    contractor_remaining: Decimal  # sisa nilai kontrak borongan (komitmen, informasi)
    projected_liquidity: Decimal   # saldo + semua akan-masuk − biaya menunggu bayar


def _month_key(d: date) -> str:
    return f"{d.year:04d}-{d.month:02d}"


@router.get("/cash-projection", response_model=CashProjection)
async def cash_projection(
    months: int = Query(6, ge=1, le=24),
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    """Proyeksi likuiditas: saldo kas sekarang + perkiraan masuk (termin pembeli per bulan +
    retensi bank) − kewajiban (biaya menunggu bayar). Menjawab 'kas cukup nggak bulan depan?'."""
    t = ctx.tenant_id
    today = date.today()
    _approved = Payment.approval_status == PaymentApprovalStatus.APPROVED

    # ── saldo kas sekarang (total semua rekening; transfer antar-rekening saling meniadakan) ──
    opening = Decimal(await db.scalar(select(func.coalesce(func.sum(CashAccount.opening_balance), 0)).where(
        CashAccount.tenant_id == t, CashAccount.is_deleted == False)) or 0)  # noqa: E712
    cash_in = Decimal(await db.scalar(select(func.coalesce(func.sum(CashBookEntry.amount), 0)).where(
        CashBookEntry.tenant_id == t, CashBookEntry.direction == CashDirection.IN)) or 0)
    cash_out = Decimal(await db.scalar(select(func.coalesce(func.sum(CashBookEntry.amount), 0)).where(
        CashBookEntry.tenant_id == t, CashBookEntry.direction == CashDirection.OUT)) or 0)
    current_cash = opening + cash_in - cash_out

    # ── termin pembeli yang belum lunas (sisa = nominal − sudah dibayar approved) ──
    sched_rows = (await db.execute(
        select(PaymentSchedule.id, PaymentSchedule.due_date, PaymentSchedule.amount).where(
            PaymentSchedule.tenant_id == t, PaymentSchedule.is_deleted == False,  # noqa: E712
            PaymentSchedule.status == ScheduleStatus.PENDING))).all()
    paid_rows = (await db.execute(
        select(Payment.schedule_id, func.coalesce(func.sum(Payment.amount), 0)).where(
            Payment.tenant_id == t, Payment.is_deleted == False, _approved,  # noqa: E712
            Payment.schedule_id.isnot(None)).group_by(Payment.schedule_id))).all()
    paid_map = {sid: Decimal(v) for sid, v in paid_rows}

    # horizon: N bulan kalender ke depan (mulai bulan berjalan)
    horizon_keys = []
    y, m = today.year, today.month
    for _ in range(months):
        horizon_keys.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m = 1; y += 1
    bucket = {k: [Decimal(0), 0] for k in horizon_keys}
    overdue = Decimal(0); beyond = Decimal(0); unscheduled = Decimal(0)
    for sid, due, amt in sched_rows:
        remaining = Decimal(amt or 0) - paid_map.get(sid, Decimal(0))
        if remaining <= 0:
            continue
        if due is None:
            unscheduled += remaining
        elif due < today:
            overdue += remaining
        else:
            k = _month_key(due)
            if k in bucket:
                bucket[k][0] += remaining; bucket[k][1] += 1
            else:
                beyond += remaining
    months_out = [CashProjMonth(month=k, termin_in=bucket[k][0], count=bucket[k][1]) for k in horizon_keys]

    # ── retensi bank belum cair (akan masuk, tak terjadwal) ──
    kpr_rows = (await db.execute(
        select(KprApplication.id, KprApplication.plafond).where(
            KprApplication.tenant_id == t, KprApplication.is_deleted == False,  # noqa: E712
            KprApplication.stage.in_((KprStage.AKAD_KREDIT, KprStage.PENCAIRAN)),
            KprApplication.plafond.isnot(None), KprApplication.plafond > 0))).all()
    retention_expected = Decimal(0)
    if kpr_rows:
        kids = [r[0] for r in kpr_rows]
        disb = {kid: Decimal(v) for kid, v in (await db.execute(
            select(Payment.kpr_id, func.coalesce(func.sum(Payment.amount), 0)).where(
                Payment.tenant_id == t, Payment.is_deleted == False, _approved,  # noqa: E712
                Payment.kpr_id.in_(kids)).group_by(Payment.kpr_id))).all()}
        for kid, plaf in kpr_rows:
            r = Decimal(plaf or 0) - disb.get(kid, Decimal(0))
            if r > 0:
                retention_expected += r

    # ── kewajiban keluar ──
    expenses_unpaid = Decimal(await db.scalar(select(func.coalesce(func.sum(Expense.amount), 0)).where(
        Expense.tenant_id == t, Expense.is_deleted == False, Expense.is_paid == False)) or 0)  # noqa: E712
    # komitmen kontraktor: Σ nilai kontrak − Σ biaya bertaut kontrak (informasi; tak masuk net krn timing tak pasti)
    contract_total = Decimal(await db.scalar(select(func.coalesce(func.sum(ContractorContract.total_value), 0)).where(
        ContractorContract.tenant_id == t, ContractorContract.is_deleted == False)) or 0)  # noqa: E712
    contract_paid = Decimal(await db.scalar(select(func.coalesce(func.sum(Expense.amount), 0)).where(
        Expense.tenant_id == t, Expense.is_deleted == False, Expense.contract_id.isnot(None))) or 0)  # noqa: E712
    contractor_remaining = contract_total - contract_paid
    if contractor_remaining < 0:
        contractor_remaining = Decimal(0)

    total_in = overdue + sum((mm.termin_in for mm in months_out), Decimal(0)) + retention_expected
    projected = current_cash + total_in - expenses_unpaid

    return CashProjection(
        current_cash=current_cash, overdue_termin=overdue, months=months_out, beyond_termin=beyond,
        unscheduled_termin=unscheduled, retention_expected=retention_expected,
        expenses_unpaid=expenses_unpaid, contractor_remaining=contractor_remaining,
        projected_liquidity=projected,
    )


# ═══════════════════════ EKUALISASI PAJAK (register + cross-check) ═══════════════════════
# Cross-check pemeriksa pajak: Penjualan (nilai AJB) = DPP PPh Final = DPP PPN.
# Developer: PPh Final subsidi 1% / komersial 2,5%; PPN subsidi BEBAS, komersial DTP (bayar 0, DPP tetap dilaporkan).
_TAX_STATUS_LABEL = {"belum": "Belum", "dibayar": "Dibayar", "validasi": "Validasi", "dtp": "DTP", "bebas": "Bebas"}
_PPN_RATE = Decimal("0.11")


class TaxEqRow(BaseModel):
    client_id: uuid.UUID
    pembeli: str
    nik: Optional[str] = None
    proyek: Optional[str] = None
    unit_label: Optional[str] = None
    kategori: str                       # subsidi | komersial
    tgl_ajb: Optional[date] = None
    nilai_ajb: Decimal                  # DPP (dari PPh, fallback kontrak)
    pph_amount: Optional[Decimal] = None
    pph_status: Optional[str] = None
    pph_ntpn: Optional[str] = None
    ppn_status: Optional[str] = None    # DTP (komersial) / Bebas (subsidi) / dll
    ppn_dpp: Optional[Decimal] = None
    ppn_dtp: Optional[Decimal] = None   # PPN terutang DTP = 11% × DPP (komersial)
    bphtb_amount: Optional[Decimal] = None
    bphtb_status: Optional[str] = None
    lengkap: bool = False               # punya catatan PPh + PPN + BPHTB


class TaxEqReport(BaseModel):
    period: str
    penjualan: Decimal          # Σ nilai AJB (peredaran usaha)
    dpp_pph: Decimal
    dpp_ppn: Decimal
    selisih_pph: Decimal        # penjualan − DPP PPh
    selisih_ppn: Decimal        # penjualan − DPP PPN
    pph_terutang: Decimal
    ppn_dtp_total: Decimal      # PPN ditanggung pemerintah (komersial)
    ppn_bebas_count: int        # jumlah unit PPN dibebaskan (subsidi)
    bphtb_total: Decimal
    incomplete: int
    rows: list[TaxEqRow]


async def _tax_eq_data(db, tenant_id, year: int, month: Optional[int]) -> TaxEqReport:
    from calendar import monthrange
    t = tenant_id
    if month:
        start = date(year, month, 1)
        end = date(year, month, monthrange(year, month)[1])
    else:
        start = date(year, 1, 1); end = date(year, 12, 31)
    # pembeli dgn tgl kontrak dalam periode (transaksi penjualan), bukan batal
    cconds = [Client.tenant_id == t, Client.is_deleted == False,  # noqa: E712
              Client.status != ClientStatus.INACTIVE,
              Client.contract_date >= start, Client.contract_date <= end]
    clients = (await db.execute(select(Client).where(*cconds).order_by(Client.contract_date))).scalars().all()
    proj_names = dict((await db.execute(select(Project.id, Project.name).where(Project.tenant_id == t))).all())
    unit_ids = {c.unit_id for c in clients if c.unit_id}
    units = {u.id: u for u in (await db.execute(select(Unit).where(Unit.id.in_(unit_ids)))).scalars().all()} if unit_ids else {}
    cids = [c.id for c in clients]
    tax_rows = (await db.execute(select(TaxRecord).where(
        TaxRecord.tenant_id == t, TaxRecord.client_id.in_(cids), TaxRecord.is_deleted == False))).scalars().all() if cids else []  # noqa: E712
    by = {}
    for r in tax_rows:
        by[(r.client_id, r.tax_type)] = r

    rows = []
    penjualan = dpp_pph = dpp_ppn = pph_terutang = ppn_dtp_total = bphtb_total = Decimal(0)
    ppn_bebas_count = 0; incomplete = 0
    for c in clients:
        pph = by.get((c.id, TaxType.PPH)); ppn = by.get((c.id, TaxType.PPN)); bphtb = by.get((c.id, TaxType.BPHTB))
        kategori = (pph.category if pph else (ppn.category if ppn else "komersial"))
        nilai_ajb = Decimal((pph.base_amount if pph and pph.base_amount is not None else
                             (ppn.base_amount if ppn and ppn.base_amount is not None else (c.contract_value or 0))) or 0)
        u = units.get(c.unit_id)
        unit_label = ("-".join(x for x in [u.block, u.unit_number] if x) or None) if u else None
        ppn_dpp = Decimal(ppn.base_amount or 0) if ppn else None
        ppn_dtp = (ppn_dpp * _PPN_RATE) if (ppn and kategori == "komersial" and (ppn.status.value if ppn.status else "") == "dtp") else None
        row = TaxEqRow(
            client_id=c.id, pembeli=c.full_name, nik=c.nik, proyek=proj_names.get(c.project_id),
            unit_label=unit_label, kategori=kategori,
            tgl_ajb=(pph.tax_date if pph else None) or c.contract_date, nilai_ajb=nilai_ajb,
            pph_amount=Decimal(pph.amount or 0) if pph else None,
            pph_status=_TAX_STATUS_LABEL.get(pph.status.value) if (pph and pph.status) else None,
            pph_ntpn=pph.ntpn if pph else None,
            ppn_status=_TAX_STATUS_LABEL.get(ppn.status.value) if (ppn and ppn.status) else None,
            ppn_dpp=ppn_dpp, ppn_dtp=ppn_dtp,
            bphtb_amount=Decimal(bphtb.amount or 0) if bphtb else None,
            bphtb_status=_TAX_STATUS_LABEL.get(bphtb.status.value) if (bphtb and bphtb.status) else None,
            lengkap=bool(pph and ppn and bphtb),
        )
        rows.append(row)
        penjualan += nilai_ajb
        if pph and pph.base_amount is not None:
            dpp_pph += Decimal(pph.base_amount)
        if pph and pph.amount is not None:
            pph_terutang += Decimal(pph.amount)
        if ppn and ppn.base_amount is not None:
            dpp_ppn += Decimal(ppn.base_amount)
        if ppn_dtp:
            ppn_dtp_total += ppn_dtp
        if ppn and kategori == "subsidi":
            ppn_bebas_count += 1
        if bphtb and bphtb.amount is not None:
            bphtb_total += Decimal(bphtb.amount)
        if not (pph and ppn and bphtb):
            incomplete += 1

    period = f"{year}" if not month else f"{year}-{month:02d}"
    return TaxEqReport(
        period=period, penjualan=penjualan, dpp_pph=dpp_pph, dpp_ppn=dpp_ppn,
        selisih_pph=penjualan - dpp_pph, selisih_ppn=penjualan - dpp_ppn,
        pph_terutang=pph_terutang, ppn_dtp_total=ppn_dtp_total, ppn_bebas_count=ppn_bebas_count,
        bphtb_total=bphtb_total, incomplete=incomplete, rows=rows)


@router.get("/tax-equalization", response_model=TaxEqReport)
async def tax_equalization(
    year: int = Query(...), month: Optional[int] = Query(None, ge=1, le=12),
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    return await _tax_eq_data(db, ctx.tenant_id, year, month)


@router.get("/tax-equalization/export")
async def tax_equalization_export(
    year: int = Query(...), month: Optional[int] = Query(None, ge=1, le=12),
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    rep = await _tax_eq_data(db, ctx.tenant_id, year, month)
    wb = openpyxl.Workbook()
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", fgColor="1E3A5F")

    def money(v):
        return float(v) if v is not None else None

    # Sheet Register
    ws = wb.active; ws.title = "Register"
    cols = ["Tgl AJB", "Pembeli", "NIK", "Proyek", "Unit", "Kategori", "Nilai AJB (DPP)",
            "PPh Jumlah", "PPh Status", "NTPN PPh", "PPN Status", "PPN DPP", "PPN Terutang DTP",
            "BPHTB Jumlah", "BPHTB Status"]
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=h); c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = 15
    r = 2
    for x in rep.rows:
        vals = [x.tgl_ajb.strftime("%d/%m/%Y") if x.tgl_ajb else "", x.pembeli, x.nik or "", x.proyek or "",
                x.unit_label or "", x.kategori, money(x.nilai_ajb), money(x.pph_amount), x.pph_status or "",
                x.pph_ntpn or "", x.ppn_status or "", money(x.ppn_dpp), money(x.ppn_dtp),
                money(x.bphtb_amount), x.bphtb_status or ""]
        for i, v in enumerate(vals, start=1):
            cc = ws.cell(row=r, column=i, value=v)
            if i in (7, 8, 12, 13, 14):
                cc.number_format = "#,##0"
        r += 1
    ws.freeze_panes = "A2"

    # Sheet Ekualisasi
    we = wb.create_sheet("Ekualisasi")
    we.column_dimensions["A"].width = 34; we.column_dimensions["B"].width = 22
    rows_eq = [
        ("EKUALISASI PAJAK — PERIODE", rep.period),
        ("", ""),
        ("Penjualan (Σ Nilai AJB)", money(rep.penjualan)),
        ("DPP PPh Final", money(rep.dpp_pph)),
        ("Selisih (Penjualan − DPP PPh)", money(rep.selisih_pph)),
        ("DPP PPN", money(rep.dpp_ppn)),
        ("Selisih (Penjualan − DPP PPN)", money(rep.selisih_ppn)),
        ("", ""),
        ("PPh Final terutang", money(rep.pph_terutang)),
        ("PPN terutang DTP (komersial)", money(rep.ppn_dtp_total)),
        ("Unit PPN dibebaskan (subsidi)", rep.ppn_bebas_count),
        ("BPHTB", money(rep.bphtb_total)),
        ("Baris belum lengkap catatan pajak", rep.incomplete),
    ]
    for i, (k, v) in enumerate(rows_eq, start=1):
        a = we.cell(row=i, column=1, value=k); a.font = Font(name="Arial", bold=(i == 1 or "Selisih" in k), size=10)
        b = we.cell(row=i, column=2, value=v)
        if isinstance(v, (int, float)) and "dibebaskan" not in k and "belum lengkap" not in k:
            b.number_format = "#,##0"
    bio = _io.BytesIO(); wb.save(bio); bio.seek(0)
    fn = f"Ekualisasi_Pajak_{rep.period}.xlsx"
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{fn}"'})
