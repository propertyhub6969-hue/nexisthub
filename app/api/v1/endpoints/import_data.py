"""Importir data dari Excel (migrasi data klien).

Alur: unduh template terisi → klien lengkapi → upload → PRATINJAU (insert/update/error)
→ TERAPKAN (upsert). Sheet pertama yang didukung: UNIT (kunci: Proyek + Blok + Nomor Unit).
Pembeli & Pembayaran menyusul.
"""
import io
import os
import re
import uuid
import zipfile
import mimetypes
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core import storage
from app.core.audit import record_audit
from app.core.unit_status import unit_status_for_client, set_unit_status
from app.api.deps import get_current_context, AuthContext
from app.models.property import Project, Unit, UnitStatus
from app.models.marketing import Client, ClientStatus, ClientPaymentType
from app.models.document import Document, DocStatus
from app.models.import_batch import ImportBatch, ImportBatchItem
from app.models.payment import Payment, PaymentSource, PaymentPurpose, PaymentMethod, PaymentApprovalStatus
from app.models.cashbook import CashBookEntry
from app.core.cashbook import sync_payment_cashbook

_MAX_FILE_BYTES = 10 * 1024 * 1024  # 10 MB / file (samakan dgn modul Dokumen)

router = APIRouter()

# ── peta status ID ↔ enum ──
STATUS_TO_LABEL = {UnitStatus.AVAILABLE: "Tersedia", UnitStatus.BOOKED: "Booking",
                   UnitStatus.SOLD: "Terjual", UnitStatus.HANDOVER: "Serah Terima"}
LABEL_TO_STATUS = {v.lower(): k for k, v in STATUS_TO_LABEL.items()}
LABEL_TO_STATUS.update({"available": UnitStatus.AVAILABLE, "booked": UnitStatus.BOOKED,
                        "sold": UnitStatus.SOLD, "handover": UnitStatus.HANDOVER})

# header UNIT → field internal (dicocokkan longgar: lower, tanpa "*", by kata kunci)
def _norm_header(h) -> str:
    return str(h or "").strip().lower().replace("*", "").strip()

def _field_of(h: str) -> Optional[str]:
    n = _norm_header(h)
    if n.startswith("proyek"): return "project"
    if n.startswith("blok"): return "block"
    if n.startswith("nomor unit") or n == "unit" or n.startswith("no unit"): return "unit_number"
    if n.startswith("tipe"): return "unit_type"
    if n.startswith("luas tanah"): return "land_area"
    if n.startswith("luas bangunan"): return "building_area"
    if n.startswith("harga"): return "price"
    if n.startswith("diskon"): return "discount"
    if n.startswith("status"): return "status"
    return None


def _to_decimal(v) -> Optional[Decimal]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace("Rp", "").replace("rp", "").strip()
    if s == "":
        return None
    # buang pemisah ribuan bila ada (1.234.567 atau 1,234,567)
    if s.count(".") > 1 or (s.count(",") > 1):
        s = s.replace(".", "").replace(",", "")
    else:
        s = s.replace(",", ".") if ("," in s and "." not in s) else s.replace(",", "")
    return Decimal(s)


def _to_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(s)


# peta label ID pembeli ↔ enum
PT_TO_LABEL = {ClientPaymentType.CASH: "Cash", ClientPaymentType.KPR: "KPR"}
LABEL_TO_PT = {"cash": ClientPaymentType.CASH, "kpr": ClientPaymentType.KPR,
               "tunai": ClientPaymentType.CASH}
CST_TO_LABEL = {ClientStatus.ACTIVE: "Aktif", ClientStatus.COMPLETED: "Selesai",
                ClientStatus.INACTIVE: "Batal"}
LABEL_TO_CST = {"aktif": ClientStatus.ACTIVE, "selesai": ClientStatus.COMPLETED,
                "batal": ClientStatus.INACTIVE, "active": ClientStatus.ACTIVE,
                "completed": ClientStatus.COMPLETED, "inactive": ClientStatus.INACTIVE,
                "nonaktif": ClientStatus.INACTIVE}


class ImportRow(BaseModel):
    row: int
    action: str            # insert | update | error
    label: str             # "Daiyan Permai / A / 036"
    errors: list[str] = []
    note: Optional[str] = None   # ringkasan perubahan (update)


class ImportPreview(BaseModel):
    sheet: str = "UNIT"
    total: int
    to_insert: int
    to_update: int
    error_count: int
    to_skip: int = 0
    rows: list[ImportRow]


class ImportCommitResult(BaseModel):
    batch_id: str
    inserted: int
    updated: int
    error_count: int
    skipped: int = 0
    rows: list[ImportRow]


async def _save_batch(db, ctx, entity, inserted_items, updated, note=None):
    """Catat 1 batch impor + record yang DITAMBAH (untuk undo). inserted_items: list (resource_id, file_key)."""
    batch = ImportBatch(tenant_id=ctx.tenant_id, user_id=ctx.user_id, entity=entity,
                        inserted=len(inserted_items), updated=updated, note=note)
    db.add(batch)
    await db.flush()
    for rid, fkey in inserted_items:
        db.add(ImportBatchItem(batch_id=batch.id, resource=entity, resource_id=rid, file_key=fkey))
    await db.flush()
    await record_audit(db, ctx.tenant_id, ctx.user_id, "IMPORT", entity, batch.id,
                       new_data={"inserted": len(inserted_items), "updated": updated})
    return batch


async def _load_maps(db: AsyncSession, tenant_id):
    """peta nama-proyek→id dan (proyek,blok,unit)→Unit."""
    projs = (await db.execute(
        select(Project.id, Project.name).where(Project.tenant_id == tenant_id))).all()
    proj_by_name = {name.strip().lower(): pid for pid, name in projs}
    units = (await db.execute(
        select(Unit).where(Unit.tenant_id == tenant_id))).scalars().all()
    unit_by_key = {}
    for u in units:
        key = (u.project_id, (u.block or "").strip().lower(), (u.unit_number or "").strip().lower())
        unit_by_key[key] = u
    return proj_by_name, unit_by_key


def _read_unit_rows(contents: bytes):
    """baca sheet UNIT → daftar (row_num, dict field mentah). Error kalau sheet tak ada."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="File bukan Excel (.xlsx) yang valid.")
    ws = None
    for name in wb.sheetnames:
        if name.strip().lower() == "unit":
            ws = wb[name]; break
    if ws is None:
        raise HTTPException(status_code=400, detail="Sheet 'UNIT' tidak ditemukan di file.")
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    col_map = {}
    for i, h in enumerate(header):
        f = _field_of(h)
        if f:
            col_map[f] = i
    if "project" not in col_map or "unit_number" not in col_map:
        raise HTTPException(status_code=400, detail="Header 'Proyek' dan 'Nomor Unit' wajib ada di sheet UNIT.")
    out = []
    rnum = 1
    for raw in rows_iter:
        rnum += 1
        def g(field):
            idx = col_map.get(field)
            return raw[idx] if (idx is not None and idx < len(raw)) else None
        d = {f: g(f) for f in ["project", "block", "unit_number", "unit_type",
                               "land_area", "building_area", "price", "discount", "status"]}
        # lewati baris kosong total
        if all((v is None or str(v).strip() == "") for v in d.values()):
            continue
        out.append((rnum, d))
    return out


def _classify(rows, proj_by_name, unit_by_key):
    """→ (parsed list of (row_num, kind, payload, unit_obj_or_None, ImportRow))"""
    parsed = []
    for rnum, d in rows:
        errors = []
        pname = str(d["project"] or "").strip()
        unum = str(d["unit_number"] or "").strip()
        block = (str(d["block"]).strip() if d["block"] not in (None, "") else None)
        label = " / ".join(x for x in [pname or "?", block or "", unum or "?"] if x != "")
        pid = proj_by_name.get(pname.lower()) if pname else None
        if not pname:
            errors.append("Proyek kosong")
        elif pid is None:
            errors.append(f"Proyek '{pname}' tidak ada di sistem")
        if not unum:
            errors.append("Nomor Unit kosong")

        vals = {}
        for f, lbl in [("land_area", "Luas Tanah"), ("building_area", "Luas Bangunan"),
                       ("price", "Harga"), ("discount", "Diskon")]:
            try:
                vals[f] = _to_decimal(d[f])
            except (InvalidOperation, ValueError):
                errors.append(f"{lbl} bukan angka: '{d[f]}'")
                vals[f] = None

        st = None
        if d["status"] not in (None, ""):
            st = LABEL_TO_STATUS.get(str(d["status"]).strip().lower())
            if st is None:
                errors.append(f"Status tidak dikenal: '{d['status']}'")
        vals["status"] = st
        vals["unit_type"] = (str(d["unit_type"]).strip() if d["unit_type"] not in (None, "") else None)

        if errors:
            parsed.append((rnum, "error", None, None, ImportRow(row=rnum, action="error", label=label, errors=errors)))
            continue

        key = (pid, (block or "").strip().lower(), unum.strip().lower())
        existing = unit_by_key.get(key)
        payload = {"project_id": pid, "block": block, "unit_number": unum, **vals}
        if existing:
            changed = []
            for f in ["unit_type", "land_area", "building_area", "price", "discount", "status"]:
                nv = vals[f]
                if nv is None:
                    continue  # kosong = jangan timpa
                cur = getattr(existing, f)
                if f in ("land_area", "building_area", "price", "discount"):
                    cur = Decimal(str(cur)) if cur is not None else None
                if cur != nv:
                    changed.append(f)
            note = ("Ubah: " + ", ".join(changed)) if changed else "Tidak ada perubahan"
            parsed.append((rnum, "update", payload, existing, ImportRow(row=rnum, action="update", label=label, note=note)))
        else:
            parsed.append((rnum, "insert", payload, None, ImportRow(row=rnum, action="insert", label=label)))
    return parsed


@router.post("/units/preview", response_model=ImportPreview)
async def preview_units(
    file: UploadFile = File(...),
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    contents = await file.read()
    rows = _read_unit_rows(contents)
    proj_by_name, unit_by_key = await _load_maps(db, ctx.tenant_id)
    parsed = _classify(rows, proj_by_name, unit_by_key)
    ins = sum(1 for p in parsed if p[1] == "insert")
    upd = sum(1 for p in parsed if p[1] == "update")
    err = sum(1 for p in parsed if p[1] == "error")
    return ImportPreview(total=len(parsed), to_insert=ins, to_update=upd, error_count=err,
                         rows=[p[4] for p in parsed])


@router.post("/units/commit", response_model=ImportCommitResult)
async def commit_units(
    file: UploadFile = File(...),
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    contents = await file.read()
    rows = _read_unit_rows(contents)
    proj_by_name, unit_by_key = await _load_maps(db, ctx.tenant_id)
    parsed = _classify(rows, proj_by_name, unit_by_key)
    inserted = updated = 0
    new_units = []
    for rnum, kind, payload, existing, rowres in parsed:
        if kind == "insert":
            u = Unit(tenant_id=ctx.tenant_id, project_id=payload["project_id"],
                     block=payload["block"], unit_number=payload["unit_number"],
                     unit_type=payload["unit_type"], land_area=payload["land_area"],
                     building_area=payload["building_area"], price=payload["price"],
                     discount=payload["discount"],
                     status=payload["status"] or UnitStatus.AVAILABLE)
            db.add(u); new_units.append(u); inserted += 1
        elif kind == "update":
            for f in ["unit_type", "land_area", "building_area", "price", "discount", "status"]:
                nv = payload[f]
                if nv is not None:
                    setattr(existing, f, nv)
            updated += 1
    batch_id = ""
    if inserted or updated:
        await db.flush()
        batch = await _save_batch(db, ctx, "units", [(u.id, None) for u in new_units], updated)
        batch_id = str(batch.id)
        await db.commit()
    err = sum(1 for p in parsed if p[1] == "error")
    return ImportCommitResult(batch_id=batch_id, inserted=inserted, updated=updated,
                              error_count=err, rows=[p[4] for p in parsed])


# ═══════════════════════ UNDUH TEMPLATE (terisi data saat ini) ═══════════════════════
_FONT = "Arial"
_hdr_fill = PatternFill("solid", fgColor="1E3A5F")
_hdr_key = PatternFill("solid", fgColor="D9C9A0")
_key_cell = PatternFill("solid", fgColor="F3EEE0")
_ex_fill = PatternFill("solid", fgColor="FFF7D6")
_thin = Side(style="thin", color="D0D0D0")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _cell(ws, r, c, v, *, bold=False, color="000000", italic=False, fill=None, align="left", numfmt=None):
    x = ws.cell(row=r, column=c, value=v)
    x.font = Font(name=_FONT, bold=bold, color=color, size=10, italic=italic)
    x.alignment = Alignment(horizontal=align, vertical="center")
    if fill: x.fill = fill
    if numfmt: x.number_format = numfmt
    x.border = _border
    return x


def _headers(ws, cols):
    for i, (title, width, is_key, req) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
        c = ws.cell(row=1, column=i, value=title + (" *" if req else ""))
        c.font = Font(name=_FONT, bold=True, size=10, color=("3A3120" if is_key else "FFFFFF"))
        c.fill = _hdr_key if is_key else _hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def _dv(ws, values, col_idx, first, last):
    d = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    ws.add_data_validation(d)
    L = get_column_letter(col_idx)
    d.add(f"{L}{first}:{L}{last}")


def _build_template(projects, units) -> bytes:
    wb = openpyxl.Workbook()
    # Petunjuk
    ws = wb.active; ws.title = "Petunjuk"; ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 30; ws.column_dimensions["C"].width = 90
    t = ws.cell(row=1, column=2, value="TEMPLATE IMPOR DATA — NEXISTHUB")
    t.font = Font(name=_FONT, bold=True, color="FFFFFF", size=13); ws.merge_cells("B1:C1")
    ws["B1"].fill = _hdr_fill; ws["C1"].fill = _hdr_fill; ws.row_dimensions[1].height = 26
    tips = [
        ("Cara kerja", "Lengkapi sheet UNIT (sudah terisi data Anda). Upload kembali → pratinjau (diperbarui/baru/error) → Terapkan."),
        ("Kolom KUNCI (krem)", "Proyek + Blok + Nomor Unit = penanda. JANGAN diubah — dipakai untuk MEMPERBARUI unit yang ada (bukan membuat dobel)."),
        ("Kolom kosong", "Isi yang kosong (tipe, luas, harga). Dibiarkan kosong = nilai lama tidak diubah."),
        ("Format angka", "Angka polos tanpa 'Rp'/titik. Contoh: 185000000."),
        ("Status (dropdown)", "Tersedia · Booking · Terjual · Serah Terima"),
        ("Nama proyek (persis)", " · ".join(projects)),
    ]
    r = 3
    for a, b in tips:
        ca = ws.cell(row=r, column=2, value=a); ca.font = Font(name=_FONT, bold=True, size=10)
        ca.alignment = Alignment(vertical="top", wrap_text=True)
        cb = ws.cell(row=r, column=3, value=b); cb.font = Font(name=_FONT, size=10)
        cb.alignment = Alignment(vertical="top", wrap_text=True); ws.row_dimensions[r].height = 30
        r += 1
    # UNIT
    wu = wb.create_sheet("UNIT"); wu.sheet_view.showGridLines = False
    ucols = [("Proyek", 20, True, True), ("Blok", 8, True, False), ("Nomor Unit", 12, True, True),
             ("Tipe", 10, False, False), ("Luas Tanah (m²)", 14, False, False),
             ("Luas Bangunan (m²)", 16, False, False), ("Harga (Rp)", 16, False, False),
             ("Diskon (Rp)", 14, False, False), ("Status", 15, False, False)]
    _headers(wu, ucols)
    row = 2
    for u in units:
        _cell(wu, row, 1, u["proyek"], fill=_key_cell)
        _cell(wu, row, 2, u["blok"], fill=_key_cell, align="center")
        _cell(wu, row, 3, u["unit"], fill=_key_cell, align="center", bold=True)
        _cell(wu, row, 4, u["tipe"], align="center")
        _cell(wu, row, 5, u["lt"], align="right", numfmt="#,##0.##")
        _cell(wu, row, 6, u["lb"], align="right", numfmt="#,##0.##")
        _cell(wu, row, 7, u["harga"], align="right", numfmt="#,##0")
        _cell(wu, row, 8, u["diskon"], align="right", numfmt="#,##0")
        _cell(wu, row, 9, u["status"], align="center")
        row += 1
    _dv(wu, ["Tersedia", "Booking", "Terjual", "Serah Terima"], 9, 2, row + 500)
    _dv(wu, projects, 1, 2, row + 500)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio.getvalue()


@router.get("/units/template")
async def download_units_template(
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    """Template Excel UNIT SUDAH TERISI data unit tenant saat ini (ekspor-terisi)."""
    projs = (await db.execute(
        select(Project.id, Project.name).where(Project.tenant_id == ctx.tenant_id).order_by(Project.name))).all()
    pmap = {pid: name for pid, name in projs}
    pnames = [name for _, name in projs]
    units = (await db.execute(
        select(Unit).where(Unit.tenant_id == ctx.tenant_id)
        .order_by(Unit.project_id, Unit.block, Unit.unit_number))).scalars().all()
    rows = []
    for u in units:
        rows.append({
            "proyek": pmap.get(u.project_id, ""), "blok": u.block or "", "unit": u.unit_number or "",
            "tipe": u.unit_type or "", "lt": float(u.land_area) if u.land_area is not None else None,
            "lb": float(u.building_area) if u.building_area is not None else None,
            "harga": float(u.price) if u.price is not None else None,
            "diskon": float(u.discount) if u.discount is not None else None,
            "status": STATUS_TO_LABEL.get(u.status, "Tersedia"),
        })
    data = _build_template(pnames, rows)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Template_Import_Unit.xlsx"'})


# ═══════════════════════ PEMBELI & KONTRAK ═══════════════════════
def _client_field_of(h: str) -> Optional[str]:
    n = _norm_header(h)
    if "ppjb" in n: return "ppjb_number"
    if "ajb" in n: return "ajb_number"
    if n.startswith("nama"): return "full_name"
    if n.startswith("nik"): return "nik"
    if n.startswith("no. hp") or n.startswith("no hp") or n.startswith("hp") or "telepon" in n or n.startswith("telp"): return "phone"
    if n.startswith("email"): return "email"
    if n.startswith("alamat"): return "address"
    if n.startswith("proyek"): return "project"
    if n.startswith("nomor unit") or n.startswith("no unit") or n == "unit": return "unit_number"
    if "pembayaran" in n or "cara beli" in n: return "payment_type"
    if "nilai kontrak" in n or n.startswith("harga") or n.startswith("kontrak"): return "contract_value"
    if n.startswith("tanggal"): return "contract_date"
    if n.startswith("status"): return "status"
    return None


def _read_client_rows(contents: bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="File bukan Excel (.xlsx) yang valid.")
    ws = None
    for name in wb.sheetnames:
        if name.strip().lower().startswith("pembeli"):
            ws = wb[name]; break
    if ws is None:
        raise HTTPException(status_code=400, detail="Sheet 'PEMBELI' tidak ditemukan di file.")
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return []
    col_map = {}
    for i, h in enumerate(header):
        f = _client_field_of(h)
        if f and f not in col_map:
            col_map[f] = i
    if "full_name" not in col_map:
        raise HTTPException(status_code=400, detail="Header 'Nama Pembeli' wajib ada di sheet PEMBELI.")
    fields = ["full_name", "nik", "phone", "email", "address", "project", "unit_number",
              "payment_type", "contract_value", "contract_date", "ppjb_number", "ajb_number", "status"]
    out = []
    rnum = 1
    for raw in it:
        rnum += 1
        d = {f: (raw[col_map[f]] if (f in col_map and col_map[f] < len(raw)) else None) for f in fields}
        if all((v is None or str(v).strip() == "") for v in d.values()):
            continue
        # lewati baris contoh
        if str(d.get("full_name") or "").strip().upper().startswith("CONTOH"):
            continue
        out.append((rnum, d))
    return out


async def _load_client_maps(db: AsyncSession, tenant_id):
    projs = (await db.execute(select(Project.id, Project.name).where(Project.tenant_id == tenant_id))).all()
    proj_by_name = {name.strip().lower(): pid for pid, name in projs}
    # unit per (project, unit_number) — bisa >1 kalau nomor sama beda blok
    units = (await db.execute(select(Unit.id, Unit.project_id, Unit.unit_number).where(Unit.tenant_id == tenant_id))).all()
    unit_by_pu = {}
    for uid, pid, unum in units:
        unit_by_pu.setdefault((pid, (unum or "").strip().lower()), []).append(uid)
    clients = (await db.execute(select(Client).where(Client.tenant_id == tenant_id, Client.is_deleted == False))).scalars().all()  # noqa: E712
    by_nik = {}
    by_unit = {}
    for c in clients:
        if c.nik:
            by_nik[c.nik.strip()] = c
        if c.unit_id and c.status != ClientStatus.INACTIVE:
            by_unit[c.unit_id] = c
    return proj_by_name, unit_by_pu, by_nik, by_unit


def _classify_clients(rows, proj_by_name, unit_by_pu, by_nik, by_unit):
    """→ list (row_num, kind, payload, existing_client, resolved_unit_id, ImportRow)."""
    parsed = []
    for rnum, d in rows:
        errors = []
        name = str(d["full_name"] or "").strip()
        if not name:
            errors.append("Nama kosong")
        nik = (str(d["nik"]).strip() if d["nik"] not in (None, "") else None)
        pname = str(d["project"] or "").strip()
        unum = (str(d["unit_number"]).strip() if d["unit_number"] not in (None, "") else None)
        label = name or "?"
        pid = proj_by_name.get(pname.lower()) if pname else None
        if pname and pid is None:
            errors.append(f"Proyek '{pname}' tidak ada")

        resolved_unit = None
        if unum:
            if pid is None:
                errors.append("Nomor Unit diisi tapi Proyek kosong/salah")
            else:
                cand = unit_by_pu.get((pid, unum.lower()), [])
                if len(cand) == 0:
                    errors.append(f"Unit '{unum}' tak ada di proyek {pname}")
                elif len(cand) > 1:
                    errors.append(f"Unit '{unum}' ambigu (ada di >1 blok) — pakai importir Unit dulu")
                else:
                    resolved_unit = cand[0]

        pt = None
        if d["payment_type"] not in (None, ""):
            pt = LABEL_TO_PT.get(str(d["payment_type"]).strip().lower())
            if pt is None:
                errors.append(f"Tipe Pembayaran tak dikenal: '{d['payment_type']}'")
        st = None
        if d["status"] not in (None, ""):
            st = LABEL_TO_CST.get(str(d["status"]).strip().lower())
            if st is None:
                errors.append(f"Status tak dikenal: '{d['status']}'")
        try:
            cval = _to_decimal(d["contract_value"])
        except (InvalidOperation, ValueError):
            errors.append(f"Nilai Kontrak bukan angka: '{d['contract_value']}'"); cval = None
        try:
            cdate = _to_date(d["contract_date"])
        except ValueError:
            errors.append(f"Tanggal Kontrak tak valid: '{d['contract_date']}' (pakai dd/mm/yyyy)"); cdate = None

        # cari existing: NIK dulu, lalu unit
        existing = None
        if nik and nik in by_nik:
            existing = by_nik[nik]
        elif resolved_unit and resolved_unit in by_unit:
            existing = by_unit[resolved_unit]

        # bentrok unit: unit sudah dipakai pembeli aktif LAIN
        if resolved_unit and resolved_unit in by_unit and (existing is None or by_unit[resolved_unit].id != existing.id):
            errors.append(f"Unit '{unum}' sudah dipakai pembeli lain ({by_unit[resolved_unit].full_name})")

        if errors:
            parsed.append((rnum, "error", None, None, None, ImportRow(row=rnum, action="error", label=label, errors=errors)))
            continue

        payload = {"full_name": name, "nik": nik,
                   "phone": (str(d["phone"]).strip() if d["phone"] not in (None, "") else None),
                   "email": (str(d["email"]).strip() if d["email"] not in (None, "") else None),
                   "address": (str(d["address"]).strip() if d["address"] not in (None, "") else None),
                   "project_id": pid, "unit_id": resolved_unit, "unit_number": unum,
                   "payment_type": pt, "contract_value": cval, "contract_date": cdate,
                   "ppjb_number": (str(d["ppjb_number"]).strip() if d["ppjb_number"] not in (None, "") else None),
                   "ajb_number": (str(d["ajb_number"]).strip() if d["ajb_number"] not in (None, "") else None),
                   "status": st}
        if existing:
            parsed.append((rnum, "update", payload, existing, resolved_unit,
                           ImportRow(row=rnum, action="update", label=label, note="Cocok " + ("NIK" if (nik and nik in by_nik) else "unit"))))
        else:
            parsed.append((rnum, "insert", payload, None, resolved_unit,
                           ImportRow(row=rnum, action="insert", label=label)))
    return parsed


@router.post("/clients/preview", response_model=ImportPreview)
async def preview_clients(
    file: UploadFile = File(...),
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    contents = await file.read()
    rows = _read_client_rows(contents)
    maps = await _load_client_maps(db, ctx.tenant_id)
    parsed = _classify_clients(rows, *maps)
    ins = sum(1 for p in parsed if p[1] == "insert")
    upd = sum(1 for p in parsed if p[1] == "update")
    err = sum(1 for p in parsed if p[1] == "error")
    return ImportPreview(sheet="PEMBELI", total=len(parsed), to_insert=ins, to_update=upd,
                         error_count=err, rows=[p[5] for p in parsed])


@router.post("/clients/commit", response_model=ImportCommitResult)
async def commit_clients(
    file: UploadFile = File(...),
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    contents = await file.read()
    rows = _read_client_rows(contents)
    maps = await _load_client_maps(db, ctx.tenant_id)
    parsed = _classify_clients(rows, *maps)
    inserted = updated = 0
    new_clients = []
    upd_fields = ["full_name", "nik", "phone", "email", "address", "project_id", "unit_id",
                  "unit_number", "payment_type", "contract_value", "contract_date",
                  "ppjb_number", "ajb_number", "status"]
    for rnum, kind, payload, existing, unit_id, rowres in parsed:
        if kind == "insert":
            c = Client(tenant_id=ctx.tenant_id, marketing_user_id=ctx.user_id,
                       status=payload["status"] or ClientStatus.ACTIVE,
                       **{k: payload[k] for k in upd_fields if k != "status"})
            db.add(c); await db.flush()
            if c.unit_id:
                await set_unit_status(db, ctx.tenant_id, c.unit_id, unit_status_for_client(c))
            new_clients.append(c); inserted += 1
        elif kind == "update":
            for f in upd_fields:
                nv = payload[f]
                if nv is not None:
                    setattr(existing, f, nv)
            await db.flush()
            if existing.unit_id:
                await set_unit_status(db, ctx.tenant_id, existing.unit_id, unit_status_for_client(existing))
            updated += 1
    batch_id = ""
    if inserted or updated:
        batch = await _save_batch(db, ctx, "clients", [(c.id, None) for c in new_clients], updated)
        batch_id = str(batch.id)
        await db.commit()
    err = sum(1 for p in parsed if p[1] == "error")
    return ImportCommitResult(batch_id=batch_id, inserted=inserted, updated=updated,
                              error_count=err, rows=[p[5] for p in parsed])


@router.get("/clients/template")
async def download_clients_template(
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    """Template PEMBELI & KONTRAK — terisi pembeli tenant saat ini (bila ada)."""
    projs = (await db.execute(
        select(Project.id, Project.name).where(Project.tenant_id == ctx.tenant_id).order_by(Project.name))).all()
    pmap = {pid: name for pid, name in projs}
    pnames = [name for _, name in projs]
    # nomor unit dari unit_id
    clients = (await db.execute(
        select(Client).where(Client.tenant_id == ctx.tenant_id, Client.is_deleted == False))).scalars().all()  # noqa: E712
    unit_ids = [c.unit_id for c in clients if c.unit_id]
    unum_map = {}
    if unit_ids:
        for uid, unum in (await db.execute(select(Unit.id, Unit.unit_number).where(Unit.id.in_(unit_ids)))).all():
            unum_map[uid] = unum
    rows = []
    for c in clients:
        rows.append({
            "nama": c.full_name or "", "nik": c.nik or "", "hp": c.phone or "", "email": c.email or "",
            "alamat": c.address or "", "proyek": pmap.get(c.project_id, ""),
            "unit": unum_map.get(c.unit_id) or (c.unit_number or ""),
            "bayar": PT_TO_LABEL.get(c.payment_type, "") if c.payment_type else "",
            "kontrak": float(c.contract_value) if c.contract_value is not None else None,
            "tgl": c.contract_date.strftime("%d/%m/%Y") if c.contract_date else "",
            "ppjb": c.ppjb_number or "", "ajb": c.ajb_number or "",
            "status": CST_TO_LABEL.get(c.status, "Aktif"),
        })
    data = _build_client_template(pnames, rows)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Template_Import_Pembeli.xlsx"'})


def _build_client_template(projects, clients) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Petunjuk"; ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 26; ws.column_dimensions["C"].width = 92
    t = ws.cell(row=1, column=2, value="TEMPLATE IMPOR — PEMBELI & KONTRAK")
    t.font = Font(name=_FONT, bold=True, color="FFFFFF", size=13); ws.merge_cells("B1:C1")
    ws["B1"].fill = _hdr_fill; ws["C1"].fill = _hdr_fill; ws.row_dimensions[1].height = 26
    tips = [
        ("Cara kerja", "Lengkapi sheet PEMBELI. Upload → pratinjau (baru/perbarui/error) → Terapkan."),
        ("Pencocokan (anti-dobel)", "Sistem cocokkan lewat NIK; kalau NIK kosong, lewat Proyek+Nomor Unit. Cocok = diperbarui; tidak = ditambah."),
        ("Nomor Unit", "Harus sudah ada (impor Unit dulu). Menautkan pembeli ke unit → status unit otomatis (Dipesan/Terjual)."),
        ("Tipe Pembayaran", "Cash · KPR"),
        ("Status", "Aktif · Selesai · Batal"),
        ("Format", "Tanggal dd/mm/yyyy · angka polos tanpa 'Rp'/titik."),
        ("Nama proyek (persis)", " · ".join(projects)),
    ]
    r = 3
    for a, b in tips:
        ca = ws.cell(row=r, column=2, value=a); ca.font = Font(name=_FONT, bold=True, size=10)
        ca.alignment = Alignment(vertical="top", wrap_text=True)
        cb = ws.cell(row=r, column=3, value=b); cb.font = Font(name=_FONT, size=10)
        cb.alignment = Alignment(vertical="top", wrap_text=True); ws.row_dimensions[r].height = 28
        r += 1
    wp = wb.create_sheet("PEMBELI"); wp.sheet_view.showGridLines = False
    cols = [("Nama Pembeli", 22, False, True), ("NIK (KTP)", 20, False, False), ("No. HP", 15, False, False),
            ("Email", 20, False, False), ("Alamat", 24, False, False), ("Proyek", 18, True, False),
            ("Nomor Unit", 12, True, False), ("Tipe Pembayaran", 15, False, False),
            ("Nilai Kontrak (Rp)", 16, False, False), ("Tanggal Kontrak", 15, False, False),
            ("No. PPJB", 14, False, False), ("No. AJB", 14, False, False), ("Status", 12, False, False)]
    _headers(wp, cols)
    row = 2
    for c in clients:
        _cell(wp, row, 1, c["nama"])
        _cell(wp, row, 2, c["nik"])
        _cell(wp, row, 3, c["hp"])
        _cell(wp, row, 4, c["email"])
        _cell(wp, row, 5, c["alamat"])
        _cell(wp, row, 6, c["proyek"], fill=_key_cell)
        _cell(wp, row, 7, c["unit"], fill=_key_cell, align="center")
        _cell(wp, row, 8, c["bayar"], align="center")
        _cell(wp, row, 9, c["kontrak"], align="right", numfmt="#,##0")
        _cell(wp, row, 10, c["tgl"], align="center")
        _cell(wp, row, 11, c["ppjb"])
        _cell(wp, row, 12, c["ajb"])
        _cell(wp, row, 13, c["status"], align="center")
        row += 1
    _dv(wp, ["Cash", "KPR"], 8, 2, row + 500)
    _dv(wp, ["Aktif", "Selesai", "Batal"], 13, 2, row + 500)
    _dv(wp, projects, 6, 2, row + 500)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio.getvalue()


# ═══════════════════════ DOKUMEN LEGALITAS UNIT (manifest + ZIP) ═══════════════════════
# doc_type kanonik (samakan dgn preset FE LegalDocuments) supaya dokumen impor tampil di UI.
DOC_CANON = {
    "shm": "Sertifikat SHM", "sertifikat shm": "Sertifikat SHM",
    "hgb": "Sertifikat HGB", "sertifikat hgb": "Sertifikat HGB",
    "slf": "SLF",
    "pbg": "IMB / PBG", "imb": "IMB / PBG", "imb/pbg": "IMB / PBG", "imb / pbg": "IMB / PBG",
    "pbb": "PBB",
}
LEGAL_DOC_TYPES = ["Sertifikat SHM", "Sertifikat HGB", "SLF", "IMB / PBG", "PBB"]
LABEL_TO_DOCST = {"belum": DocStatus.BELUM, "proses": DocStatus.PROSES, "terbit": DocStatus.TERBIT}
DOCST_TO_LABEL = {DocStatus.BELUM: "Belum", DocStatus.PROSES: "Proses", DocStatus.TERBIT: "Terbit"}


def _doc_field_of(h: str) -> Optional[str]:
    n = _norm_header(h)
    if n.startswith("jenis"): return "doc_type"
    if n.startswith("proyek"): return "project"
    if n.startswith("blok"): return "block"
    if n.startswith("nomor unit") or n.startswith("no unit") or n == "unit": return "unit_number"
    if "nomor dokumen" in n or n.startswith("no. dok") or n.startswith("no dok") or n.startswith("nop"): return "doc_number"
    if n.startswith("luas tanah") or n == "lt" or n.startswith("lt "): return "land_area"
    if n.startswith("alamat") or "objek" in n: return "address"
    if "masa berlaku" in n or n.startswith("berlaku") or "expiry" in n or "kedaluwarsa" in n: return "expiry_date"
    if n.startswith("tanggal"): return "doc_date"
    if n.startswith("status"): return "status"
    if "nama file" in n or n == "file" or n.startswith("file"): return "file_name"
    return None


def _read_doc_rows(contents: bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="File bukan Excel (.xlsx) yang valid.")
    ws = None
    for name in wb.sheetnames:
        if name.strip().lower().startswith("dokumen"):
            ws = wb[name]; break
    if ws is None:
        raise HTTPException(status_code=400, detail="Sheet 'DOKUMEN' tidak ditemukan di file.")
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return []
    col_map = {}
    for i, h in enumerate(header):
        f = _doc_field_of(h)
        if f and f not in col_map:
            col_map[f] = i
    if "doc_type" not in col_map or "unit_number" not in col_map:
        raise HTTPException(status_code=400, detail="Header 'Jenis Dokumen' & 'Nomor Unit' wajib ada di sheet DOKUMEN.")
    fields = ["doc_type", "project", "block", "unit_number", "doc_number", "land_area",
              "address", "expiry_date", "doc_date", "status", "file_name"]
    out = []
    rnum = 1
    for raw in it:
        rnum += 1
        d = {f: (raw[col_map[f]] if (f in col_map and col_map[f] < len(raw)) else None) for f in fields}
        if all((v is None or str(v).strip() == "") for v in d.values()):
            continue
        if str(d.get("doc_type") or "").strip().upper().startswith("CONTOH"):
            continue
        out.append((rnum, d))
    return out


async def _load_doc_maps(db: AsyncSession, tenant_id):
    projs = (await db.execute(select(Project.id, Project.name, Project.id).where(Project.tenant_id == tenant_id))).all()
    proj_by_name = {name.strip().lower(): pid for pid, name, _ in projs}
    units = (await db.execute(select(Unit.id, Unit.project_id, Unit.block, Unit.unit_number).where(Unit.tenant_id == tenant_id))).all()
    unit_by_key = {}   # (pid, block, unum) -> unit_id
    unit_by_pu = {}    # (pid, unum) -> [unit_id]
    for uid, pid, block, unum in units:
        bl = (block or "").strip().lower(); un = (unum or "").strip().lower()
        unit_by_key[(pid, bl, un)] = uid
        unit_by_pu.setdefault((pid, un), []).append(uid)
    docs = (await db.execute(
        select(Document).where(Document.tenant_id == tenant_id, Document.is_deleted == False,  # noqa: E712
                               Document.unit_id.isnot(None)))).scalars().all()
    existing = {}
    for dc in docs:
        existing[(dc.unit_id, dc.doc_type)] = dc
    return proj_by_name, unit_by_key, unit_by_pu, existing


def _classify_docs(rows, proj_by_name, unit_by_key, unit_by_pu, existing):
    """→ list (row_num, kind, payload, existing_doc, ImportRow). Manifest-only —
    keberadaan file di ZIP dicek saat commit (agar pratinjau tak perlu unggah ZIP)."""
    parsed = []
    seen = {}  # (unit_id, doc_type) dalam batch → agar dobel di file jadi update
    for rnum, d in rows:
        errors = []
        raw_type = str(d["doc_type"] or "").strip()
        canon = DOC_CANON.get(raw_type.lower())
        if not raw_type:
            errors.append("Jenis Dokumen kosong")
        elif canon is None:
            errors.append(f"Jenis Dokumen tak dikenal: '{raw_type}' (SHM/HGB/SLF/PBG/PBB)")
        pname = str(d["project"] or "").strip()
        unum = str(d["unit_number"] or "").strip()
        block = (str(d["block"]).strip() if d["block"] not in (None, "") else None)
        pid = proj_by_name.get(pname.lower()) if pname else None
        label = f"{canon or raw_type or '?'} — {pname or '?'}/{unum or '?'}"
        if not pname:
            errors.append("Proyek kosong")
        elif pid is None:
            errors.append(f"Proyek '{pname}' tidak ada")
        if not unum:
            errors.append("Nomor Unit kosong")

        unit_id = None
        if pid is not None and unum:
            if block is not None:
                unit_id = unit_by_key.get((pid, block.lower(), unum.lower()))
                if unit_id is None:
                    errors.append(f"Unit {block}/{unum} tak ada di {pname}")
            else:
                cand = unit_by_pu.get((pid, unum.lower()), [])
                if len(cand) == 0:
                    errors.append(f"Unit '{unum}' tak ada di {pname}")
                elif len(cand) > 1:
                    errors.append(f"Unit '{unum}' ambigu (>1 blok) — isi kolom Blok")
                else:
                    unit_id = cand[0]

        st = None
        if d["status"] not in (None, ""):
            st = LABEL_TO_DOCST.get(str(d["status"]).strip().lower())
            if st is None:
                errors.append(f"Status tak dikenal: '{d['status']}'")
        try:
            ddate = _to_date(d["doc_date"])
        except ValueError:
            errors.append(f"Tanggal tak valid: '{d['doc_date']}'"); ddate = None
        try:
            la = _to_decimal(d["land_area"])
        except (InvalidOperation, ValueError):
            errors.append(f"Luas Tanah bukan angka: '{d['land_area']}'"); la = None
        try:
            edate = _to_date(d["expiry_date"])
        except ValueError:
            errors.append(f"Masa Berlaku tak valid: '{d['expiry_date']}'"); edate = None
        addr = (str(d["address"]).strip() if d["address"] not in (None, "") else None)

        fname = (str(d["file_name"]).strip() if d["file_name"] not in (None, "") else None)

        if errors:
            parsed.append((rnum, "error", None, None, ImportRow(row=rnum, action="error", label=label, errors=errors)))
            continue

        payload = {"unit_id": unit_id, "unit_number": unum, "doc_type": canon,
                   "name": (str(d["doc_number"]).strip() if d["doc_number"] not in (None, "") else None),
                   "doc_date": ddate, "status": st, "land_area": la, "address": addr, "expiry_date": edate, "file_name": fname}
        key = (unit_id, canon)
        ex = existing.get(key) or seen.get(key)
        note = (f"Lampirkan: {fname}" if fname else "Metadata (file dicocokkan dari ZIP via nomor unit)")
        if ex is not None:
            parsed.append((rnum, "update", payload, ex, ImportRow(row=rnum, action="update", label=label, note=note)))
        else:
            seen[key] = True
            parsed.append((rnum, "insert", payload, None, ImportRow(row=rnum, action="insert", label=label, note=note)))
    return parsed


def _build_automatch_index(zip_names):
    """Dari basename→[fullname] bikin indeks utk auto-match by nomor unit.
    stem_index: stem-nama-file (persis) → set(fullname); token_index: token dlm nama → set(fullname).
    Angka dinormalkan (001 ↔ 1) supaya '001.pdf' cocok ke unit '1' dan sebaliknya."""
    stem_index, token_index = {}, {}
    for base, fulls in zip_names.items():
        stem = os.path.splitext(base)[0]
        stem_keys = {stem}
        if stem.isdigit():
            stem_keys.add(str(int(stem)))
        for k in stem_keys:
            stem_index.setdefault(k, set()).update(fulls)
        for tok in re.split(r"[^a-z0-9]+", stem):
            if not tok:
                continue
            token_index.setdefault(tok, set()).update(fulls)
            if tok.isdigit():
                token_index.setdefault(str(int(tok)), set()).update(fulls)
    return stem_index, token_index


def _match_files_for_unit(unum: str, stem_index, token_index):
    """Cari fullname file utk nomor unit ini. Prioritas: stem = nomor unit; fallback: token."""
    if not unum:
        return []
    cands = {unum.lower()}
    if unum.isdigit():
        cands.add(str(int(unum)))
    hits = set()
    for c in cands:
        hits |= stem_index.get(c, set())
    if hits:
        return list(hits)
    for c in cands:
        hits |= token_index.get(c, set())
    return list(hits)


def _read_zip_names(archive_bytes: Optional[bytes]):
    if not archive_bytes:
        return {}
    try:
        zf = zipfile.ZipFile(io.BytesIO(archive_bytes))
    except Exception:
        raise HTTPException(status_code=400, detail="Arsip ZIP tidak valid.")
    names = {}
    for n in zf.namelist():
        if n.endswith("/"):
            continue
        base = os.path.basename(n).lower()
        names.setdefault(base, []).append(n)
    return names


@router.post("/documents/preview", response_model=ImportPreview)
async def preview_documents(
    manifest: UploadFile = File(...),
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    """Pratinjau MANIFEST saja (tanpa ZIP) — ringan & cepat. Keberadaan file dicek saat Terapkan."""
    contents = await manifest.read()
    rows = _read_doc_rows(contents)
    maps = await _load_doc_maps(db, ctx.tenant_id)
    parsed = _classify_docs(rows, *maps)
    ins = sum(1 for p in parsed if p[1] == "insert")
    upd = sum(1 for p in parsed if p[1] == "update")
    err = sum(1 for p in parsed if p[1] == "error")
    return ImportPreview(sheet="DOKUMEN", total=len(parsed), to_insert=ins, to_update=upd,
                         error_count=err, rows=[p[4] for p in parsed])


@router.post("/documents/commit", response_model=ImportCommitResult)
async def commit_documents(
    manifest: UploadFile = File(...),
    archive: Optional[UploadFile] = File(None),
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    contents = await manifest.read()
    rows = _read_doc_rows(contents)
    arc = await archive.read() if archive is not None else None
    zf = zipfile.ZipFile(io.BytesIO(arc)) if arc else None
    zip_names = _read_zip_names(arc)
    stem_index, token_index = _build_automatch_index(zip_names) if zf else ({}, {})
    proj_by_name, unit_by_key, unit_by_pu, existing = await _load_doc_maps(db, ctx.tenant_id)
    # project_id tiap unit (utk isi Document.project_id)
    unit_pid = {}
    for (pid, bl, un), uid in unit_by_key.items():
        unit_pid[uid] = pid
    parsed = _classify_docs(rows, proj_by_name, unit_by_key, unit_by_pu, existing)
    inserted = updated = 0
    attached = missing = 0
    created = {}  # (unit_id, doc_type) → Document baru dlm batch ini
    for rnum, kind, payload, ex, rowres in parsed:
        if kind == "error":
            continue
        key = (payload["unit_id"], payload["doc_type"])
        doc = ex or created.get(key)
        if doc is None:
            doc = Document(tenant_id=ctx.tenant_id, unit_id=payload["unit_id"],
                           project_id=unit_pid.get(payload["unit_id"]), doc_type=payload["doc_type"],
                           status=payload["status"] or DocStatus.TERBIT)
            db.add(doc); await db.flush(); created[key] = doc; inserted += 1
        else:
            updated += 1
        if payload["name"] is not None:
            doc.name = payload["name"]
        if payload["doc_date"] is not None:
            doc.doc_date = payload["doc_date"]
        if payload["status"] is not None:
            doc.status = payload["status"]
        if payload["address"] is not None:
            doc.address = payload["address"]
        if payload["expiry_date"] is not None:
            doc.expiry_date = payload["expiry_date"]
        if payload["land_area"] is not None:
            doc.land_area = payload["land_area"]
            # LT dari dokumen legalitas = sumber valid → sinkron ke Unit.land_area (spt modul Dokumen)
            u = (await db.execute(select(Unit).where(
                Unit.id == doc.unit_id, Unit.tenant_id == ctx.tenant_id))).scalar_one_or_none()
            if u is not None:
                u.land_area = payload["land_area"]
        # ── lampirkan file dari ZIP ──
        # Nama File diisi → pakai itu. Kosong → AUTO-MATCH via nomor unit (nama file = nomor unit).
        # Metadata tetap tersimpan walau file tak ada; masalah file = CATATAN per-baris (tak menggagalkan).
        async def _attach(full, suffix):
            data = zf.read(full)
            if len(data) > _MAX_FILE_BYTES:
                rowres.note = (rowres.note or "") + " — file >10MB dilewati"
                return False
            fn = os.path.basename(full)
            key_obj = storage.build_key(ctx.tenant_id, "documents", doc.id, fn)
            await storage.put(key_obj, data, mimetypes.guess_type(fn)[0])
            doc.file_key = key_obj
            doc.file_data = None
            doc.file_name = fn
            doc.file_type = mimetypes.guess_type(fn)[0] or "application/octet-stream"
            doc.file_size = len(data)
            rowres.note = (rowres.note or "") + suffix
            return True

        fname = payload["file_name"]
        if fname:
            if not zf:
                rowres.note = (rowres.note or "") + " — ZIP tak diunggah, file dilewati"
            else:
                hit = zip_names.get(os.path.basename(fname).lower())
                if hit is None:
                    rowres.note = (rowres.note or "") + f" — file '{fname}' tak ada di ZIP"; missing += 1
                elif len(hit) > 1:
                    rowres.note = (rowres.note or "") + f" — nama file '{fname}' dobel di ZIP"; missing += 1
                elif await _attach(hit[0], f" — file {os.path.basename(fname)} terlampir"):
                    attached += 1
                else:
                    missing += 1
        elif zf is not None:
            m = _match_files_for_unit(payload["unit_number"], stem_index, token_index)
            if len(m) == 1:
                if await _attach(m[0], f" — auto: {os.path.basename(m[0])}"):
                    attached += 1
                else:
                    missing += 1
            elif len(m) == 0:
                rowres.note = (rowres.note or "") + f" — tak ada file utk unit {payload['unit_number']}"; missing += 1
            else:
                rowres.note = (rowres.note or "") + f" — {len(m)} file cocok utk unit {payload['unit_number']} (isi Nama File utk pastikan)"; missing += 1
    batch_id = ""
    if inserted or updated:
        batch = await _save_batch(db, ctx, "documents",
                                  [(doc.id, doc.file_key) for doc in created.values()], updated)
        batch_id = str(batch.id)
        await db.commit()
    err = sum(1 for p in parsed if p[1] == "error")
    return ImportCommitResult(batch_id=batch_id, inserted=inserted, updated=updated,
                              error_count=err, rows=[p[4] for p in parsed])


@router.get("/documents/template")
async def download_documents_template(
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    """Template manifest DOKUMEN legalitas unit — terisi dokumen legalitas yang sudah ada (bila ada)."""
    projs = (await db.execute(
        select(Project.id, Project.name).where(Project.tenant_id == ctx.tenant_id).order_by(Project.name))).all()
    pmap = {pid: name for pid, name in projs}
    pnames = [name for _, name in projs]
    docs = (await db.execute(
        select(Document).where(Document.tenant_id == ctx.tenant_id, Document.is_deleted == False,  # noqa: E712
                               Document.unit_id.isnot(None), Document.doc_type.in_(LEGAL_DOC_TYPES)))).scalars().all()
    uids = [dc.unit_id for dc in docs if dc.unit_id]
    umap = {}
    if uids:
        for uid, block, unum, pid in (await db.execute(
                select(Unit.id, Unit.block, Unit.unit_number, Unit.project_id).where(Unit.id.in_(uids)))).all():
            umap[uid] = (block or "", unum or "", pmap.get(pid, ""))
    rows = []
    for dc in docs:
        blk, unum, pnm = umap.get(dc.unit_id, ("", "", ""))
        rows.append({"jenis": dc.doc_type, "proyek": pnm, "blok": blk, "unit": unum,
                     "nomor": dc.name or "", "lt": float(dc.land_area) if dc.land_area is not None else None,
                     "alamat": dc.address or "",
                     "berlaku": dc.expiry_date.strftime("%d/%m/%Y") if dc.expiry_date else "",
                     "tgl": dc.doc_date.strftime("%d/%m/%Y") if dc.doc_date else "",
                     "status": DOCST_TO_LABEL.get(dc.status, "Terbit"),
                     "file": dc.file_name or ""})
    data = _build_doc_template(pnames, rows)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Template_Import_Dokumen.xlsx"'})


def _build_doc_template(projects, docs) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Petunjuk"; ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 26; ws.column_dimensions["C"].width = 94
    t = ws.cell(row=1, column=2, value="TEMPLATE IMPOR — DOKUMEN LEGALITAS UNIT")
    t.font = Font(name=_FONT, bold=True, color="FFFFFF", size=13); ws.merge_cells("B1:C1")
    ws["B1"].fill = _hdr_fill; ws["C1"].fill = _hdr_fill; ws.row_dimensions[1].height = 26
    tips = [
        ("Cara kerja", "Isi manifest DOKUMEN (1 baris = 1 dokumen unit). Untuk melampirkan scan: tulis Nama File, kumpulkan semua file jadi 1 ZIP, upload manifest + ZIP. Nomor/tanggal saja tanpa file juga boleh (Nama File dikosongkan)."),
        ("Jenis Dokumen", "SHM · HGB · SLF · PBG (IMB) · PBB"),
        ("Kunci pencocokan", "Proyek + (Blok) + Nomor Unit + Jenis → dokumen unit diperbarui bila sudah ada, atau dibuat baru. Isi Blok bila nomor unit dobel antar-blok."),
        ("Lampirkan scan (2 cara)", "OTOMATIS: beri nama file scan = NOMOR UNIT (mis. 001.pdf, 036.pdf), zip semua, kosongkan kolom 'Nama File' → sistem mencocokkan sendiri. MANUAL: isi kolom 'Nama File' dengan nama file persis di ZIP. Maks 10MB/file."),
        ("Kolom per jenis (isi yang relevan, sisanya kosong)",
         "Luas Tanah → SHM/HGB (ikut memperbarui LT unit) · Alamat Objek (NOP) → PBB · Masa Berlaku → SLF/PBG."),
        ("Status", "Belum · Proses · Terbit (default Terbit)"),
        ("Format", "Tanggal & Masa Berlaku dd/mm/yyyy · angka LT polos (mis. 72)."),
        ("Nama proyek (persis)", " · ".join(projects)),
    ]
    r = 3
    for a, b in tips:
        ca = ws.cell(row=r, column=2, value=a); ca.font = Font(name=_FONT, bold=True, size=10)
        ca.alignment = Alignment(vertical="top", wrap_text=True)
        cb = ws.cell(row=r, column=3, value=b); cb.font = Font(name=_FONT, size=10)
        cb.alignment = Alignment(vertical="top", wrap_text=True); ws.row_dimensions[r].height = 30
        r += 1
    wd = wb.create_sheet("DOKUMEN"); wd.sheet_view.showGridLines = False
    cols = [("Jenis Dokumen", 16, False, True), ("Proyek", 18, True, True), ("Blok", 8, True, False),
            ("Nomor Unit", 12, True, True), ("Nomor Dokumen", 20, False, False),
            ("Luas Tanah (m²)", 13, False, False), ("Alamat Objek (PBB)", 24, False, False),
            ("Masa Berlaku", 14, False, False), ("Tanggal", 14, False, False),
            ("Status", 12, False, False), ("Nama File", 26, False, False)]
    _headers(wd, cols)
    row = 2
    if not docs:
        # baris contoh (dihapus otomatis saat impor karena diawali CONTOH)
        ex = ["SHM", projects[0] if projects else "", "", "036", "12.34.56", 72, "", "", "20/05/2025", "Terbit", "Daiyan_036_SHM.pdf"]
        for i, v in enumerate(ex, start=1):
            _cell(wd, row, i, v, fill=_ex_fill, italic=True, align="center" if i in (1, 3, 4, 6, 8, 9, 10) else "left")
        row += 1
    for dc in docs:
        _cell(wd, row, 1, dc["jenis"], align="center")
        _cell(wd, row, 2, dc["proyek"], fill=_key_cell)
        _cell(wd, row, 3, dc["blok"], fill=_key_cell, align="center")
        _cell(wd, row, 4, dc["unit"], fill=_key_cell, align="center")
        _cell(wd, row, 5, dc["nomor"])
        _cell(wd, row, 6, dc["lt"], align="right", numfmt="#,##0.##")
        _cell(wd, row, 7, dc["alamat"])
        _cell(wd, row, 8, dc["berlaku"], align="center")
        _cell(wd, row, 9, dc["tgl"], align="center")
        _cell(wd, row, 10, dc["status"], align="center")
        _cell(wd, row, 11, dc["file"])
        row += 1
    _dv(wd, ["SHM", "HGB", "SLF", "PBG", "PBB"], 1, 2, row + 500)
    _dv(wd, ["Belum", "Proses", "Terbit"], 10, 2, row + 500)
    _dv(wd, projects, 2, 2, row + 500)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio.getvalue()


# ═══════════════════════ RIWAYAT & UNDO BATCH IMPOR ═══════════════════════
class ImportBatchRow(BaseModel):
    id: uuid.UUID
    entity: str
    inserted: int
    updated: int
    created_at: datetime
    undone_at: Optional[datetime] = None
    can_undo: bool


class UndoResult(BaseModel):
    deleted: int
    files_removed: int
    entity: str


@router.get("/batches", response_model=list[ImportBatchRow])
async def list_batches(
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    rows = (await db.execute(
        select(ImportBatch).where(ImportBatch.tenant_id == ctx.tenant_id)
        .order_by(ImportBatch.created_at.desc()).limit(50))).scalars().all()
    return [ImportBatchRow(
        id=b.id, entity=b.entity, inserted=b.inserted, updated=b.updated,
        created_at=b.created_at, undone_at=b.undone_at,
        can_undo=(b.inserted > 0 and b.undone_at is None)) for b in rows]


@router.post("/batches/{batch_id}/undo", response_model=UndoResult)
async def undo_batch(
    batch_id: uuid.UUID,
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    """Batalkan satu batch impor (Level 1): HAPUS record yang DITAMBAH batch itu + file MinIO-nya.
    Yang di-UPDATE tidak dikembalikan (butuh Level 2). Aman utuh (atomik) — gagal = tak ada yang berubah."""
    from datetime import timezone
    from sqlalchemy.exc import IntegrityError

    b = (await db.execute(select(ImportBatch).where(
        ImportBatch.id == batch_id, ImportBatch.tenant_id == ctx.tenant_id))).scalar_one_or_none()
    if b is None:
        raise HTTPException(status_code=404, detail="Batch tidak ditemukan")
    if b.undone_at is not None:
        raise HTTPException(status_code=400, detail="Batch ini sudah dibatalkan.")

    items = (await db.execute(select(ImportBatchItem).where(ImportBatchItem.batch_id == b.id))).scalars().all()
    deleted = 0
    file_keys = []
    try:
        for it in items:
            if it.resource == "units":
                obj = (await db.execute(select(Unit).where(
                    Unit.id == it.resource_id, Unit.tenant_id == ctx.tenant_id))).scalar_one_or_none()
                if obj is not None:
                    await db.delete(obj); deleted += 1
            elif it.resource == "clients":
                obj = (await db.execute(select(Client).where(
                    Client.id == it.resource_id, Client.tenant_id == ctx.tenant_id))).scalar_one_or_none()
                if obj is not None:
                    uid = obj.unit_id
                    await db.delete(obj); await db.flush(); deleted += 1
                    # kembalikan unit ke Tersedia bila tak ada pembeli aktif lain di unit itu
                    if uid:
                        other = (await db.execute(select(Client.id).where(
                            Client.tenant_id == ctx.tenant_id, Client.unit_id == uid,
                            Client.status != ClientStatus.INACTIVE, Client.is_deleted == False))).first()  # noqa: E712
                        if other is None:
                            await set_unit_status(db, ctx.tenant_id, uid, UnitStatus.AVAILABLE)
            elif it.resource == "documents":
                obj = (await db.execute(select(Document).where(
                    Document.id == it.resource_id, Document.tenant_id == ctx.tenant_id))).scalar_one_or_none()
                if obj is not None:
                    await db.delete(obj); deleted += 1
                if it.file_key:
                    file_keys.append(it.file_key)
            elif it.resource == "payments":
                obj = (await db.execute(select(Payment).where(
                    Payment.id == it.resource_id, Payment.tenant_id == ctx.tenant_id))).scalar_one_or_none()
                if obj is not None:
                    ce = (await db.execute(select(CashBookEntry).where(
                        CashBookEntry.tenant_id == ctx.tenant_id, CashBookEntry.source_type == "payment",
                        CashBookEntry.source_id == obj.id))).scalar_one_or_none()
                    if ce is not None:
                        await db.delete(ce)
                    await db.delete(obj); deleted += 1
        b.undone_at = datetime.now(timezone.utc)
        b.undone_by_id = ctx.user_id
        await record_audit(db, ctx.tenant_id, ctx.user_id, "UNDO_IMPORT", b.entity, b.id,
                           new_data={"deleted": deleted})
        await db.flush()
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(status_code=400,
                            detail="Tak bisa membatalkan: sebagian data sudah dipakai/direferensikan data lain.")

    # hapus objek MinIO setelah commit DB berhasil (best-effort)
    removed = 0
    for k in file_keys:
        try:
            await storage.delete(k); removed += 1
        except Exception:
            pass
    return UndoResult(deleted=deleted, files_removed=removed, entity=b.entity)


# ═══════════════════════ PEMBAYARAN (dari Pembeli) ═══════════════════════
# Fase 1: source=PEMBELI, status langsung APPROVED (data historis). Anti-dobel hibrida:
# No. Referensi bila diisi, kalau kosong dedup (pembeli+tanggal+jumlah).
LABEL_TO_PURPOSE = {
    "dp": PaymentPurpose.DP, "uang muka": PaymentPurpose.DP,
    "booking fee": PaymentPurpose.BOOKING_FEE, "booking": PaymentPurpose.BOOKING_FEE,
    "cicilan": PaymentPurpose.CICILAN_TERMIN, "termin": PaymentPurpose.CICILAN_TERMIN,
    "angsuran": PaymentPurpose.CICILAN_TERMIN, "cicilan termin": PaymentPurpose.CICILAN_TERMIN,
    "pelunasan": PaymentPurpose.PELUNASAN_TERMIN, "lunas": PaymentPurpose.LUNAS_UNIT,
    "lunas unit": PaymentPurpose.LUNAS_UNIT, "cash keras": PaymentPurpose.LUNAS_UNIT,
}
PURPOSE_TO_LABEL = {
    PaymentPurpose.DP: "DP", PaymentPurpose.BOOKING_FEE: "Booking Fee",
    PaymentPurpose.CICILAN_TERMIN: "Cicilan", PaymentPurpose.PELUNASAN_TERMIN: "Pelunasan",
    PaymentPurpose.LUNAS_UNIT: "Lunas Unit",
}
LABEL_TO_METHOD = {"transfer": PaymentMethod.TRANSFER, "tunai": PaymentMethod.TUNAI,
                   "cash": PaymentMethod.TUNAI, "lainnya": PaymentMethod.LAINNYA}


def _pay_field_of(h: str) -> Optional[str]:
    n = _norm_header(h)
    if n.startswith("nik"): return "nik"
    if n.startswith("nama"): return "name"
    if n.startswith("proyek"): return "project"
    if n.startswith("nomor unit") or n.startswith("no unit") or n == "unit": return "unit_number"
    if n.startswith("tanggal"): return "date"
    if n.startswith("jumlah") or n.startswith("nominal") or n.startswith("nilai"): return "amount"
    if n.startswith("jenis"): return "purpose"
    if n.startswith("metode") or n.startswith("cara bayar"): return "method"
    if "referensi" in n or "kwitansi" in n or n.startswith("no. ref") or n.startswith("no ref"): return "reference"
    if n.startswith("keterangan") or n.startswith("catatan"): return "notes"
    return None


def _read_pay_rows(contents: bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="File bukan Excel (.xlsx) yang valid.")
    ws = None
    for name in wb.sheetnames:
        if name.strip().lower().startswith("pembayaran"):
            ws = wb[name]; break
    if ws is None:
        raise HTTPException(status_code=400, detail="Sheet 'PEMBAYARAN' tidak ditemukan di file.")
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return []
    col_map = {}
    for i, h in enumerate(header):
        f = _pay_field_of(h)
        if f and f not in col_map:
            col_map[f] = i
    if "amount" not in col_map:
        raise HTTPException(status_code=400, detail="Header 'Jumlah' wajib ada di sheet PEMBAYARAN.")
    fields = ["nik", "name", "project", "unit_number", "date", "amount", "purpose", "method", "reference", "notes"]
    out = []
    rnum = 1
    for raw in it:
        rnum += 1
        d = {f: (raw[col_map[f]] if (f in col_map and col_map[f] < len(raw)) else None) for f in fields}
        if all((v is None or str(v).strip() == "") for v in d.values()):
            continue
        if str(d.get("name") or "").strip().upper().startswith("CONTOH"):
            continue
        out.append((rnum, d))
    return out


async def _load_pay_maps(db: AsyncSession, tenant_id):
    proj_by_name, unit_by_pu, by_nik, by_unit = await _load_client_maps(db, tenant_id)
    pays = (await db.execute(
        select(Payment.client_id, Payment.payment_date, Payment.amount, Payment.receipt_number)
        .where(Payment.tenant_id == tenant_id, Payment.is_deleted == False))).all()  # noqa: E712
    existing_receipts = set()
    existing_combos = set()
    for cid, pdate, amt, rcpt in pays:
        if rcpt:
            existing_receipts.add(str(rcpt).strip())
        existing_combos.add((str(cid), pdate.isoformat() if pdate else "", _money_key(amt)))
    return proj_by_name, unit_by_pu, by_nik, by_unit, existing_receipts, existing_combos


def _money_key(v) -> str:
    try:
        return f"{Decimal(v or 0):.2f}"
    except (InvalidOperation, TypeError, ValueError):
        return "0.00"


def _classify_payments(rows, proj_by_name, unit_by_pu, by_nik, by_unit, existing_receipts, existing_combos):
    parsed = []
    seen_receipt = set()
    seen_combo = set()
    for rnum, d in rows:
        errors = []
        nik = (str(d["nik"]).strip() if d["nik"] not in (None, "") else None)
        pname = str(d["project"] or "").strip()
        unum = (str(d["unit_number"]).strip() if d["unit_number"] not in (None, "") else None)
        # cari pembeli: NIK dulu, lalu Proyek+Unit (pembeli aktif di unit itu)
        client = None
        if nik and nik in by_nik:
            client = by_nik[nik]
        elif pname and unum:
            pid = proj_by_name.get(pname.lower())
            if pid is not None:
                cand = unit_by_pu.get((pid, unum.lower()), [])
                if len(cand) == 1:
                    client = by_unit.get(cand[0])
        cname = client.full_name if client else (str(d["name"] or "").strip() or nik or "?")
        if client is None:
            errors.append("Pembeli tak ditemukan (isi NIK, atau Proyek+Nomor Unit yang benar)")
        try:
            amount = _to_decimal(d["amount"])
        except (InvalidOperation, ValueError):
            errors.append(f"Jumlah bukan angka: '{d['amount']}'"); amount = None
        if amount is None or amount <= 0:
            errors.append("Jumlah kosong / tidak valid")
        try:
            pdate = _to_date(d["date"])
        except ValueError:
            errors.append(f"Tanggal tak valid: '{d['date']}'"); pdate = None
        purpose = PaymentPurpose.CICILAN_TERMIN
        if d["purpose"] not in (None, ""):
            purpose = LABEL_TO_PURPOSE.get(str(d["purpose"]).strip().lower())
            if purpose is None:
                errors.append(f"Jenis tak dikenal: '{d['purpose']}'")
        method = PaymentMethod.TRANSFER
        if d["method"] not in (None, ""):
            method = LABEL_TO_METHOD.get(str(d["method"]).strip().lower(), PaymentMethod.TRANSFER)
        ref = (str(d["reference"]).strip() if d["reference"] not in (None, "") else None)
        notes = (str(d["notes"]).strip() if d["notes"] not in (None, "") else None)
        label = f"{cname} — {_id_rp(amount)}"

        if errors:
            parsed.append((rnum, "error", None, None, ImportRow(row=rnum, action="error", label=label, errors=errors)))
            continue
        combo = (str(client.id), pdate.isoformat() if pdate else "", _money_key(amount))
        if ref and (ref in existing_receipts or ref in seen_receipt):
            parsed.append((rnum, "skip", None, None, ImportRow(row=rnum, action="skip", label=label, note=f"Dobel — No. Referensi '{ref}' sudah ada")))
            continue
        if not ref and (combo in existing_combos or combo in seen_combo):
            parsed.append((rnum, "skip", None, None, ImportRow(row=rnum, action="skip", label=label, note="Dobel — pembeli+tanggal+jumlah sama sudah ada")))
            continue
        if ref:
            seen_receipt.add(ref)
        seen_combo.add(combo)
        payload = {"client_id": client.id, "amount": amount, "payment_date": pdate,
                   "purpose": purpose, "method": method, "reference": ref, "notes": notes}
        parsed.append((rnum, "insert", payload, client, ImportRow(row=rnum, action="insert", label=label,
                       note=PURPOSE_TO_LABEL.get(purpose, purpose.value))))
    return parsed


def _id_rp(v):
    try:
        return "Rp " + f"{int(v):,}".replace(",", ".")
    except (TypeError, ValueError):
        return "Rp -"


@router.post("/payments/preview", response_model=ImportPreview)
async def preview_payments(
    file: UploadFile = File(...),
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    contents = await file.read()
    rows = _read_pay_rows(contents)
    maps = await _load_pay_maps(db, ctx.tenant_id)
    parsed = _classify_payments(rows, *maps)
    ins = sum(1 for p in parsed if p[1] == "insert")
    skp = sum(1 for p in parsed if p[1] == "skip")
    err = sum(1 for p in parsed if p[1] == "error")
    return ImportPreview(sheet="PEMBAYARAN", total=len(parsed), to_insert=ins, to_update=0,
                         error_count=err, to_skip=skp, rows=[p[4] for p in parsed])


@router.post("/payments/commit", response_model=ImportCommitResult)
async def commit_payments(
    file: UploadFile = File(...),
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    contents = await file.read()
    rows = _read_pay_rows(contents)
    maps = await _load_pay_maps(db, ctx.tenant_id)
    parsed = _classify_payments(rows, *maps)
    seq = int(await db.scalar(select(func.count()).select_from(Payment).where(Payment.tenant_id == ctx.tenant_id)) or 0)
    inserted = 0
    new_pays = []
    for rnum, kind, payload, client, rowres in parsed:
        if kind != "insert":
            continue
        if payload["reference"]:
            receipt = payload["reference"]
        else:
            seq += 1
            receipt = f"KW-{seq:06d}"
        pay = Payment(
            tenant_id=ctx.tenant_id, client_id=payload["client_id"], amount=payload["amount"],
            payment_date=payload["payment_date"], source=PaymentSource.PEMBELI,
            purpose=payload["purpose"], method=payload["method"], receipt_number=receipt,
            notes=payload["notes"], approval_status=PaymentApprovalStatus.APPROVED,
            approver_id=ctx.user_id, approved_at=datetime.utcnow())
        db.add(pay)
        await db.flush()
        await sync_payment_cashbook(db, ctx.tenant_id, pay)
        new_pays.append(pay)
        inserted += 1
    skipped = sum(1 for p in parsed if p[1] == "skip")
    err = sum(1 for p in parsed if p[1] == "error")
    batch_id = ""
    if inserted:
        batch = await _save_batch(db, ctx, "payments", [(p.id, None) for p in new_pays], 0)
        batch_id = str(batch.id)
        await db.commit()
    return ImportCommitResult(batch_id=batch_id, inserted=inserted, updated=0,
                              error_count=err, skipped=skipped, rows=[p[4] for p in parsed])


@router.get("/payments/template")
async def download_payments_template(
    ctx: AuthContext = Depends(get_current_context), db: AsyncSession = Depends(get_db),
):
    """Template PEMBAYARAN (dari pembeli). Kosong + contoh — pembayaran itu kejadian, tak di-ekspor."""
    projs = (await db.execute(
        select(Project.name).where(Project.tenant_id == ctx.tenant_id).order_by(Project.name))).all()
    pnames = [n for (n,) in projs]
    data = _build_pay_template(pnames)
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Template_Import_Pembayaran.xlsx"'})


def _build_pay_template(projects) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Petunjuk"; ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3; ws.column_dimensions["B"].width = 24; ws.column_dimensions["C"].width = 94
    t = ws.cell(row=1, column=2, value="TEMPLATE IMPOR — PEMBAYARAN (DARI PEMBELI)")
    t.font = Font(name=_FONT, bold=True, color="FFFFFF", size=13); ws.merge_cells("B1:C1")
    ws["B1"].fill = _hdr_fill; ws["C1"].fill = _hdr_fill; ws.row_dimensions[1].height = 26
    tips = [
        ("Cara kerja", "Isi sheet PEMBAYARAN (1 baris = 1 pembayaran). Upload → pratinjau → Terapkan. Pembayaran impor langsung DISETUJUI (data historis) & masuk Buku Kas."),
        ("Pencocokan pembeli", "Lewat NIK; kalau NIK kosong, lewat Proyek + Nomor Unit (pembeli aktif di unit itu)."),
        ("Anti-dobel", "Isi No. Referensi (mis. no. kwitansi) sebagai kunci unik. Kalau kosong, sistem menandai dobel bila pembeli+tanggal+jumlah sama sudah ada."),
        ("Jenis", "DP · Booking Fee · Cicilan · Pelunasan · Lunas Unit"),
        ("Metode", "Transfer · Tunai · Lainnya (default Transfer)"),
        ("Format", "Tanggal dd/mm/yyyy · Jumlah angka polos tanpa 'Rp'/titik."),
        ("Catatan", "Fase ini hanya pembayaran DARI PEMBELI. Pencairan Bank/KPR menyusul."),
        ("Nama proyek (persis)", " · ".join(projects)),
    ]
    r = 3
    for a, b in tips:
        ca = ws.cell(row=r, column=2, value=a); ca.font = Font(name=_FONT, bold=True, size=10)
        ca.alignment = Alignment(vertical="top", wrap_text=True)
        cb = ws.cell(row=r, column=3, value=b); cb.font = Font(name=_FONT, size=10)
        cb.alignment = Alignment(vertical="top", wrap_text=True); ws.row_dimensions[r].height = 28
        r += 1
    wp = wb.create_sheet("PEMBAYARAN"); wp.sheet_view.showGridLines = False
    cols = [("NIK (KTP)", 20, False, False), ("Nama Pembeli", 22, False, False), ("Proyek", 18, True, False),
            ("Nomor Unit", 12, True, False), ("Tanggal Bayar", 14, False, True), ("Jumlah (Rp)", 16, False, True),
            ("Jenis", 14, False, False), ("Metode", 12, False, False), ("No. Referensi", 16, False, False),
            ("Keterangan", 22, False, False)]
    _headers(wp, cols)
    ex = ["6371054910010004", "CONTOH — hapus baris ini", projects[0] if projects else "", "036",
          "20/05/2025", 50000000, "DP", "Transfer", "KW-001", "uang muka"]
    for i, v in enumerate(ex, start=1):
        _cell(wp, 2, i, v, fill=_ex_fill, italic=True, numfmt="#,##0" if i == 6 else None,
              align="right" if i == 6 else ("center" if i in (5, 7, 8) else "left"))
    _dv(wp, ["DP", "Booking Fee", "Cicilan", "Pelunasan", "Lunas Unit"], 7, 2, 800)
    _dv(wp, ["Transfer", "Tunai", "Lainnya"], 8, 2, 800)
    _dv(wp, projects, 3, 2, 800)
    for rr in range(3, 60):
        wp.cell(row=rr, column=3).fill = _key_cell
        wp.cell(row=rr, column=4).fill = _key_cell
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio.getvalue()
